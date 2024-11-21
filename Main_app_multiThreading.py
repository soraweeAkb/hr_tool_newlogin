from PyQt5.QtCore import QTimer, QThread, pyqtSignal
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QMessageBox, QTableWidgetItem, QDialog, QFileDialog
from PyQt5 import QtCore
from PyQt5.QtCore import QDate, QTime, Qt

import sys
APP = QApplication(sys.argv)

DESKTOP = QApplication.desktop()

WIDTH = DESKTOP.width()

if WIDTH <= 1366:
    from ui.login import Ui_loginWindow
    from ui.mainwindow import Ui_MainWindow
    from ui.password import Ui_PassWindow
    from ui.timecard import Ui_TimeCard
    from ui.askforleave import Ui_AskForLeave
    from ui.otapplication import Ui_OTApplication
    from ui.bookmeetingroom import Ui_BookMeetingRoom
    from ui.approve_panel import Ui_ApprovePanel
    from ui.applylateclockin import Ui_ApplyLateClockIn
    from ui.forgetrecord import Ui_ForgetRecord

    from ui.main_admin import Ui_AdminMain
    from ui.staff_manage import Ui_StaffManage
    from ui.team_structure import Ui_TeamStructure
    from ui.login_pass import Ui_LoginPass
    from ui.calendar import Ui_CalendarSetting

    W_LEAVE=831
    H_LEAVE=550
    W_OT=858
    H_OT=526

    W_APPROVE=1065
    H_APPROVE=672
else:
    from ui_highDPI.login import Ui_loginWindow
    from ui_highDPI.mainwindow import Ui_MainWindow
    from ui_highDPI.password import Ui_PassWindow
    from ui_highDPI.timecard import Ui_TimeCard
    from ui_highDPI.askforleave import Ui_AskForLeave
    from ui_highDPI.otapplication import Ui_OTApplication
    from ui_highDPI.bookmeetingroom import Ui_BookMeetingRoom
    from ui_highDPI.approve_panel import Ui_ApprovePanel
    from ui_highDPI.applylateclockin import Ui_ApplyLateClockIn
    from ui_highDPI.forgetrecord import Ui_ForgetRecord
    W_LEAVE = 1067
    H_LEAVE = 722
    W_OT = 1067
    H_OT = 701

    W_APPROVE = 1296
    H_APPROVE = 776

from modules.NTP_time import NTP_DateTime
from modules.Mail_Sender import MailSender

import datetime
import time
import csv
import os
import openpyxl as xl

import pymysql


class loginWindow(QMainWindow, Ui_loginWindow):
    def __init__(self):
        super(loginWindow, self).__init__()
        self.setupUi(self)

        self.label_7.setText('Developed in 2020  Ver.1.1')
        self.label_2.setText('HR Information System V1.1')

        self.id = ''
        self.pushButton.clicked.connect(self.login)
        self.pushButton_2.clicked.connect(self.quit)
        #self.pushButton_2.clicked.connect(self.temp)

        self.init_default()

    def init_default(self):
        info_line = []
        try:
            with open('info.csv', 'r') as file:
                reader = csv.reader(file)
                for line in reader:
                    info_line.append(line)
            file.close()
        except:
            pass

        if info_line!=[]:
            if info_line[0][3]=='1':
                self.lineEdit.setText(str(info_line[0][0]))
                self.lineEdit_2.setText(str(info_line[0][1]))
                self.comboBox.setCurrentText(str(info_line[0][2]))
                self.checkBox.setChecked(True)

    def temp(self):
        self.pushButton_2.setAttribute(Qt.WA_UnderMouse, False)
        global DB
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='TEST_AKT1')
        except pymysql.err.OperationalError:
            QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            return

        self.cursor = DB.cursor()

        self.id = '3001'
        self.password = 'akt1765323'
        sql = """SELECT PASSWORD, PRIORITY FROM login_pass WHERE ID=%s"""
        self.cursor.execute(sql, (self.id))
        pass_check = self.cursor.fetchall()
        if pass_check == ():
            QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            # print('Wrong user id!')
            self.cursor.close()

            return
        if self.password == pass_check[0][0]:
            if self.comboBox.currentText() == 'Administrator':
                if pass_check[0][1] == 'admin':
                    self.to_main_win(mode='admin')
                else:
                    QMessageBox.warning(self, 'warning',
                                        'Sorry, you do not have the administration authority, please select the "Normal" mode.')

            else:
                self.to_main_win(mode='normal')

        else:
            QMessageBox.warning(self, 'Warning', 'Wrong password!')
            # print('Wrong password!')

        self.cursor.close()

    def login(self):
        self.pushButton.setAttribute(Qt.WA_UnderMouse, False)

        global DB
        try:
            DB.close()
        except:
            pass
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='TEST_AKT1')
        except pymysql.err.OperationalError:
            QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            return

        self.cursor_version = DB.cursor()
        SQL = """SELECT VERSION FROM version_control WHERE ID=%s"""
        self.cursor_version.execute(SQL, (1))
        results=self.cursor_version.fetchall()
        self.cursor_version.close()
        if results == ():
            QMessageBox.information(self, 'Info' , 'Sorry, the system is under maintenance, please try it later...')
            return
        version=results[0][0]
        if version==None:
            QMessageBox.information(self, 'Info', 'Sorry, the system and database is under maintenance, please try it later...')
            return

        current_version = 1.1
        if current_version < version:
            QMessageBox.information(self, 'Version Too Old', f'Sorry, the version you are using is too old, please update to version {version} first!')
            return

        self.cursor = DB.cursor()

        self.id = self.lineEdit.text()
        self.password = self.lineEdit_2.text()
        sql = """SELECT PASSWORD, PRIORITY FROM login_pass WHERE ID=%s"""
        self.cursor.execute(sql, (self.id))
        pass_check = self.cursor.fetchall()
        if pass_check == ():
            QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            # print('Wrong user id!')
            self.cursor.close()

            return
        if self.password == pass_check[0][0]:
            if self.comboBox.currentText() == 'Administrator':
                if pass_check[0][1] == 'admin':
                    self.to_main_win(mode='admin')
                else:
                    QMessageBox.warning(self, 'warning',
                                        'Sorry, you do not have the administration authority, please select the "Normal" mode.')

            else:
                self.to_main_win(mode='normal')

        else:
            QMessageBox.warning(self, 'Warning', 'Wrong password!')
            # print('Wrong password!')

        self.cursor.close()

    def to_main_win(self, mode):
        global ID
        ID = self.id

        global MainWindow
        MainWindow.show()
        # print(mode!='admin')
        if mode != 'admin':
            MainWindow.pushButton_5.setEnabled(False)
            MainWindow.pushButton_5.setStyleSheet("QPushButton\n"
                                                  "{\n"
                                                  "    /*字体为微软雅黑*/\n"
                                                  "    font-family:Microsoft Yahei;\n"
                                                  "    /*字体大小为20点*/\n"
                                                  "    font-size:15pt;\n"
                                                  "    /*字体颜色为白色*/    \n"
                                                  "    color:white;\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(133, 173, 173);\n"
                                                  "    /*边框圆角半径为8像素*/ \n"
                                                  "    border-radius:10px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮停留态*/\n"
                                                  "QPushButton:hover\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(58, 0, 175);\n"
                                                  "    padding-left:-3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:-3px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮按下态*/\n"
                                                  "QPushButton:pressed\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(57, 0, 122);\n"
                                                  "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                  "    padding-left:3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:3px;\n"
                                                  "}")
        else:
            MainWindow.pushButton_5.setEnabled(True)
            MainWindow.pushButton_5.setStyleSheet("QPushButton\n"
                                                  "{\n"
                                                  "    /*字体为微软雅黑*/\n"
                                                  "    font-family:Microsoft Yahei;\n"
                                                  "    /*字体大小为20点*/\n"
                                                  "    font-size:15pt;\n"
                                                  "    /*字体颜色为白色*/    \n"
                                                  "    color:white;\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(85, 0, 255);\n"
                                                  "    /*边框圆角半径为8像素*/ \n"
                                                  "    border-radius:10px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮停留态*/\n"
                                                  "QPushButton:hover\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(58, 0, 175);\n"
                                                  "    padding-left:-3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:-3px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮按下态*/\n"
                                                  "QPushButton:pressed\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(57, 0, 122);\n"
                                                  "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                  "    padding-left:3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:3px;\n"
                                                  "}")
        MainWindow.init_db()

        global loginWindow
        loginWindow.destroy()
        TimeCard.initial()

        if self.checkBox.isChecked():
            with open('info.csv', 'w', newline='') as file:
                writer=csv.writer(file)
                writer.writerow([self.lineEdit.text(), self.lineEdit_2.text(),self.comboBox.currentText(), '1'])
            file.close()
        else:
            try:
                os.remove('info.csv')
            except:
                pass

    def quit(self):
        if DB != None:
            DB.close()
        sys.exit()

    def closeEvent(self, event):
        if DB != None:
            DB.close()
        sys.exit()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.pushButton_7.clicked.connect(self.quit)
        self.pushButton_6.clicked.connect(self.to_login)
        self.pushButton_8.clicked.connect(self.to_passchange)
        self.pushButton.clicked.connect(self.to_timecard)
        self.pushButton_2.clicked.connect(self.to_askforleave)
        self.pushButton_3.clicked.connect(self.to_otapplication)
        self.pushButton_4.clicked.connect(self.to_bookmeetingroom)
        self.pushButton_9.clicked.connect(self.to_approvepanel)
        self.pushButton_5.clicked.connect(self.to_main_admin)

    def to_main_admin(self):
        self.pushButton_5.setAttribute(Qt.WA_UnderMouse, False)
        AdminMain.show()
        MainWindow.setVisible(False)

    def to_approvepanel(self):
        self.pushButton_9.setAttribute(Qt.WA_UnderMouse,False)
        ApprovePanel.show()
        ApprovePanel.initializing()
        MainWindow.destroy()

    def to_bookmeetingroom(self):
        self.pushButton_4.setAttribute(Qt.WA_UnderMouse, False)
        BookMeetingRoom.show()
        BookMeetingRoom.initializing()
        MainWindow.destroy()

    def to_otapplication(self):
        self.pushButton_3.setAttribute(Qt.WA_UnderMouse, False)
        OTApplication.show()
        OTApplication.initializing()
        MainWindow.destroy()

    def to_askforleave(self):
        self.pushButton_2.setAttribute(Qt.WA_UnderMouse, False)
        AskForLeave.show()
        AskForLeave.initializing()
        MainWindow.destroy()

    def to_timecard(self):
        self.pushButton.setAttribute(Qt.WA_UnderMouse, False)
        TimeCard.show()
        TimeCard.startTimer()
        MainWindow.destroy()

    def to_passchange(self):
        self.pushButton_8.setAttribute(Qt.WA_UnderMouse, False)
        PassWindow.show()
        PassWindow.initializing()
        MainWindow.setEnabled(False)

    def to_login(self):
        self.pushButton_6.setAttribute(Qt.WA_UnderMouse, False)
        MainWindow.destroy()
        TimeCard.timerStop()
        loginWindow.show()

    def quit(self):
        a = QMessageBox.question(self, 'Query', 'Are you sure to quit the system?', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            return

        # print(12345)
        if DB != None:
            DB.close()
        sys.exit()

    def init_db(self):
        self.cursor = DB.cursor()
        sql = """SELECT NAME, POSITION FROM akt_staff WHERE ID=%s"""
        self.cursor.execute(sql, (ID))
        res = self.cursor.fetchall()
        self.name = res[0][0]
        self.label.setText(f'Hi, {self.name}, welcome!')
        self.position = res[0][1]
        # print(self.position, type(self.position))
        if self.position.strip() != 'Leader' and self.position.strip() != 'DM' and self.position.strip() != 'MD':
            # print(111)
            self.pushButton_9.setEnabled(False)
            self.pushButton_9.setStyleSheet("QPushButton\n"
                                            "{\n"
                                            "    /*字体为微软雅黑*/\n"
                                            "    font-family:Microsoft Yahei;\n"
                                            "    /*字体大小为20点*/\n"
                                            "    font-size:15pt;\n"
                                            "    /*字体颜色为白色*/    \n"
                                            "    color:white;\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(133, 173, 173);\n"
                                            "    /*边框圆角半径为8像素*/ \n"
                                            "    border-radius:10px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮停留态*/\n"
                                            "QPushButton:hover\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(58, 0, 175);\n"
                                            "    padding-left:-3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:-3px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮按下态*/\n"
                                            "QPushButton:pressed\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(57, 0, 122);\n"
                                            "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                            "    padding-left:3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:3px;\n"
                                            "}")
        else:
            self.pushButton_9.setEnabled(True)
            self.pushButton_9.setStyleSheet("QPushButton\n"
                                            "{\n"
                                            "    /*字体为微软雅黑*/\n"
                                            "    font-family:Microsoft Yahei;\n"
                                            "    /*字体大小为20点*/\n"
                                            "    font-size:15pt;\n"
                                            "    /*字体颜色为白色*/    \n"
                                            "    color:white;\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(85,0,255);\n"
                                            "    /*边框圆角半径为8像素*/ \n"
                                            "    border-radius:10px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮停留态*/\n"
                                            "QPushButton:hover\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(58, 0, 175);\n"
                                            "    padding-left:-3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:-3px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮按下态*/\n"
                                            "QPushButton:pressed\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:rgb(57, 0, 122);\n"
                                            "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                            "    padding-left:3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:3px;\n"
                                            "}")

        self.cursor.close()

    def closeEvent(self, event):
        a = QMessageBox.question(self, 'Query', 'Are you sure to quit the system?', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            event.ignore()
            return

        self.cursor.close()
        if DB != None:
            DB.close()
        sys.exit()


class WorkThread(QThread):
    trigger = pyqtSignal(str)

    def __int__(self,calendarWidget):
        super(WorkThread, self).__init__()
        self.calendarWidget=calendarWidget

    def run(self):
        qdt = self.calendarWidget.selectedDate()
        dt_for_query = datetime.datetime.strptime(qdt.toString('dd/MM/yyyy'), '%d/%m/%Y')
        sql = """SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
        self.cursor_queryhours = DB.cursor()
        try:
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        except:
            reconnect_DB(self)
            self.cursor_queryhours = DB.cursor()
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        res = self.cursor_queryhours.fetchall()
        self.cursor_queryhours.close()
        if res == ():
            self.textEdit_5.setText('0')
        else:
            ot_hours = str(res[0][0])
            self.textEdit_5.setText(ot_hours)

        dt = qdt.toString('yyyyMMdd')
        user_id = str(ID)
        serial = user_id + dt
        sql = """SELECT CLOCK_IN, CLOCK_OUT, DAY_LAG, OUT_1, IN_1, OUT_2, IN_2 FROM time_card WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except:
            self.cursor = DB.cursor()
            try:
                self.cursor.execute(sql, (serial))
            except:
                reconnect_DB(self)
                self.cursor = DB.cursor()
                self.cursor.execute(sql, (serial))

        result = self.cursor.fetchall()
        if result == ():
            self.textEdit.setText('-')
            self.textEdit_2.setText('-')
            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')

            self.textEdit_6.setText('-')
            self.textEdit_7.setText('-')
            self.textEdit_8.setText('-')
            self.textEdit_9.setText('-')
            return

        day_lag = result[0][2]

        clock_out = result[0][1]

        if clock_out == None:
            clock_out = '-'

        clock_in = result[0][0]

        if clock_in == None:
            clock_in = '-'

        try:
            out1 = result[0][3]
        except:
            out1 = '-'
        if out1 == None:
            out1 = '-'
        try:
            in1 = result[0][4]
        except:
            in1 = '-'
        if in1 == None:
            in1 = '-'
        try:
            out2 = result[0][5]
        except:
            out2 = '-'
        if out2 == None:
            out2 = '-'
        try:
            in2 = result[0][6]
        except:
            in2 = '-'
        if in2 == None:
            in2 = '-'

        if clock_out == '-' or clock_in == '-':
            work_time = '-'
            over_time = '-'
        # ---------------------计算实际作业时间
        else:
            if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                # 1
                if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    work_time = (clock_out - clock_in).seconds
                # 2
                elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 3
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 3600
                # 4
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                    work_time -= 3600
                # 5
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 4500

            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                # 6
                if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = 0
                # 7
                elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                # 8
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230',
                        '%Y%m%d%H%M')).seconds
                # 9
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                # 10
                if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                # 11
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 12
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                # 13
                if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = 0
                # 14
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                        '%Y%m%d%H%M')).seconds
            else:
                work_time = (clock_out - clock_in).seconds

            # ------------------计算离岗累计时间
            space_time1 = 0
            space_time2 = 0
            if out1 != '-' and in1 != '-':
                # 1
                if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time1 = (in1 - out1).seconds
                # 2
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 3
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds - 3600
                # 4
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds - 3600
                # 5
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 3600 - 900
                # 6
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = 0
                # 7
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds
                # 11
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 12
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 900
                # 13
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = 0
                # 14
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time1 = (in1 - out1).seconds

            # -----------------out2 in2
            if out2 != '-' and in2 != '-':
                # 1
                if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time2 = (in2 - out2).seconds
                # 2
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 3
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds - 3600
                # 4
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds - 3600
                # 5
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 3600 - 900
                # 6
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = 0
                # 7
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds
                # 11
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 12
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 900
                # 13
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = 0
                # 14
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time2 = (in2 - out2).seconds

            work_time -= space_time1
            work_time -= space_time2
            # --------------------------如过午夜加上一天
            if day_lag == 1:
                work_time += 86400

            over_time = work_time - 28800
            if over_time < 0:
                over_time = 0

        if work_time != '-':
            work_time = round(int(work_time) / 3600, 2)
        if over_time != '-':
            over_time = round(int(over_time) / 3600, 2)
        if clock_in != '-':
            self.textEdit.setText(clock_in.strftime('%H:%M:%S'))
        else:
            self.textEdit.setText(clock_in)
        if clock_out != '-':
            self.textEdit_2.setText(clock_out.strftime('%H:%M:%S'))
        else:
            self.textEdit_2.setText(clock_out)
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))
        if out1 != '-':
            self.textEdit_6.setText(out1.strftime('%H:%M:%S'))
        else:
            self.textEdit_6.setText(out1)
        if in1 != '-':
            self.textEdit_7.setText(in1.strftime('%H:%M:%S'))
        else:
            self.textEdit_7.setText(in1)

        if out2 != '-':
            self.textEdit_8.setText(out2.strftime('%H:%M:%S'))
        else:
            self.textEdit_8.setText(out2)
        if in2 != '-':
            self.textEdit_9.setText(in2.strftime('%H:%M:%S'))
        else:
            self.textEdit_9.setText(in2)

        self.trigger.emit(str(i))

class TimeCard(QMainWindow, Ui_TimeCard):
    def __init__(self):
        super(TimeCard, self).__init__()
        self.setupUi(self)

        self.textEdit.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_2.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_3.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_4.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_5.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_6.setEnabled(False)
        self.textEdit_7.setEnabled(False)
        self.textEdit_8.setEnabled(False)
        self.textEdit_9.setEnabled(False)

        self.NTP = NTP_DateTime()
        self.pushButton_6.clicked.connect(self.quit)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.showtime)

        self.pushButton.clicked.connect(self.clock_in)
        self.pushButton_2.clicked.connect(self.clock_out)
        self.pushButton_5.clicked.connect(self.cancel_clock_out)
        self.calendarWidget.clicked.connect(self.select_date)
        #self.calendarWidget.clicked.connect(self.startTimer)
        self.pushButton_7.clicked.connect(self.to_applylateclockin)
        self.pushButton_3.clicked.connect(self.to_forgetrecord)
        self.pushButton_4.clicked.connect(self.export_excel)

    def export_excel(self):
        yearmonth = str(self.calendarWidget.yearShown()) + "%02d" % self.calendarWidget.monthShown()

        self.cursor_exc=DB.cursor()
        SQL="""SELECT * FROM time_card WHERE SERIAL>%s and SERIAL<%s"""
        try:
            self.cursor_exc.execute(SQL, (int(str(ID)+yearmonth+'00'), int(str(ID)+yearmonth+'32') ))
        except:
            reconnect_DB(self)
            self.cursor_exc = DB.cursor()
            self.cursor_exc.execute(SQL, (int(str(ID) + yearmonth + '00'), int(str(ID) + yearmonth + '32')))

        res=self.cursor_exc.fetchall()
        self.cursor_exc.close()
        #print(res)
        if res==():
            QMessageBox.information(self, 'Empty Record', f'The time card record in the year-month you selected({yearmonth}) is empty!')
            return

        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s timecard({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        #print(a[0])
        wb=xl.Workbook()
        ws=wb.active
        headers=['Series No', 'User ID', 'Clock In', 'Clock Out', 'Out 1', 'In 1', 'Out 2', 'In 2', 'Day Lag', 'Actual Work Time', 'Over Time', 'Approved OT']
        ws.append(headers)
        for each_line in res:
            line_head=each_line
            data_set=calculate_worktime(data_line=each_line)
            worktime=data_set[0]
            overtime=data_set[1]
            approved_ot=data_set[2]
            line_end=[worktime, overtime, approved_ot]
            line_full=list(line_head)+line_end
            ws.append(line_full)
            #print(line_full)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def to_forgetrecord(self):
        self.pushButton_3.setAttribute(Qt.WA_UnderMouse, False)
        ForgetRecord.show()
        ForgetRecord.initialize()
        TimeCard.destroy()

    def to_applylateclockin(self):
        self.pushButton_7.setAttribute(Qt.WA_UnderMouse, False)
        ApplyLateClockIn.show()
        ApplyLateClockIn.initialize()
        TimeCard.destroy()

    def select_date(self):
        qdt = self.calendarWidget.selectedDate()
        dt_for_query=datetime.datetime.strptime(qdt.toString('dd/MM/yyyy'),'%d/%m/%Y')
        sql="""SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
        self.cursor_queryhours=DB.cursor()
        try:
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        except:
            reconnect_DB(self)
            self.cursor_queryhours=DB.cursor()
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        res=self.cursor_queryhours.fetchall()
        self.cursor_queryhours.close()
        if res==():
            self.textEdit_5.setText('0')
        else:
            ot_hours=str(res[0][0])
            self.textEdit_5.setText(ot_hours)

        dt = qdt.toString('yyyyMMdd')
        user_id = str(ID)
        serial = user_id + dt
        sql = """SELECT CLOCK_IN, CLOCK_OUT, DAY_LAG, OUT_1, IN_1, OUT_2, IN_2 FROM time_card WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except:
            self.cursor = DB.cursor()
            try:
                self.cursor.execute(sql, (serial))
            except:
                reconnect_DB(self)
                self.cursor=DB.cursor()
                self.cursor.execute(sql, (serial))

        result = self.cursor.fetchall()
        if result == ():
            self.textEdit.setText('-')
            self.textEdit_2.setText('-')
            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')

            self.textEdit_6.setText('-')
            self.textEdit_7.setText('-')
            self.textEdit_8.setText('-')
            self.textEdit_9.setText('-')
            return

        day_lag = result[0][2]

        clock_out = result[0][1]

        if clock_out == None:
            clock_out = '-'

        clock_in = result[0][0]

        if clock_in == None:
            clock_in = '-'

        try:
            out1 = result[0][3]
        except:
            out1 = '-'
        if out1 == None:
            out1 = '-'
        try:
            in1 = result[0][4]
        except:
            in1 = '-'
        if in1 == None:
            in1 = '-'
        try:
            out2 = result[0][5]
        except:
            out2 = '-'
        if out2 == None:
            out2 = '-'
        try:
            in2 = result[0][6]
        except:
            in2 = '-'
        if in2 == None:
            in2 = '-'

        if clock_out == '-' or clock_in=='-':
            work_time = '-'
            over_time = '-'
        # ---------------------计算实际作业时间
        else:
            if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                #1
                if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    work_time = (clock_out - clock_in).seconds
                #2
                elif int(clock_in.strftime('%Y%m%d') + '1130')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                #3
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 3600
                #4
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                    work_time -= 3600
                #5
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 4500

            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                #6
                if int(clock_in.strftime('%Y%m%d') + '1130')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                    work_time= 0
                #7
                elif int(clock_in.strftime('%Y%m%d') + '1230')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                              '%Y%m%d%H%M')).seconds
                #8
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time=(datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                              '%Y%m%d%H%M')).seconds
                #9
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                    work_time-=900

            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                #10
                if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                #11
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                #12
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time-=900

            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                #13
                if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time=0
                #14
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                        '%Y%m%d%H%M')).seconds
            else:
                work_time = (clock_out - clock_in).seconds

            # ------------------计算离岗累计时间
            space_time1 = 0
            space_time2 = 0
            if out1 != '-' and in1 != '-':
                # 1
                if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time1 = (in1 - out1).seconds
                # 2
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 3
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds - 3600
                # 4
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds - 3600
                # 5
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 3600 - 900
                # 6
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = 0
                # 7
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds
                # 11
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 12
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 900
                # 13
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = 0
                # 14
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time1 = (in1 - out1).seconds

            # -----------------out2 in2
            if out2 != '-' and in2 != '-':
                # 1
                if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time2 = (in2 - out2).seconds
                # 2
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 3
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds - 3600
                # 4
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds - 3600
                # 5
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 3600 - 900
                # 6
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = 0
                # 7
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds
                # 11
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 12
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 900
                # 13
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = 0
                # 14
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time2 = (in2 - out2).seconds

            work_time -= space_time1
            work_time -= space_time2
            # --------------------------如过午夜加上一天
            if day_lag == 1:
                work_time += 86400

            over_time = work_time - 28800
            if over_time < 0:
                over_time = 0

        if work_time != '-':
            work_time = round(int(work_time) / 3600, 2)
        if over_time != '-':
            over_time = round(int(over_time) / 3600, 2)
        if clock_in != '-':
            self.textEdit.setText(clock_in.strftime('%H:%M:%S'))
        else:
            self.textEdit.setText(clock_in)
        if clock_out != '-':
            self.textEdit_2.setText(clock_out.strftime('%H:%M:%S'))
        else:
            self.textEdit_2.setText(clock_out)
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))
        if out1 != '-':
            self.textEdit_6.setText(out1.strftime('%H:%M:%S'))
        else:
            self.textEdit_6.setText(out1)
        if in1 != '-':
            self.textEdit_7.setText(in1.strftime('%H:%M:%S'))
        else:
            self.textEdit_7.setText(in1)

        if out2 != '-':
            self.textEdit_8.setText(out2.strftime('%H:%M:%S'))
        else:
            self.textEdit_8.setText(out2)
        if in2 != '-':
            self.textEdit_9.setText(in2.strftime('%H:%M:%S'))
        else:
            self.textEdit_9.setText(in2)

    def refresh_panel(self):
        dt = self.t.strftime('%Y%m%d')
        dt_for_query = datetime.datetime.strptime(self.t.strftime('%d/%m/%Y'),'%d/%m/%Y')
        sql = """SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
        self.cursor_queryhours = DB.cursor()
        try:
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        except:
            reconnect_DB(self)
            self.cursor_queryhours = DB.cursor()
            self.cursor_queryhours.execute(sql, (ID, dt_for_query))
        res = self.cursor_queryhours.fetchall()
        self.cursor_queryhours.close()
        if res == ():
            self.textEdit_5.setText('0')
        else:
            ot_hours = str(res[0][0])
            self.textEdit_5.setText(ot_hours)

        user_id = str(ID)
        serial = user_id + dt
        sql = """SELECT CLOCK_IN, CLOCK_OUT, DAY_LAG, OUT_1, IN_1, OUT_2, IN_2 FROM time_card WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except:
            self.cursor = DB.cursor()
            try:
                self.cursor.execute(sql, (serial))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor=DB.cursor()
                self.cursor.execute(sql, (serial))

        result = self.cursor.fetchall()
        if result == ():
            self.textEdit.setText('-')
            self.textEdit_2.setText('-')
            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')

            self.textEdit_6.setText('-')
            self.textEdit_7.setText('-')
            self.textEdit_8.setText('-')
            self.textEdit_9.setText('-')
            return

        day_lag = result[0][2]

        clock_out = result[0][1]
        if clock_out == None:
            clock_out = '-'

        clock_in = result[0][0]
        try:
            out1 = result[0][3]
        except:
            out1 = '-'
        if out1 == None:
            out1 = '-'
        try:
            in1 = result[0][4]
        except:
            in1='-'
        if in1 == None:
            in1 = '-'
        try:
            out2 = result[0][5]
        except:
            out2='-'
        if out2 == None:
            out2 = '-'
        try:
            in2 = result[0][6]
        except:
            in2='-'
        if in2 == None:
            in2 = '-'

        if clock_out == '-':
            work_time = '-'
            over_time = '-'
        else:
            if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                # 1
                if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    work_time = (clock_out - clock_in).seconds
                # 2
                elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 3
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 3600
                # 4
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                    work_time -= 3600
                # 5
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 4500

            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                # 6
                if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = 0
                # 7
                elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                # 8
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230',
                        '%Y%m%d%H%M')).seconds
                # 9
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                # 10
                if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                # 11
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 12
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                # 13
                if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = 0
                # 14
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                        '%Y%m%d%H%M')).seconds
            else:
                work_time = (clock_out - clock_in).seconds

            # ------------------计算离岗累计时间
            space_time1 = 0
            space_time2 = 0
            if out1 != '-' and in1 != '-':
                # 1
                if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time1 = (in1 - out1).seconds
                # 2
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 3
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds - 3600
                # 4
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds - 3600
                # 5
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 3600 - 900
                # 6
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = 0
                # 7
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds
                # 11
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 12
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 900
                # 13
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = 0
                # 14
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time1 = (in1 - out1).seconds

            # -----------------out2 in2
            if out2 != '-' and in2 != '-':
                # 1
                if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time2 = (in2 - out2).seconds
                # 2
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 3
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds - 3600
                # 4
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds - 3600
                # 5
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 3600 - 900
                # 6
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = 0
                # 7
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds
                # 11
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 12
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 900
                # 13
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = 0
                # 14
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time2 = (in2 - out2).seconds

            work_time -= space_time1
            work_time -= space_time2
            # --------------------------如过午夜加上一天
            if day_lag == 1:
                work_time += 86400

            over_time = work_time - 28800
            if over_time < 0:
                over_time = 0

        if work_time!='-':
            work_time = round(int(work_time) / 3600, 2)
        if over_time!='-':
            over_time = round(int(over_time) / 3600, 2)

        self.textEdit.setText(clock_in.strftime('%H:%M:%S'))
        if clock_out != '-':
            self.textEdit_2.setText(clock_out.strftime('%H:%M:%S'))
        else:
            self.textEdit_2.setText(clock_out)
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))

        if out1 != '-':
            self.textEdit_6.setText(out1.strftime('%H:%M:%S'))
        else:
            self.textEdit_6.setText(out1)
        if in1 != '-':
            self.textEdit_7.setText(in1.strftime('%H:%M:%S'))
        else:
            self.textEdit_7.setText(in1)

        if out2 != '-':
            self.textEdit_8.setText(out2.strftime('%H:%M:%S'))
        else:
            self.textEdit_8.setText(out2)
        if in2 != '-':
            self.textEdit_9.setText(in2.strftime('%H:%M:%S'))
        else:
            self.textEdit_9.setText(in2)

    def clock_in(self):
        a=QMessageBox.question(self, 'Confirmation', 'Are you sure to clock in?')
        if a==QMessageBox.No:
            return
        user_id = str(ID)
        serial = user_id + self.t.strftime('%Y%m%d')
        in_time = self.t
        # serial = user_id + (self.t-datetime.timedelta(days=1)).strftime('%Y%m%d')
        # in_time = self.t-datetime.timedelta(days=1)
        self.cursor_clock = DB.cursor()
        sql = """INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN) VALUES
        (%s,%s,%s)"""
        try:
            try:
                self.cursor_clock.execute(sql, (serial, user_id, in_time))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_clock=DB.cursor()
                self.cursor_clock.execute(sql, (serial, user_id, in_time))
                DB.commit()
            QMessageBox.information(self, 'Info', 'Clock-in completed!')
        except pymysql.err.IntegrityError:
            DB.rollback()
            sql = """SELECT CLOCK_OUT, OUT_1, OUT_2 FROM time_card WHERE SERIAL=%s"""
            self.cursor_clock.execute(sql, (serial))
            res = self.cursor_clock.fetchall()
            if res[0][0] == None:  # CLOCK_OUT==None
                QMessageBox.critical(self, 'Info', 'Sorry, you have already clocked in.')
            else:
                if res[0][1] == None:  # OUT_1==None
                    sql = """UPDATE time_card SET OUT_1=%s, CLOCK_OUT=%s, IN_1=%s WHERE SERIAL=%s"""
                    self.cursor_clock.execute(sql, (res[0][0], None, in_time, serial))
                    DB.commit()
                    QMessageBox.information(self, 'Info', 'Clock-in completed!')
                elif res[0][2] == None:  # OUT_2=None
                    sql = """UPDATE time_card SET OUT_2=%s, CLOCK_OUT=%s, IN_2=%s WHERE SERIAL=%s"""
                    self.cursor_clock.execute(sql, (res[0][0] ,None, in_time, serial))
                    DB.commit()
                    QMessageBox.information(self, 'Info', 'Clock-in completed!')
                else:
                    QMessageBox.critical(self, 'Warning', 'Sorry, you can not clock in more than 3 times per day.')

        self.cursor_clock.close()
        self.refresh_panel()

    def clock_out(self):
        a = QMessageBox.question(self, 'Confirmation', 'Are you sure to clock out?')
        if a == QMessageBox.No:
            return
        user_id = str(ID)
        if not self.checkBox.isChecked():
            serial = user_id + self.t.strftime('%Y%m%d')
            dt = self.t.strftime('%d/%m/%Y')
            day_lag = None
        else:
            serial = user_id + (self.t - datetime.timedelta(days=1)).strftime('%Y%m%d')
            dt = (self.t - datetime.timedelta(days=1)).strftime('%d/%m/%Y')
            day_lag = 1

        out_time = self.t
        self.cursor_clock = DB.cursor()
        sql = """SELECT SERIAL, CLOCK_OUT FROM time_card WHERE SERIAL=%s"""
        try:
            a = self.cursor_clock.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_clock=DB.cursor()
            a = self.cursor_clock.execute(sql, (serial))

        if a == 0:
            QMessageBox.critical(self, 'Warning', f'Clock-out failed! You did not clocked in for today({dt}).')
            self.cursor_clock.close()
            return

        temp = self.cursor_clock.fetchall()[0][1]
        if temp != None:
            QMessageBox.critical(self, 'Warning',
                                 f'You have already clocked out, if you want to clock out again, please clock in first, or cancel clock-out first.')
            self.cursor_clock.close()
            return

        sql = """UPDATE time_card SET CLOCK_OUT=%s, DAY_LAG=%s WHERE SERIAL=%s"""
        self.cursor_clock.execute(sql, (out_time, day_lag, serial))
        DB.commit()
        QMessageBox.information(self, 'Info', 'Clock-out completed!')
        self.cursor_clock.close()
        self.refresh_panel()

    def cancel_clock_out(self):
        user_id = str(ID)
        if not self.checkBox.isChecked():
            serial = user_id + self.t.strftime('%Y%m%d')
            dt = self.t.strftime('%d/%m/%Y')

        else:
            serial = user_id + (self.t - datetime.timedelta(days=1)).strftime('%Y%m%d')
            dt = (self.t - datetime.timedelta(days=1)).strftime('%d/%m/%Y')

        sql = """SELECT CLOCK_OUT FROM time_card WHERE SERIAL=%s"""
        try:
            a = self.cursor.execute(sql, (serial))
        except pymysql.err.ProgrammingError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            a = self.cursor.execute(sql, (serial))
        if a == 0:
            QMessageBox.critical(self, 'Warning', f'You did not clocked out for the latest time.')
            return

        temp = self.cursor.fetchall()[0][0]
        if temp == None:
            QMessageBox.critical(self, 'Warning', f'You did not clocked out for the latest time.')
            return

        target_time = temp.strftime('%d/%m/%Y %H:%M:%S')
        res = QMessageBox.question(self, 'Query', f'Are you sure to cancel the clock-out record: {target_time}?')
        if res == QMessageBox.No:
            return

        sql = """UPDATE time_card SET CLOCK_OUT=%s, DAY_LAG=%s WHERE SERIAL=%s"""
        self.cursor.execute(sql, (None, None, serial))
        DB.commit()
        QMessageBox.information(self, 'Info',
                                f'The latest clock-out record({target_time}) has been canceled successfully')
        self.refresh_panel()

    def initial(self):
        self.label_2.setText(f'Name: {MainWindow.name}  Staff ID:{ID}')
        self.cursor = DB.cursor()
        self.startTimer()
        self.refresh_panel()

    def timerStop(self):
        self.timer.stop()

    def showtime(self):
        if (self.t - self.t0).seconds == 600:
            self.t0 = self.NTP.get_datetime()
            self.t = self.t0
            # print('Reset time!', self.t0.strftime('%H:%M:%S'))
        self.t += datetime.timedelta(seconds=1)
        _time = self.t.strftime('%H:%M:%S')
        # print(_time)
        self.lcdNumber.display(_time)
        _date = self.t.strftime('%d/%m/%Y')
        self.label_9.setText(_date)

    def startTimer(self):
        self.t0 = self.NTP.get_datetime()
        self.t = self.t0
        self.timer.start(1000)

    def quit(self):
        TimeCard.close()

    def closeEvent(self, event):
        self.cursor.close()
        MainWindow.show()


class AskForLeave(QMainWindow, Ui_AskForLeave):
    def __init__(self):
        super(AskForLeave, self).__init__()
        self.setupUi(self)

        self.dateEdit_5.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_2.setDate(QDate.currentDate())
        self.dateEdit_7.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit_6.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_6.setCurrentText(month)

        self.geoTab = self.tabWidget.geometry()
        self.geoTable = self.tableWidget.geometry()
        self.geoBTN7 = self.pushButton_7.geometry()
        self.geoBTN16 = self.pushButton_16.geometry()
        self.geoBTN2 = self.pushButton_2.geometry()
        self.geoBTN4 = self.pushButton_4.geometry()
        self.geoFrame = self.frame.geometry()

        self.pushButton_4.clicked.connect(self.quit)
        self.checkBox.clicked.connect(self.is_multiple)
        self.pushButton.clicked.connect(self.request_per_leave)
        self.checkBox_2.clicked.connect(self.is_query)
        self.dateEdit_7.dateChanged.connect(self.show_IDs)
        self.comboBox_2.currentTextChanged.connect(self.input_id)
        self.lineEdit_2.textChanged.connect(self.refresh_data)
        self.pushButton_5.clicked.connect(self.del_single_request)
        self.pushButton_3.clicked.connect(self.request_go_home)
        self.checkBox_4.clicked.connect(self.is_query_hometown)
        self.lineEdit_3.textChanged.connect(self.get_id_hometown)
        self.comboBox_3.currentTextChanged.connect(self.refresh_hometown)
        self.pushButton_6.clicked.connect(self.del_hometown_request)
        self.pushButton_16.clicked.connect(self.download_data)
        self.pushButton_7.clicked.connect(self.cancel_request_on_table)
        self.pushButton_2.clicked.connect(self.export_excel)

        self.dateEdit_6.dateTimeChanged.connect(self.download_data)
        self.comboBox_6.currentTextChanged.connect(self.download_data)
        self.pushButton_16.setVisible(False)

        self.tabWidget.currentChanged.connect(self.tableview)

    def export_excel(self):
        try:
            first_cell=self.tableWidget.item(0, 0).text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Table', 'Warning: The table is empty, no need to export the excel file!')
            return

        yearmonth=self.dateEdit_6.text()+self.comboBox_6.currentText()
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s Leave Application({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return
        excel_path = a[0]

        data_frame=[]
        for i in range(self.tableWidget.rowCount()):
            data_line=[]
            for j in range(self.tableWidget.columnCount()):
                data_line.append(self.tableWidget.item(i, j).text())
            data_frame.append(data_line)


        wb=xl.Workbook()
        ws=wb.active
        ws.append(['Request ID', 'Staff ID', 'Staff Name', 'Type', 'Request Date-Time', 'Request Date',
                   'Leave From', 'Time to Start', 'Leave Until', 'Time to End', 'If Single Day', 'Duration','Remarks',
                   'Leader','DM','MD','HR'])

        for line in data_frame:
            ws.append(line)

        try:
            wb.save(filename=excel_path)
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info','Excel file has been exported successfully!')

    def tableview(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        # print(self.tabWidget.currentIndex())
        if self.tabWidget.currentIndex() == 2:
            width0 = W_LEAVE
            height0 = H_LEAVE
            self.desktop = QApplication.desktop()
            screen_count=self.desktop.screenCount() #2 screen version updating
            geometry = self.desktop.geometry()
            if screen_count==1:                     #2 screen version updating
                width1 = geometry.width()           #2 screen version updating
            else:                                   #2 screen version updating
                width1 = geometry.width()/2         #2 screen version updating
            height1 = geometry.height()
            # print(width0, height0, width1, height1)
            x = width1 - width0
            y = height1 - height0 - 100

            self.setMaximumSize(16777215, 16777215)
            self.showMaximized()

            move_widgets_for_frame(self.tabWidget, x, y)
            move_widgets_for_frame(self.tableWidget, x, y)
            move_widgets_only_vertical(self.pushButton_7, x, y)
            move_widgets_only_horizontal(self.pushButton_16, x, y)
            move_widgets_only_vertical(self.pushButton_2, x, y)
            move_widgets_both(self.pushButton_4, x, y)
            extend_widgets_horizontal(self.frame, x, y)
        else:
            self.tabWidget.setGeometry(self.geoTab.x(), self.geoTab.y(), self.geoTab.width(), self.geoTab.height())
            self.tableWidget.setGeometry(self.geoTable.x(), self.geoTable.y(), self.geoTable.width(),
                                         self.geoTable.height())
            self.pushButton_7.setGeometry(self.geoBTN7.x(), self.geoBTN7.y(), self.geoBTN7.width(),
                                          self.geoBTN7.height())
            self.pushButton_16.setGeometry(self.geoBTN16.x(), self.geoBTN16.y(), self.geoBTN16.width(),
                                           self.geoBTN16.height())
            self.pushButton_4.setGeometry(self.geoBTN4.x(), self.geoBTN4.y(), self.geoBTN4.width(),
                                          self.geoBTN4.height())
            self.pushButton_2.setGeometry(self.geoBTN2)
            self.frame.setGeometry(self.geoFrame)
            self.showNormal()
            self.setMaximumSize(W_LEAVE, H_LEAVE)

    def initializing(self):
        self.dateEdite_style_highlight = "QDateEdit{ font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 138, 163);}"
        self.comboBox_style_highlight = "QComboBox{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:9pt;background-color:rgb(255, 138, 163)}"
        self.textEdit_style_highlight = "QTextEdit{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 138, 163);}"
        self.radio_style_highlight = "QRadioButton{font-family:Microsoft Yahei;font-size:12pt;color:rgb(74, 74, 74); background-color:rgb(255, 138, 163);}"
        self.lineEdite_style_highlight = "QLineEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:black;   font-size:12pt;	background-color:rgb(200, 200, 200);}QLineEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QLineEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"

        self.dateEdite_style_normal = "QDateEdit{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 255, 255);}QLineEdit:hover{background-color:rgb(229, 242, 255);}QLineEdit:focus{background-color:rgb(229, 242, 255);}"
        self.comboBox_style_normal = "QComboBox    {font-family:Microsoft Yahei;    /*字体大小为20点*/	color:rgb(74, 74, 74);   font-size:9pt;}QComboBox:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QComboBox:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"
        self.textEdit_style_normal = "QTextEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTextEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QTextEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"
        self.radio_style_normal = "QRadioButton{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/    font-size:12pt;    /*字体颜色为白色*/        color:rgb(74, 74, 74);    /*背景颜色*/  }/*按钮停留态*/QRadioButton:hover{    /*背景颜色*/      background-color:rgb(21, 175, 0);	font-size:14pt;    padding-left:-3pt;	font-weight:bold;	color:white;    /*上内边距为3像素，让按下时字向下移动3像素*/      padding-top:-3pt;}/*按钮按下态*/QRadioButton:pressed{    /*背景颜色*/      background-color:rgb(0, 113, 0);	font-weight:bold;	font-size:14pt;	color:white;    /*左内边距为3像素，让按下时字向右移动3像素*/      padding-left:3pt;    /*上内边距为3像素，让按下时字向下移动3像素*/      padding-top:3pt;}"
        self.lineEdite_style_normal = "QLineEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:grey;   font-size:12pt;	background-color:rgb(255, 255, 255);}QLineEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QLineEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"

        self.cursor = DB.cursor()
        sql = """SELECT AN_DAYS, SICK_DAYS, HOME_TOWN FROM akt_staff WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (str(ID)))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (str(ID)))

        res = self.cursor.fetchall()
        self.annual_remain=res[0][0]
        self.label_10.setText(f'{self.annual_remain} DAYS')
        self.sick_remain=res[0][1]
        self.label_18.setText(f'{self.sick_remain} DAYS')
        self.is_query()
        self.home_town = res[0][2]
        self.label_25.setText(f'Remains {self.home_town} time(s) for traveling to hometown in this year.')

        self.download_data()

    def cancel_request_on_table(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return
        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 16).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return
        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.is_query()
        self.is_query_hometown()

        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.download_data()

    def download_data(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_6.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit_6.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_6.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit_6.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + self.comboBox_6.currentText() + '01',
                                                  '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit_6.text() + str(int(self.comboBox_6.currentText()) + 1) + '01', '%Y%m%d')

        sql = """SELECT * FROM leave_request WHERE USER_ID=%s AND (APPLY_DTTM>=%s AND APPLY_DTTM<%s)"""
        try:
            self.cursor.execute(sql, (ID, date_min, date_max))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID, date_min, date_max))

        res = self.cursor.fetchall()
        if res == ():
            #QMessageBox.information(self, 'Empty Record', 'No record in the selected time range.')
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            for j in range(len(res[0])):
                value=res[i][j]
                if j in [13, 14, 15, 16]:
                    if res[i][j]==None:
                        value='Unconfirmed'
                    elif res[i][j]==1:
                        value='OK'
                    else:
                        value='Declined'
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(value)))

        for i in range(17):
            if i == 12:
                continue
            self.tableWidget.resizeColumnToContents(i)

    def del_hometown_request(self):
        if self.comboBox_3.currentText() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.comboBox_3.currentText())
        sql = """SELECT SERIAL, TYPE, APPLY_DTTM, START_DT,START_LEN,END_DT, END_LEN, DURING, REMARKS, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()
        request_id = res[0][0]
        type = res[0][1]
        apply_dt = res[0][2]
        start_dt = res[0][3]
        start_len = res[0][4]
        end_dt = res[0][5]
        end_len = res[0][6]
        during = res[0][7]
        destination = res[0][8]
        leader = res[0][9]
        dm = res[0][10]
        md = res[0][11]
        hr = res[0][12]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(hr)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        end_dt_func = lambda y: None if y == None else y.strftime("%d/%m/%Y")
        msm = f'Are you sure to delete this request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {apply_dt}\n' \
              f'Destination: {destination}\n' \
              f'Type: {type}\n' \
              f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Start from: {start_len}\n' \
              f'Until(Date): {end_dt_func(end_dt)}\n' \
              f'Until(time): {end_len}\n' \
              f'Duration: {during} day(s)\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'MD approving: {approve_func(md)}\n' \
              f'HR approving: {approve_func(hr)}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        self.cursor.execute(sql, (serial))
        DB.commit()
        self.is_query_hometown()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been deleted successfully!')

    def default_css_hometown(self):
        self.lineEdit.setStyleSheet(self.lineEdite_style_normal)
        self.dateEdit_3.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit_4.setStyleSheet(self.dateEdite_style_normal)
        self.pushButton_12.setText('―')
        self.pushButton_13.setText('―')
        self.pushButton_14.setText('―')
        self.pushButton_15.setText('―')

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_12.setStyleSheet(btn_yellow_css)
        self.pushButton_13.setStyleSheet(btn_yellow_css)
        self.pushButton_14.setStyleSheet(btn_yellow_css)
        self.pushButton_15.setStyleSheet(btn_yellow_css)

        self.label_38.setText('WAITING...')
        self.label_38.setStyleSheet("QLabel{color:rgb(255, 170, 0);}")

    def refresh_hometown(self):
        self.default_css_hometown()
        if self.comboBox_3.currentText() == '':
            return
        sql = """SELECT REMARKS, START_DT, END_DT, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (self.comboBox_3.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (self.comboBox_3.currentText()))

        res = self.cursor.fetchall()
        destination = res[0][0]
        start_dt = res[0][1]
        end_dt = res[0][2]
        leader = res[0][3]
        dm = res[0][4]
        md = res[0][5]
        hr = res[0][6]

        self.lineEdit.setText(destination)
        self.lineEdit.setStyleSheet(self.lineEdite_style_highlight)
        self.dateEdit_3.setDate(start_dt)
        self.dateEdit_3.setStyleSheet(self.dateEdite_style_highlight)
        self.dateEdit_4.setDate(end_dt)
        self.dateEdit_4.setStyleSheet(self.dateEdite_style_highlight)

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"

        if leader == None:
            self.pushButton_12.setText('―')
            self.pushButton_12.setStyleSheet(btn_yellow_css)
        elif leader == 1:
            self.pushButton_12.setText('〇')
            self.pushButton_12.setStyleSheet(btn_green_css)
        else:
            self.pushButton_12.setText('×')
            self.pushButton_12.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_13.setText('―')
            self.pushButton_13.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_13.setText('〇')
            self.pushButton_13.setStyleSheet(btn_green_css)
        else:
            self.pushButton_13.setText('×')
            self.pushButton_13.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_14.setText('―')
            self.pushButton_14.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_14.setText('〇')
            self.pushButton_14.setStyleSheet(btn_green_css)
        else:
            self.pushButton_14.setText('×')
            self.pushButton_14.setStyleSheet(btn_red_css)

        if hr == None:
            self.pushButton_15.setText('―')
            self.pushButton_15.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_15.setText('〇')
            self.pushButton_15.setStyleSheet(btn_green_css)
        else:
            self.pushButton_15.setText('×')
            self.pushButton_15.setStyleSheet(btn_red_css)

        if self.pushButton_12.text() == '×' or self.pushButton_13.text() == '×' or self.pushButton_14.text() == '×' or self.pushButton_15.text() == '×':
            self.label_38.setText('DECLINED')
            self.label_38.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_15.text() == '〇':
            self.label_38.setText('OK')
            self.label_38.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_38.setText('WAITING...')
            self.label_38.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def get_id_hometown(self):
        self.comboBox_3.clear()
        if self.lineEdit_3.text() == '':
            self.cursor_2 = DB.cursor()
            sql = """SELECT SERIAL FROM leave_request WHERE USER_ID=%s AND TYPE=%s"""
            try:
                self.cursor_2.execute(sql, (ID, 'hometown'))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_2=DB.cursor()
                self.cursor_2.execute(sql, (ID, 'hometown'))

            res = self.cursor_2.fetchall()
            if res == ():
                self.cursor_2.close()
                return

            id_lst = []
            for each in res:
                id_lst.append(str(each[0]))
            self.comboBox_3.addItems(id_lst)
            self.cursor_2.close()
            return

        year = self.lineEdit_3.text()
        try:
            dttm0 = datetime.datetime.strptime(year + '0101', '%Y%m%d')
            dttm1 = datetime.datetime.strptime(str(int(year) + 1) + '0101', '%Y%m%d')
        except:
            return
        self.cursor_2 = DB.cursor()
        sql = """SELECT SERIAL FROM leave_request WHERE TYPE=%s AND APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID=%s"""
        self.cursor_2.execute(sql, ('hometown', dttm0, dttm1, ID))
        res = self.cursor_2.fetchall()
        if res == ():
            self.cursor_2.close()
            return
        id_lst = []
        for each in res:
            id_lst.append(str(each[0]))
        self.comboBox_3.addItems(id_lst)
        self.cursor_2.close()

    def query_mode_hometown(self):
        self.lineEdit.setEnabled(False)
        self.dateEdit_3.setEnabled(False)
        self.dateEdit_4.setEnabled(False)
        self.pushButton_3.setEnabled(False)

        self.comboBox_3.setEnabled(True)
        self.lineEdit_3.setEnabled(True)
        self.pushButton_6.setEnabled(True)
        self.pushButton_12.setEnabled(True)
        self.pushButton_13.setEnabled(True)
        self.pushButton_14.setEnabled(True)
        self.pushButton_15.setEnabled(True)

        self.get_id_hometown()

    def apply_mode_hometown(self):
        self.lineEdit.setEnabled(True)
        self.dateEdit_3.setEnabled(True)
        self.dateEdit_4.setEnabled(True)
        self.pushButton_3.setEnabled(True)

        self.comboBox_3.setEnabled(False)
        self.lineEdit_3.setEnabled(False)
        self.pushButton_6.setEnabled(False)
        self.pushButton_12.setEnabled(False)
        self.pushButton_13.setEnabled(False)
        self.pushButton_14.setEnabled(False)
        self.pushButton_15.setEnabled(False)

        self.default_css_hometown()

    def is_query_hometown(self):
        self.default_css_hometown()
        if self.checkBox_4.isChecked():
            self.query_mode_hometown()
        else:
            self.apply_mode_hometown()
            self.panel_to_default()

    def request_go_home(self):
        if self.home_town==0:
            QMessageBox.warning(self, 'Request Denied', 'Request denied: 0 time remains for traveling to hometown in this year!')
            return

        if self.lineEdit.text() == '':
            QMessageBox.critical(self, 'Error', 'Please input the "Destination"(Your hometown)!')
            return

        start_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        end_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        if int(start_dt.strftime('%Y%m%d')) >= int(end_dt.strftime('%Y%m%d')):
            QMessageBox.critical(self, 'Date Error', 'Date error: end date must be later than start date!')
            return

        id = ID
        name = MainWindow.name
        type = 'Hometown'
        apply_dttm = TimeCard.t
        apply_dt = TimeCard.t.strftime('%d/%m/%Y')
        start_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        start_len = 'all'
        end_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        end_len = 'all'
        single_day = 0
        during = (end_dt - start_dt).days + 1
        remarks = self.lineEdit.text()

        msm = f'Submit this hometown traveling request?\n' \
              f'Leave type: Hometown traveling\n' \
              f'Start from: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Until: {end_dt.strftime("%d/%m/%Y")}\n' \
              f'Duration: {during} days'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = res[0][2]
                current_po = 'MD'

        sql = """INSERT INTO leave_request (USER_ID, USER_NAME, TYPE, APPLY_DTTM, APPLY_DT,START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        self.cursor.execute(sql, (
        id, name, type, apply_dttm, apply_dt, start_dt, start_len, end_dt, end_len, single_day, during, remarks,
        current_to, current_po))
        DB.commit()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='leave')
        # ===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Hometown traveling request has been sent successfully! Please wait for the approvements.')
        self.panel_to_default()

    def del_single_request(self):
        if self.lineEdit_2.text() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.lineEdit_2.text())
        sql = """SELECT SERIAL, TYPE, START_DT, START_LEN, END_DT, END_LEN, DURING, REMARKS, LEADER, DM, MD, HR, APPLY_DTTM FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()

        request_id = res[0][0]
        type = res[0][1]
        start_dt = res[0][2]
        start_len = res[0][3]
        end_dt = res[0][4]
        end_len = res[0][5]
        during = res[0][6]
        remarks = res[0][7]
        leader = res[0][8]
        dm = res[0][9]
        md = res[0][10]
        hr = res[0][11]
        request_date = res[0][12]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(hr)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        end_dt_func = lambda y: None if y == None else y.strftime("%d/%m/%Y")
        msm = f'Are you sure to delete this request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {request_date}\n' \
              f'Leave type: {type}\n' \
              f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Start from: {start_len}\n' \
              f'Until(Date): {end_dt_func(end_dt)}\n' \
              f'Until(time): {end_len}\n' \
              f'Duration: {during} day(s)\n' \
              f'Remarks: {remarks}\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'MD approving: {approve_func(md)}\n' \
              f'HR approving: {approve_func(hr)}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        self.cursor.execute(sql, (serial))
        DB.commit()
        self.lineEdit_2.setText('')
        self.input_id()
        self.show_IDs()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been deleted successfully!')

    def input_id(self):
        txt = self.comboBox_2.currentText()
        #if txt == '':
            #return
        self.lineEdit_2.setText(txt)

    def default_css(self):
        self.dateEdit_5.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit_2.setStyleSheet(self.dateEdite_style_normal)
        self.comboBox_5.setStyleSheet(self.comboBox_style_normal)
        self.comboBox.setStyleSheet(self.comboBox_style_normal)
        self.comboBox_4.setStyleSheet(self.comboBox_style_normal)
        self.textEdit.setStyleSheet(self.textEdit_style_normal)
        self.radioButton.setStyleSheet(self.radio_style_normal)
        self.radioButton_2.setStyleSheet(self.radio_style_normal)
        self.radioButton_3.setStyleSheet(self.radio_style_normal)
        self.radioButton_4.setStyleSheet(self.radio_style_normal)
        self.radioButton_5.setStyleSheet(self.radio_style_normal)
        self.radioButton_6.setStyleSheet(self.radio_style_normal)
        self.radioButton_7.setStyleSheet(self.radio_style_normal)

        self.pushButton_8.setText('―')
        self.pushButton_9.setText('―')
        self.pushButton_10.setText('―')
        self.pushButton_11.setText('―')

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_8.setStyleSheet(btn_yellow_css)
        self.pushButton_9.setStyleSheet(btn_yellow_css)
        self.pushButton_10.setStyleSheet(btn_yellow_css)
        self.pushButton_11.setStyleSheet(btn_yellow_css)

        self.label_21.setText('WAITING...')
        self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def refresh_data(self):
        self.default_css()

        self.cursor_2 = DB.cursor()
        sql = """SELECT TYPE, START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, REMARKS, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_2.execute(sql, (self.comboBox_2.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_2=DB.cursor()
            self.cursor_2.execute(sql, (self.comboBox_2.currentText()))

        res = self.cursor_2.fetchall()

        self.cursor_2.close()

        if res == ():
            return
        type = res[0][0]
        start_date = res[0][1]
        start_len = res[0][2]
        end_date = res[0][3]
        end_len = res[0][4]
        single_day = res[0][5]
        remarks = res[0][6]
        leader = res[0][7]
        dm = res[0][8]
        md = res[0][9]
        hr = res[0][10]

        len_func = lambda x: 'Morning leave' if x == 'morning' else (
            'Afternoon leave' if x == 'afternoon' else 'All day leave')
        if single_day == 1:
            self.checkBox.setChecked(False)

            self.dateEdit_5.setDate(start_date)
            self.dateEdit_5.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox_5.setCurrentText(len_func(start_len))
            self.comboBox_5.setStyleSheet(self.comboBox_style_highlight)

        else:
            self.checkBox.setChecked(True)

            self.dateEdit.setDate(start_date)
            self.dateEdit.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox.setCurrentText(len_func(start_len))
            self.comboBox.setStyleSheet(self.comboBox_style_highlight)

            self.dateEdit_2.setDate(end_date)
            self.dateEdit_2.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox_4.setCurrentText(len_func(end_len))
            self.comboBox_4.setStyleSheet(self.comboBox_style_highlight)

        self.textEdit.setText(remarks)
        self.textEdit.setStyleSheet(self.textEdit_style_highlight)

        if type == 'Annual leave':
            self.radioButton.setChecked(True)
            self.radioButton.setStyleSheet(self.radio_style_highlight)
        elif type == 'Leave without pay':
            self.radioButton_5.setChecked(True)
            self.radioButton_5.setStyleSheet(self.radio_style_highlight)
        elif type == 'Sick leave':
            self.radioButton_3.setChecked(True)
            self.radioButton_3.setStyleSheet(self.radio_style_highlight)
        elif type == 'Compensatory leave':
            self.radioButton_6.setChecked(True)
            self.radioButton_6.setStyleSheet(self.radio_style_highlight)
        elif type == 'Maternity leave':
            self.radioButton_2.setChecked(True)
            self.radioButton_2.setStyleSheet(self.radio_style_highlight)
        elif type == 'Personal leave':
            self.radioButton_7.setChecked(True)
            self.radioButton_7.setStyleSheet(self.radio_style_highlight)
        else:
            self.radioButton_4.setChecked(True)
            self.radioButton_4.setStyleSheet(self.radio_style_highlight)

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"
        if leader == None:
            self.pushButton_8.setText('―')
            self.pushButton_8.setStyleSheet(btn_yellow_css)

        elif leader == 1:
            self.pushButton_8.setText('〇')
            self.pushButton_8.setStyleSheet(btn_green_css)
        else:
            self.pushButton_8.setText('×')
            self.pushButton_8.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_9.setText('―')
            self.pushButton_9.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_9.setText('〇')
            self.pushButton_9.setStyleSheet(btn_green_css)
        else:
            self.pushButton_9.setText('×')
            self.pushButton_9.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_10.setText('―')
            self.pushButton_10.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_10.setText('〇')
            self.pushButton_10.setStyleSheet(btn_green_css)
        else:
            self.pushButton_10.setText('×')
            self.pushButton_10.setStyleSheet(btn_red_css)

        if hr == None:
            self.pushButton_11.setText('―')
            self.pushButton_11.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_11.setText('〇')
            self.pushButton_11.setStyleSheet(btn_green_css)
        else:
            self.pushButton_11.setText('×')
            self.pushButton_11.setStyleSheet(btn_red_css)

        if self.pushButton_8.text() == '×' or self.pushButton_9.text() == '×' or self.pushButton_10.text() == '×' or self.pushButton_11.text() == '×':
            self.label_21.setText('DECLINED')
            self.label_21.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_11.text() == '〇':
            self.label_21.setText('OK')
            self.label_21.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_21.setText('WAITING...')
            self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def show_IDs(self):
        self.comboBox_2.clear()
        sql = """SELECT SERIAL FROM leave_request WHERE USER_ID=%s AND APPLY_DT=%s AND TYPE!=%s"""
        try:
            self.cursor.execute(sql, (ID, self.dateEdit_7.text(), 'hometown'))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID, self.dateEdit_7.text(), 'hometown'))

        res = self.cursor.fetchall()
        if res == ():
            return
        id_lst = []
        for each in res:
            id_lst.append(str(each[0]))
        self.comboBox_2.addItems(id_lst)

    def request_per_leave(self):
        if self.checkBox.isChecked():
            start_dt = datetime.datetime.strptime(self.dateEdit.text(), '%d/%m/%Y')
            if self.comboBox.currentText() == 'All day leave':
                start_len = 'all'
            else:
                start_len = 'afternoon'

            end_dt = datetime.datetime.strptime(self.dateEdit_2.text(), '%d/%m/%Y')
            if self.comboBox_4.currentText() == 'All day leave':
                end_len = 'all'
            else:
                end_len = 'morning'

            if int(start_dt.strftime('%Y%m%d')) >= int(end_dt.strftime('%Y%m%d')):
                QMessageBox.critical(self, 'Date Error', 'Date error: end date must be later than start date!')
                return

            single = 0
            during = (end_dt - start_dt).days + 1
            if start_len == 'afternoon':
                during -= 0.5
            if end_len == 'morning':
                during -= 0.5

        else:
            start_dt = datetime.datetime.strptime(self.dateEdit_5.text(), '%d/%m/%Y')
            # print(self.comboBox_5.currentText())
            if self.comboBox_5.currentText() == 'All day leave':
                start_len = 'all'
            elif self.comboBox_5.currentText() == 'Morning leave':
                start_len = 'morning'
            else:
                start_len = 'afternoon'

            single = 1
            if start_len == 'all':
                during = 1
            else:
                during = 0.5

            end_dt = None
            end_len = None
            # print(during)

        remarks = self.textEdit.toPlainText()
        id = ID
        name = MainWindow.name

        if self.radioButton.isChecked():
            if float(self.annual_remain) < during:
                QMessageBox.critical(self, 'Warning',
                                     "Amount of remaining annual leave days is not enough, please change the type of leave that you are requesting.")
                return
        if self.radioButton_3.isChecked():
            if float(self.sick_remain) < during:
                QMessageBox.critical(self, 'Warning',
                                     "Amount of remaining sick leave days is not enough, please change the type of leave that you are requesting.")
                return

        if self.radioButton.isChecked():
            type = 'Annual leave'
        elif self.radioButton_5.isChecked():
            type = 'Leave without pay'
        elif self.radioButton_3.isChecked():
            type = 'Sick leave'
        elif self.radioButton_6.isChecked():
            type = 'Compensatory leave'
        elif self.radioButton_2.isChecked():
            type = 'Maternity leave'
        elif self.radioButton_7.isChecked():
            type = 'Personal leave'
        else:
            type = 'Sterilisation leave'

        if single == 1:
            msm = f'Submit this leave request?\n' \
                  f'Leave type: {type}\n' \
                  f'Leave date: {start_dt.strftime("%d/%m/%Y")}\n' \
                  f'Leave time: {start_len}\n' \
                  f'Duration: {during} day\n' \
                  f'Remarks: {remarks}'
        else:
            msm = f'Submit this leave request?\n' \
                  f'Leave type: {type}\n' \
                  f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
                  f'Start from: {start_len}\n' \
                  f'Until(Date): {end_dt.strftime("%d/%m/%Y")}\n' \
                  f'Until(time): {end_len}\n' \
                  f'Duration: {during} day(s)\n' \
                  f'Remarks: {remarks}'

        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = res[0][2]
                current_po = 'MD'

        sql = """INSERT INTO leave_request (USER_ID, USER_NAME, TYPE, APPLY_DTTM, APPLY_DT,START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        self.cursor.execute(sql, (
        id, name, type, TimeCard.t, TimeCard.t.strftime('%d/%m/%Y'), start_dt, start_len, end_dt, end_len, single,
        during, remarks, current_to, current_po))
        DB.commit()

        #Mail sending==============================================
        info_lst=query_email(id=current_to)
        if info_lst==-1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='leave')
        #===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been sent successfully! Please wait for the approvements.')

        self.panel_to_default()

    def query_mode(self):
        self.dateEdit_5.setEnabled(False)
        self.checkBox.setEnabled(False)
        self.dateEdit.setEnabled(False)
        self.dateEdit_2.setEnabled(False)
        self.comboBox_5.setEnabled(False)
        self.comboBox.setEnabled(False)
        self.comboBox_4.setEnabled(False)
        self.textEdit.setEnabled(False)
        self.radioButton.setEnabled(False)
        self.radioButton_2.setEnabled(False)
        self.radioButton_3.setEnabled(False)
        self.radioButton_4.setEnabled(False)
        self.radioButton_5.setEnabled(False)
        self.radioButton_6.setEnabled(False)
        self.pushButton.setEnabled(False)

        self.comboBox_2.setEnabled(True)
        self.dateEdit_7.setEnabled(True)
        self.pushButton_5.setEnabled(True)

        self.lineEdit_2.setText('')
        self.input_id()
        self.show_IDs()

    def apply_mode(self):
        self.dateEdit_5.setEnabled(True)
        self.checkBox.setEnabled(True)
        self.dateEdit.setEnabled(True)
        self.dateEdit_2.setEnabled(True)
        self.comboBox_5.setEnabled(True)
        self.comboBox.setEnabled(True)
        self.comboBox_4.setEnabled(True)
        self.textEdit.setEnabled(True)
        self.radioButton.setEnabled(True)
        self.radioButton_2.setEnabled(True)
        self.radioButton_3.setEnabled(True)
        self.radioButton_4.setEnabled(True)
        self.radioButton_5.setEnabled(True)
        self.radioButton_6.setEnabled(True)
        self.pushButton.setEnabled(True)

        self.comboBox_2.setEnabled(False)
        self.dateEdit_7.setEnabled(False)
        self.pushButton_5.setEnabled(False)

        self.is_multiple()

    def is_query(self):
        self.default_css()
        if self.checkBox_2.isChecked():
            self.query_mode()
        else:
            self.apply_mode()
            self.panel_to_default()
            self.default_css()

    def is_multiple(self):
        if self.checkBox.isChecked():
            self.dateEdit.setEnabled(True)
            self.dateEdit_2.setEnabled(True)
            self.comboBox.setEnabled(True)
            self.comboBox_4.setEnabled(True)

            self.dateEdit_5.setEnabled(False)
            self.comboBox_5.setEnabled(False)
        else:
            self.dateEdit.setEnabled(False)
            self.dateEdit_2.setEnabled(False)
            self.comboBox.setEnabled(False)
            self.comboBox_4.setEnabled(False)

            self.dateEdit_5.setEnabled(True)
            self.comboBox_5.setEnabled(True)

    def panel_to_default(self):
        self.dateEdit_5.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_2.setDate(QDate.currentDate())
        self.dateEdit_7.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit_6.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_6.setCurrentText(month)
        self.textEdit.setText('')
        self.lineEdit.setText('')
        self.radioButton.setChecked(True)
        self.checkBox_2.setChecked(False)
        self.apply_mode()
        self.checkBox_4.setChecked(False)
        self.apply_mode_hometown()


    def quit(self):
        AskForLeave.close()

    def closeEvent(self, event):
        try:
            self.cursor.close()
        except:
            pass
        MainWindow.show()


class OTApplication(QMainWindow, Ui_OTApplication):
    def __init__(self):
        super(OTApplication, self).__init__()
        self.setupUi(self)

        self.label_31.setText('')
        self.timeEdit.setTime(QTime.fromString('17:15', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('18:15', 'hh:mm'))
        self.initHours()
        self.textEdit_5.setReadOnly(True)
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.geoTab = self.tabWidget.geometry()
        self.geoTable = self.tableWidget.geometry()
        self.geoBTN6 = self.pushButton_6.geometry()
        self.geoBTN2 = self.pushButton_2.geometry()
        self.geoBTN4 = self.pushButton_4.geometry()
        self.geoFrame2 = self.frame_2.geometry()
        self.geoBTN16 = self.pushButton_16.geometry()

        self.pushButton.clicked.connect(self.submit_request)
        self.timeEdit.timeChanged.connect(self.initHours)
        self.timeEdit_2.timeChanged.connect(self.initHours)
        self.checkBox.clicked.connect(self.is_query)
        self.pushButton_4.clicked.connect(self.quit)
        self.dateEdit_4.dateChanged.connect(self.show_key)
        self.comboBox_2.currentTextChanged.connect(self.refresh_panel)
        self.pushButton_5.clicked.connect(self.cancel_apply)
        self.pushButton_16.clicked.connect(self.download_data)
        self.pushButton_6.clicked.connect(self.cancel_request_on_table)
        self.pushButton_2.clicked.connect(self.export_excel)

        self.dateEdit.dateTimeChanged.connect(self.download_data)
        self.comboBox_3.currentTextChanged.connect(self.download_data)
        self.pushButton_16.setVisible(False)

        self.tabWidget.currentChanged.connect(self.tableview)

    def export_excel(self):
        try:
            first_cell = self.tableWidget.item(0, 0).text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Table', 'Warning: The table is empty, no need to export the excel file!')
            return

        yearmonth = self.dateEdit.text() + self.comboBox_3.currentText()
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s OT Application({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return
        excel_path = a[0]

        data_frame = []
        for i in range(self.tableWidget.rowCount()):
            data_line = []
            for j in range(self.tableWidget.columnCount()):
                data_line.append(self.tableWidget.item(i, j).text())
            data_frame.append(data_line)

        wb = xl.Workbook()
        ws = wb.active
        ws.append(['Request ID', 'Staff ID', 'Staff Name', 'Request Date-Time', 'Request Date',
                   'Date of OT', 'Start Time', 'End Time', 'OT Hours', 'Leader', 'DM', 'MD', 'HR','Remarks'])

        for line in data_frame:
            ws.append(line)

        try:
            wb.save(filename=excel_path)
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been exported successfully!')

    def tableview(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        if self.tabWidget.currentIndex() == 1:
            width0 = W_OT
            height0 = H_OT
            self.desktop = QApplication.desktop()
            screen_count = self.desktop.screenCount()  #2 screen version updating
            geometry = self.desktop.geometry()
            if screen_count == 1:                      #2 screen version updating
                width1 = geometry.width()              #2 screen version updating
            else:                                      #2 screen version updating
                width1 = geometry.width()/2            #2 screen version updating
            height1 = geometry.height()
            x = width1 - width0
            y = height1 - height0 - 100

            self.setMaximumSize(16777215, 16777215)
            self.showMaximized()

            move_widgets_for_frame(self.tabWidget, x, y)
            move_widgets_for_frame(self.tableWidget, x, y)
            move_widgets_only_vertical(self.pushButton_6, x, y)
            move_widgets_only_vertical(self.pushButton_2, x, y)
            move_widgets_only_horizontal(self.pushButton_16, x, y)
            move_widgets_both(self.pushButton_4, x, y)
            extend_widgets_horizontal(self.frame_2, x, y)
        else:
            self.tabWidget.setGeometry(self.geoTab.x(), self.geoTab.y(), self.geoTab.width(), self.geoTab.height())
            self.tableWidget.setGeometry(self.geoTable.x(), self.geoTable.y(), self.geoTable.width(),
                                         self.geoTable.height())
            self.pushButton_6.setGeometry(self.geoBTN6.x(), self.geoBTN6.y(), self.geoBTN6.width(),
                                          self.geoBTN6.height())
            self.pushButton_2.setGeometry(self.geoBTN2.x(), self.geoBTN2.y(), self.geoBTN2.width(),
                                          self.geoBTN2.height())
            self.pushButton_4.setGeometry(self.geoBTN4)
            self.pushButton_16.setGeometry(self.geoBTN16.x(), self.geoBTN16.y(), self.geoBTN16.width(),
                                           self.geoBTN16.height())
            self.frame_2.setGeometry(self.geoFrame2)
            self.showNormal()
            self.setMaximumSize(W_OT, H_OT)

    def hours_css_control(self, mode):
        if mode == 0:
            self.textEdit_5.setStyleSheet(
                'QTextEdit{ font-family:Microsoft Yahei; color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 251, 201);}')
            self.label_31.setText('')
        else:
            self.textEdit_5.setStyleSheet(
                'QTextEdit{  font-family:Microsoft Yahei;	color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 120, 120);}')
            self.label_31.setText('Please input the correct time range!')

    def initHours(self):
        seconds = self.timeEdit.time().secsTo(self.timeEdit_2.time())
        hours = seconds / 3600
        hours = round(hours, 2)
        self.textEdit_5.setText(str(hours))
        if hours < 0:
            self.hours_css_control(mode=1)
        else:
            self.hours_css_control(mode=0)

    def initializing(self):
        self.apply_mode()

        self.dateEdit_css_highlighted = 'QDateEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'
        self.timeEdit_css_highlighted = 'QTimeEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'
        self.textEdit_css_highlighted = 'QTextEdit{    font-family:Microsoft Yahei;   color:rgb(50, 50, 50);   font-size:12pt;   background-color:rgb(255, 163, 144);}'
        self.remarks_css_highlighted = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'

        self.dateEdit_css_normal = 'QDateEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255,255);}QDateEdit:hover{   background-color:rgb(229, 242, 255);}QDateEdit:focus{   background-color:rgb(229, 242, 255);}'
        self.timeEdit_css_normal = 'QTimeEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTimeEdit:hover{   background-color:rgb(229, 242, 255);}QTimeEdit:focus{   background-color:rgb(229, 242, 255);}'
        self.textEdit_css_normal = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 251, 201);}'
        self.remarks_css_normal = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTextEdit:hover{   background-color:rgb(229, 242, 255);}QTextEdit:focus{   background-color:rgb(229, 242, 255);}'

        self.download_data()

    def cancel_request_on_table(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 12).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.show_key()

        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.download_data()
        self.cursor.close()

    def download_data(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM ot_request WHERE OT_DT>=%s AND OT_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            #QMessageBox.information(self, 'Empty Record', 'No record in the selected time range.')
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            for j in range(len(res[0])):
                value=res[i][j]
                if j in [9,10,11,12]:
                    if res[i][j]==None:
                        value='Unconfirmed'
                    elif res[i][j]==1:
                        value='OK'
                    else:
                        value='Declined'

                self.tableWidget.setItem(i, j, QTableWidgetItem(str(value)))

        for i in [0, 3, 9, 10, 11, 12, 13]:
            # if i==13:
            #   continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()

    def cancel_apply(self):
        if self.comboBox_2.currentText() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.comboBox_2.currentText())
        self.cursor = DB.cursor()

        sql = """SELECT SERIAL, OT_DT, START_TM, END_TM, DURING, REMARKS, LEADER, DM, MD, HR, APPLY_DTTM FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()

        request_id = res[0][0]
        ot_dt = res[0][1]
        start_tm = res[0][2]
        end_tm = res[0][3]
        during = res[0][4]
        remarks = res[0][5]
        leader = res[0][6]
        dm = res[0][7]
        md = res[0][8]
        hr = res[0][9]
        request_date = res[0][10]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(hr)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        msm = f'Are you sure to delete this OT request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {request_date}\n' \
              f'OT date: {ot_dt}\n' \
              f'Start time: {start_tm}\n' \
              f'End time: {end_tm}\n' \
              f'Duration: {during} hour(s)\n' \
              f'Remarks: {remarks}\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'MD approving: {approve_func(md)}\n' \
              f'HR approving: {approve_func(hr)}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            self.cursor.close()
            return

        sql = """DELETE FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))
            DB.commit()

        self.show_key()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Selected OT request has been deleted successfully!')

    def highlight_on(self):
        self.dateEdit_3.setStyleSheet(self.dateEdit_css_highlighted)
        self.timeEdit.setStyleSheet(self.timeEdit_css_highlighted)
        self.timeEdit_2.setStyleSheet(self.timeEdit_css_highlighted)
        self.textEdit_5.setStyleSheet(self.textEdit_css_highlighted)
        self.textEdit.setStyleSheet(self.remarks_css_highlighted)

    def highlight_off(self):
        self.dateEdit_3.setStyleSheet(self.dateEdit_css_normal)
        self.timeEdit.setStyleSheet(self.timeEdit_css_normal)
        self.timeEdit_2.setStyleSheet(self.timeEdit_css_normal)
        self.textEdit_5.setStyleSheet(self.textEdit_css_normal)
        self.textEdit.setStyleSheet(self.remarks_css_normal)

        self.pushButton_8.setText('―')
        self.pushButton_9.setText('―')
        self.pushButton_10.setText('―')
        self.pushButton_11.setText('―')
        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_8.setStyleSheet(btn_yellow_css)
        self.pushButton_9.setStyleSheet(btn_yellow_css)
        self.pushButton_10.setStyleSheet(btn_yellow_css)
        self.pushButton_11.setStyleSheet(btn_yellow_css)

        self.label_21.setText('WAITING...')
        self.label_21.setStyleSheet('QLabel{color:rgb(255, 170, 0);}')

    def refresh_panel(self):
        if self.comboBox_2.currentText() == '':
            self.highlight_off()
            return
        self.cursor = DB.cursor()
        sql = """SELECT OT_DT, START_TM, END_TM, DURING, REMARKS,LEADER, DM, MD, HR FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (self.comboBox_2.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (self.comboBox_2.currentText()))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return
        ot_dt = res[0][0]
        start_tm = res[0][1]
        end_tm = res[0][2]
        during = res[0][3]
        remarks = res[0][4]
        leader = res[0][5]
        dm = res[0][6]
        md = res[0][7]
        hr = res[0][8]

        self.dateEdit_3.setDate(ot_dt)
        self.timeEdit.setDateTime(datetime.datetime.strptime(str(start_tm), '%H:%M:%S'))
        self.timeEdit_2.setDateTime(datetime.datetime.strptime(str(end_tm), '%H:%M:%S'))
        self.textEdit_5.setText(str(during))
        self.textEdit.setText(remarks)
        self.highlight_on()
        self.cursor.close()

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"
        if leader == None:
            self.pushButton_8.setText('―')
            self.pushButton_8.setStyleSheet(btn_yellow_css)

        elif leader == 1:
            self.pushButton_8.setText('〇')
            self.pushButton_8.setStyleSheet(btn_green_css)
        else:
            self.pushButton_8.setText('×')
            self.pushButton_8.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_9.setText('―')
            self.pushButton_9.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_9.setText('〇')
            self.pushButton_9.setStyleSheet(btn_green_css)
        else:
            self.pushButton_9.setText('×')
            self.pushButton_9.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_10.setText('―')
            self.pushButton_10.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_10.setText('〇')
            self.pushButton_10.setStyleSheet(btn_green_css)
        else:
            self.pushButton_10.setText('×')
            self.pushButton_10.setStyleSheet(btn_red_css)

        if hr == None:
            self.pushButton_11.setText('―')
            self.pushButton_11.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_11.setText('〇')
            self.pushButton_11.setStyleSheet(btn_green_css)
        else:
            self.pushButton_11.setText('×')
            self.pushButton_11.setStyleSheet(btn_red_css)

        if self.pushButton_8.text() == '×' or self.pushButton_9.text() == '×' or self.pushButton_10.text() == '×' or self.pushButton_11.text() == '×':
            self.label_21.setText('DECLINED')
            self.label_21.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_11.text() == '〇':
            self.label_21.setText('OK')
            self.label_21.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_21.setText('WAITING...')
            self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def show_key(self):
        self.comboBox_2.clear()
        ot_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        self.cursor2 = DB.cursor()
        sql = """SELECT SERIAL FROM ot_request WHERE OT_DT=%s AND USER_ID=%s"""
        try:
            self.cursor2.execute(sql, (ot_dt, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor2=DB.cursor()
            self.cursor2.execute(sql, (ot_dt, ID))

        res = self.cursor2.fetchall()
        if res == ():
            self.cursor2.close()
            return
        keys = []
        for each in res:
            keys.append(str(each[0]))
        self.comboBox_2.addItems(keys)
        self.cursor2.close()

    def query_mode(self):
        self.dateEdit_3.setEnabled(False)
        self.timeEdit.setEnabled(False)
        self.timeEdit_2.setEnabled(False)
        self.textEdit_5.setEnabled(False)
        self.pushButton.setEnabled(False)
        self.textEdit.setEnabled(False)

        self.comboBox_2.setEnabled(True)
        self.pushButton_5.setEnabled(True)
        self.dateEdit_4.setEnabled(True)

        self.show_key()

    def apply_mode(self):
        self.dateEdit_3.setEnabled(True)
        self.timeEdit.setEnabled(True)
        self.timeEdit_2.setEnabled(True)
        self.textEdit_5.setEnabled(True)
        self.pushButton.setEnabled(True)
        self.textEdit.setEnabled(True)

        self.comboBox_2.setEnabled(False)
        self.pushButton_5.setEnabled(False)
        self.dateEdit_4.setEnabled(False)

    def is_query(self):
        if self.checkBox.isChecked():
            self.query_mode()
        else:
            self.apply_mode()
            self.to_default_panel()
            self.highlight_off()

    def submit_request(self):
        user_id = ID
        user_name = MainWindow.name
        apply_dttm = TimeCard.t
        apply_dt = TimeCard.t.strftime('%d/%m/%Y')
        ot_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        start_tm = datetime.datetime.strptime(self.timeEdit.text(), '%H:%M')
        end_tm = datetime.datetime.strptime(self.timeEdit_2.text(), '%H:%M')
        during = float(self.textEdit_5.toPlainText())
        remarks = self.textEdit.toPlainText()

        self.cursor_confirm=DB.cursor()
        sql="""SELECT * FROM ot_request WHERE USER_ID=%s AND OT_DT=%s"""
        try:
            self.cursor_confirm.execute(sql, (ID, ot_dt))
        except:
            reconnect_DB(self)
            self.cursor_confirm=DB.cursor()
            self.cursor_confirm.execute(sql, (ID, ot_dt))
        if_exist=self.cursor_confirm.fetchall()
        self.cursor_confirm.close()
        if if_exist!=():
            QMessageBox.critical(self, 'Request Denied', f'Request denied: You have already submitted the request for the date of {self.dateEdit_3.text()}! If you want to request again, please cancel the old request first.')
            return

        msm = f'Submit this OT application?\n' \
              f'OT date: {ot_dt}\n' \
              f'Start from: {start_tm.strftime("%H:%M")}\n' \
              f'Until: {end_tm.strftime("%H:%M")}\n' \
              f'Duration: {during} hour(s)\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = res[0][2]
                current_po = 'MD'

        sql = """INSERT INTO ot_request (USER_ID, USER_NAME, APPLY_DTTM, APPLY_DT, OT_DT, START_TM, END_TM, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        self.cursor.execute(sql, (
        user_id, user_name, apply_dttm, apply_dt, ot_dt, start_tm, end_tm, during, remarks, current_to, current_po))
        DB.commit()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='ot')
        # ===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'OT application has been submitted! Please wait for the approvements.')
        self.cursor.close()
        self.to_default_panel()

    def to_default_panel(self):
        self.label_31.setText('')
        self.timeEdit.setTime(QTime.fromString('17:15', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('18:15', 'hh:mm'))
        self.initHours()
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)
        self.textEdit.setText('')

    def quit(self):
        OTApplication.close()

    def closeEvent(self, event):
        MainWindow.show()


class BookMeetingRoom(QMainWindow, Ui_BookMeetingRoom):
    def __init__(self):
        super(BookMeetingRoom, self).__init__()
        self.setupUi(self)

        self.tableWidget.setRowCount(20)
        self.label_31.setText(' ')
        self.timeEdit.setTime(QTime.fromString('09:00', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('09:30', 'hh:mm'))

        self.timeEdit.timeChanged.connect(self.is_valid_time)
        self.timeEdit_2.timeChanged.connect(self.is_valid_time)
        self.pushButton_5.clicked.connect(self.quit)
        self.pushButton_3.clicked.connect(self.submit_request)
        self.calendarWidget.clicked.connect(self.show_on_table)
        self.pushButton_4.clicked.connect(self.cancel_booking)

    def initializing(self):
        self.show_on_table()

    def cancel_booking(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Warning', 'Please select the booking record that you want to cancel.')
            return

        try:
            booking_id = self.tableWidget.item(index, 0).text()
        except:
            QMessageBox.critical(self, 'Empty Selection',
                                 'Empty selection! Please select the booking record that you want to cancel.')
            return

        self.cursor = DB.cursor()

        SQL = """SELECT USER_ID FROM book_meeting_room WHERE SERIAL=%s"""
        try:
            self.cursor.execute(SQL, (booking_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(SQL, (booking_id))

        res = self.cursor.fetchall()
        user_id = res[0][0]
        if user_id != ID:
            QMessageBox.critical(self, 'Access Denied',
                                 'Warning: You can only cancel the bookings submitted by yourself.')
            self.cursor.close()
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected booking?\n'
                                                       f'Booking ID: {booking_id}')
        if a == QMessageBox.No:
            self.cursor.close()
            return

        sql = """DELETE FROM book_meeting_room WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (booking_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (booking_id))
            DB.commit()
        self.cursor.close()

        self.show_on_table()
        QMessageBox.information(self, 'Info', f'Request ID: {booking_id}, has been canceled successfully!')

    def submit_request(self):
        if self.label_31.text() == 'Please input the correct time range!':
            QMessageBox.critical(self, 'Time Error', 'Please input the correct time range first!')
            return

        apply_dttm = TimeCard.t
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')
        start_time = datetime.datetime.strptime(self.timeEdit.time().toString('hh:mm'), '%H:%M')
        end_time = datetime.datetime.strptime(self.timeEdit_2.time().toString('hh:mm'), '%H:%M')
        user_id = ID
        applier = MainWindow.name
        division = self.get_userDIV()
        contents = self.textEdit.toPlainText()

        judge = self.judge_validation(start_tm=start_time, end_tm=end_time)
        if judge == False:
            QMessageBox.warning(self, 'Warning',
                                'The time range you selected is occupied by another meeting! Please change the time range and submit the booking again.')
            return

        msm = f'Are you sure to submit the booking?\n' \
              f'Meeting Date: {meeting_dt.strftime("%d/%m/%Y")}\n' \
              f'Start Time: {start_time.strftime("%H:%M")}\n' \
              f'Finish Time: {end_time.strftime("%H:%M")}\n' \
              f'Meeting Contents: {contents}\n'

        a = QMessageBox.question(self, 'Query', msm)
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """INSERT INTO book_meeting_room (APPLY_DTTM, MEETING_DT, START_TM, END_TM, USER_ID, USER_NAME, DIVISION, CONTENTS) VALUES (%s, %s,%s,%s,%s,%s,%s,%s)"""
        try:
            self.cursor.execute(sql, (apply_dttm, meeting_dt, start_time, end_time, user_id, applier, division, contents))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql,
                                (apply_dttm, meeting_dt, start_time, end_time, user_id, applier, division, contents))
            DB.commit()

        QMessageBox.information(self, 'Info', 'Meeting room booking has been completed successfully!')
        self.cursor.close()
        self.show_on_table()

    def judge_validation(self, start_tm, end_tm):
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')
        self.cursor_judge = DB.cursor()
        sql = """SELECT START_TM, END_TM FROM book_meeting_room WHERE MEETING_DT=%s"""
        try:
            self.cursor_judge.execute(sql, (meeting_dt))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_judge=DB.cursor()
            self.cursor_judge.execute(sql, (meeting_dt))

        res = self.cursor_judge.fetchall()
        self.cursor_judge.close()
        if res == ():
            return True
        for each in res:
            if not (end_tm < datetime.datetime.strptime(str(each[0]),
                                                        '%H:%M:%S') or start_tm > datetime.datetime.strptime(
                    str(each[1]), '%H:%M:%S')):
                return False
        return True

    def show_on_table(self):
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        #print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM book_meeting_room WHERE MEETING_DT=%s"""
        try:
            self.cursor.execute(sql, (meeting_dt))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (meeting_dt))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res) + 20)
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 3, 4, 6, 7, 8]:
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    # self.tableWidget.sortItems(2, QtCore.Qt.AscendingOrder)

    def get_userDIV(self):
        self.cursor_staff = DB.cursor()
        sql = """SELECT DIVISION FROM akt_staff WHERE ID=%s"""
        try:
            self.cursor_staff.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_staff=DB.cursor()
            self.cursor_staff.execute(sql, (ID))

        res = self.cursor_staff.fetchall()
        self.cursor_staff.close()
        return res[0][0]

    def is_valid_time(self):
        if self.timeEdit_2.time() <= self.timeEdit.time():
            self.label_31.setText('Please input the correct time range!')
        else:
            self.label_31.setText(' ')

    def quit(self):
        BookMeetingRoom.close()

    def closeEvent(self, event):
        MainWindow.show()


class PassWindow(QWidget, Ui_PassWindow):

    def __init__(self):
        super(PassWindow, self).__init__()
        self.setupUi(self)

        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)

        self.pushButton_2.clicked.connect(self.change_psswd)
        self.pushButton_3.clicked.connect(self.quit)

    def initializing(self):
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_4.setText('')

    def change_psswd(self):
        for each_char in self.lineEdit_3.text():
            if ord(each_char) not in range(32, 127):
                QMessageBox.warning(self, 'Warning', 'Please only use A-Z, a-z, 0-9 and keyboard symbols(such as "!","$","_"... and so on...) for your new password!')
                return
        for each_char in self.lineEdit_4.text():
            if ord(each_char) not in range(32, 127):
                QMessageBox.warning(self, 'Warning',
                                    'Please only use A-Z, a-z, 0-9 or keyboard symbols(such as "!","$","_"... and so on...) for your new password!')
                return

        if self.lineEdit_2.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your old password!')
            return
        if self.lineEdit_3.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your new password!')
            return
        if self.lineEdit_4.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your new password again!')
            return
        if self.lineEdit_3.text() != self.lineEdit_4.text():
            QMessageBox.warning(self, 'Warning',
                                'The new passwords that you input are not the same, please check it again!')
            return

        self.cursor = DB.cursor()
        sql = """SELECT PASSWORD FROM login_pass WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        pw_old = self.cursor.fetchall()[0][0]
        if self.lineEdit_2.text() != pw_old:
            QMessageBox.warning(self, 'Warning', 'The old password is not correct!')
            self.cursor.close()
            return

        sql = """UPDATE login_pass SET PASSWORD=%s WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (self.lineEdit_3.text(), ID))
            DB.commit()
            QMessageBox.information(self, 'Info', 'Your password has been changed successfully!')
            self.quit()
        except:
            DB.rollback()
            QMessageBox.critical(self, 'Warning', 'Operation failed! Please check the network.')
        self.cursor.close()

    def quit(self):
        PassWindow.close()

    def closeEvent(self, event):
        MainWindow.setEnabled(True)


class ApprovePanel(QMainWindow, Ui_ApprovePanel):
    def __init__(self):
        super(ApprovePanel, self).__init__()
        self.setupUi(self)
        self.binding_btn()

    def binding_btn(self):
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.dateEdit.setEnabled(False)
        self.comboBox_3.setEnabled(False)

        self.pushButton_8.clicked.connect(self.quit)
        self.tableWidget_3.clicked.connect(self.show_leave_panel)
        self.tableWidget_2.clicked.connect(self.show_ot_panel)
        self.tableWidget_4.clicked.connect(self.show_late_panel)
        self.tableWidget_8.clicked.connect(self.show_forget_panel)

        self.pushButton_2.clicked.connect(self.leave_accepted)
        self.pushButton_5.clicked.connect(self.leave_declined)
        #self.pushButton_5.clicked.connect(self.show_leave_history)
        self.pushButton.clicked.connect(self.ot_accepted)
        self.pushButton_4.clicked.connect(self.ot_declined)
        self.pushButton_3.clicked.connect(self.late_accepted)
        self.pushButton_6.clicked.connect(self.late_declined)
        self.pushButton_15.clicked.connect(self.forget_accepted)
        self.pushButton_14.clicked.connect(self.forget_declined)

        self.checkBox.stateChanged.connect(self.mode_switch)

        self.dateEdit.dateChanged.connect(self.show_leave_history)
        self.dateEdit.dateChanged.connect(self.show_ot_history)
        self.dateEdit.dateChanged.connect(self.show_late_history)
        self.dateEdit.dateChanged.connect(self.show_forget_history)

        self.comboBox_3.currentTextChanged.connect(self.show_leave_history)
        self.comboBox_3.currentTextChanged.connect(self.show_ot_history)
        self.comboBox_3.currentTextChanged.connect(self.show_late_history)
        self.comboBox_3.currentTextChanged.connect(self.show_forget_history)

    def initializing(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        width0 = W_APPROVE
        height0 = H_APPROVE
        self.desktop = QApplication.desktop()
        screen_count = self.desktop.screenCount()  #2 screen version updating
        geometry = self.desktop.geometry()
        if screen_count == 1:                      #2 screen version updating
            width1 = geometry.width()              #2 screen version updating
        else:                                      #2 screen version updating
            width1 = geometry.width()/2            #2 screen version updating
        height1 = geometry.height()
        x = width1 - width0
        y = height1 - height0 - 100

        self.setMaximumSize(16777215, 16777215)
        self.setMinimumSize(width1, height1)
        self.showMaximized()

        move_widgets_for_frame(self.tabWidget, x, y)

        move_widgets_for_frame(self.tableWidget_2, x, y)
        move_widgets_for_frame(self.tableWidget_3, x, y)
        move_widgets_for_frame(self.tableWidget_4, x, y)
        move_widgets_for_frame(self.tableWidget_8, x, y)

        move_widgets_only_vertical(self.pushButton, x, y)
        move_widgets_only_vertical(self.pushButton_2, x, y)
        move_widgets_only_vertical(self.pushButton_3, x, y)
        move_widgets_only_vertical(self.pushButton_15, x, y)

        move_widgets_both(self.pushButton_4, x, y)
        move_widgets_both(self.pushButton_5, x, y)
        move_widgets_both(self.pushButton_6, x, y)
        move_widgets_both(self.pushButton_14, x, y)

        move_widgets_only_horizontal(self.pushButton_8, x, y)
        extend_widgets_horizontal(self.widget, x, y)

        extend_widgets_horizontal(self.pushButton, 50, y)
        extend_widgets_horizontal(self.pushButton_2, 50, y)
        extend_widgets_horizontal(self.pushButton_4, 50, y)
        extend_widgets_horizontal(self.pushButton_5, 50, y)
        extend_widgets_horizontal(self.pushButton_3, 50, y)
        extend_widgets_horizontal(self.pushButton_15, 50, y)
        extend_widgets_horizontal(self.pushButton_6, 50, y)
        extend_widgets_horizontal(self.pushButton_14, 50, y)

        extend_widgets_vertical(self.textEdit_2, x, y)
        extend_widgets_vertical(self.textEdit, x, y)
        extend_widgets_vertical(self.textEdit_3, x, y)
        extend_widgets_vertical(self.textEdit_40, x, y)

        self.show_leave_contents()
        self.show_ot_contents()
        self.show_late_contents()
        self.show_forget_contents()

    def mode_switch(self):
        if self.checkBox.isChecked():
            self.show_leave_history()
            self.show_ot_history()
            self.show_late_history()
            self.show_forget_history()

            self.dateEdit.setEnabled(True)
            self.comboBox_3.setEnabled(True)

            self.pushButton_2.setEnabled(False)
            self.pushButton_5.setEnabled(False)
            self.pushButton.setEnabled(False)
            self.pushButton_4.setEnabled(False)
            self.pushButton_3.setEnabled(False)
            self.pushButton_6.setEnabled(False)
            self.pushButton_15.setEnabled(False)
            self.pushButton_14.setEnabled(False)
        else:
            self.show_leave_contents()
            self.show_ot_contents()
            self.show_late_contents()
            self.show_forget_contents()

            self.dateEdit.setEnabled(False)
            self.comboBox_3.setEnabled(False)

            self.pushButton_2.setEnabled(True)
            self.pushButton_5.setEnabled(True)
            self.pushButton.setEnabled(True)
            self.pushButton_4.setEnabled(True)
            self.pushButton_3.setEnabled(True)
            self.pushButton_6.setEnabled(True)
            self.pushButton_15.setEnabled(True)
            self.pushButton_14.setEnabled(True)

    def show_leave_contents(self):
        self.tableWidget_3.clearContents()
        self.tableWidget_3.setRowCount(0)
        self.tableWidget_3.setColumnCount(11)
        self.cursor_filling = DB.cursor()
        SQL = """SELECT * from leave_request WHERE CURRENT_TO=%s"""
        try:
            self.cursor_filling.execute(SQL, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_filling=DB.cursor()
            self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_3.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 6, 7, 8, 9, 11, 12]:
                self.tableWidget_3.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                #if col != 4 and col != 5:
                    #self.tableWidget_3.resizeColumnToContents(col)
                col += 1
            self.tableWidget_3.resizeRowToContents(i)

        self.cursor_filling.close()

    def show_ot_contents(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.setColumnCount(9)
        self.cursor_filling = DB.cursor()
        SQL = """SELECT * from ot_request WHERE CURRENT_TO=%s"""
        try:
            self.cursor_filling.execute(SQL, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_filling=DB.cursor()
            self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_2.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 5, 6, 7, 8, 13]:
                self.tableWidget_2.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                # self.tableWidget_2.resizeColumnToContents(col)
                col += 1
            self.tableWidget_2.resizeRowToContents(i)

        self.cursor_filling.close()

    def show_late_contents(self):
        self.tableWidget_4.clearContents()
        self.tableWidget_4.setRowCount(0)
        self.tableWidget_4.setColumnCount(7)
        self.cursor_filling = DB.cursor()
        SQL = """SELECT * from apply_late WHERE CURRENT_TO=%s"""
        try:
            self.cursor_filling.execute(SQL, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_filling = DB.cursor()
            self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_4.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 5, 6]:
                self.tableWidget_4.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                if col == 4:
                    self.tableWidget_4.resizeColumnToContents(col)
                col += 1
            self.tableWidget_4.resizeRowToContents(i)

        self.cursor_filling.close()

    def show_forget_contents(self):
        self.tableWidget_8.clearContents()
        self.tableWidget_8.setRowCount(0)
        self.tableWidget_8.setColumnCount(12)
        self.cursor_filling = DB.cursor()
        SQL = """SELECT * from forget_record WHERE CURRENT_TO=%s"""
        try:
            self.cursor_filling.execute(SQL, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_filling = DB.cursor()
            self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_8.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                value=str(res[i][j])
                if j in [5,6,7,8,9,10]:
                    if res[i][j]!=None:
                        value=res[i][j].strftime("%H:%M:%S")

                self.tableWidget_8.setItem(i, col, QTableWidgetItem(value))
                #if col != 4 and col != 5:
                #self.tableWidget_8.resizeColumnToContents(col)
                col += 1
            self.tableWidget_8.resizeRowToContents(i)

        self.cursor_filling.close()
#-----------------------------------------------------
    def show_leave_history(self):
        self.tableWidget_3.clearContents()
        self.tableWidget_3.setRowCount(0)
        self.tableWidget_3.setColumnCount(15)
        self.tableWidget_3.setHorizontalHeaderLabels(['No.','Staff Name','Staff ID','Type','Submitted On','Start Date','Start Time','End Date','End Time','Duration(Day)','Remarks','Leader','DM','MD','HR'])
        cur = DB.cursor()
        SQL = """SELECT ID from team_stru WHERE LEADER_ID=%s OR DM_ID=%s OR MD_ID=%s"""
        try:
            cur.execute(SQL, (ID, ID, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (ID, ID, ID))
        res = cur.fetchall()
        if res == ():
            cur.close()
            return
        staff_lst=[]
        for i in range(len(res)):
            staff_lst.append(res[i][0])

        self.staff_lst=staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        SQL = """SELECT * FROM leave_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
        try:
            cur.execute(SQL, (date_min, date_max, staff_lst))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_3.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4, 6, 7, 8, 9, 11, 12,13,14,15,16]:
                value = str(res[i][j])
                if j in [13, 14,15,16]:
                    if res[i][j] == None:
                        value = "Unconfirmed"
                    elif res[i][j] == 1:
                        value = "OK"
                    else:
                        value = "Declined"

                self.tableWidget_3.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            #if i == 10:
                #continue
            #self.tableWidget_3.resizeColumnToContents(i)
            self.tableWidget_3.resizeRowToContents(i)
        cur.close()

    def show_ot_history(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.setColumnCount(13)
        self.tableWidget_2.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'OT Date', 'Start Time', 'Finish Time',
            'Duration(h)', 'Remarks', 'Leader', 'DM', 'MD', 'HR'])
        cur = DB.cursor()

        staff_lst=self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        SQL = """SELECT * FROM ot_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
        try:
            cur.execute(SQL, (date_min, date_max, staff_lst))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_2.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 5, 6, 7, 8, 13,9,10,11,12]:
                value = str(res[i][j])
                if j in [9, 10, 11, 12]:
                    if res[i][j] == None:
                        value = "Unconfirmed"
                    elif res[i][j] == 1:
                        value = "OK"
                    else:
                        value = "Declined"

                self.tableWidget_2.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            #if i == 8:
                #continue
            #self.tableWidget_2.resizeColumnToContents(i)
            self.tableWidget_2.resizeRowToContents(i)
        cur.close()

    def show_late_history(self):
        self.tableWidget_4.clearContents()
        self.tableWidget_4.setRowCount(0)
        self.tableWidget_4.setColumnCount(10)
        self.tableWidget_4.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'Date of Late Clock-In', 'Clock-In Time', 'Remarks', 'Leader','DM','HR'])
        cur = DB.cursor()

        staff_lst = self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        SQL = """SELECT * FROM apply_late WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
        try:
            cur.execute(SQL, (date_min, date_max, staff_lst))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_4.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4,5, 6, 7, 8, 9]:
                value = str(res[i][j])
                if j in [7,8,9]:
                    if res[i][j] == None:
                        value = "Unconfirmed"

                self.tableWidget_4.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        self.tableWidget_4.resizeColumnToContents(4)
        for i in range(len(res)):
            # if i == 8:
            # continue
            # self.tableWidget_4.resizeColumnToContents(i)
            self.tableWidget_4.resizeRowToContents(i)
        cur.close()

    def show_forget_history(self):
        self.tableWidget_8.clearContents()
        self.tableWidget_8.setRowCount(0)
        self.tableWidget_8.setColumnCount(15)
        self.tableWidget_8.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'Clock Date', 'Clock In', 'Clock Out',
             'Out-1', 'In-1','Out-2','In-2','Remarks','Leader','DM','HR'])
        cur = DB.cursor()

        staff_lst = self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        SQL = """SELECT * FROM forget_record WHERE REQUEST_DTTM>=%s AND REQUEST_DTTM<%s AND USER_ID IN %s"""
        try:
            cur.execute(SQL, (date_min, date_max, staff_lst))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_8.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4, 5, 6, 7, 8, 9,10,11,12,13,14]:
                value = str(res[i][j])
                if j in [12, 13, 14]:
                    if res[i][j] == None:
                        value = "Unconfirmed"

                self.tableWidget_8.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            # if i == 8:
            # continue
            # self.tableWidget_8.resizeColumnToContents(i)
            self.tableWidget_8.resizeRowToContents(i)
        cur.close()

#-----------------------------------------------------
    def show_leave_panel(self):
        self.textEdit_11.setText('')
        self.textEdit_12.setText('')
        self.textEdit_16.setText('')
        self.textEdit_13.setText('')
        self.textEdit_14.setText('')
        self.textEdit_15.setText('')
        self.textEdit_2.setText('')

        index = self.tableWidget_3.currentRow()
        try:
            staff_name = self.tableWidget_3.item(index, 1).text()
            staff_id = self.tableWidget_3.item(index, 2).text()
            type = self.tableWidget_3.item(index, 3).text()
            start_dt = self.tableWidget_3.item(index, 5).text()
            start_len = self.tableWidget_3.item(index, 6).text()
            end_dt = self.tableWidget_3.item(index, 7).text()
            end_len = self.tableWidget_3.item(index, 8).text()
            during = self.tableWidget_3.item(index, 9).text()
            remarks = self.tableWidget_3.item(index, 10).text()
        except:
            return

        self.textEdit_11.setText(staff_name)
        self.textEdit_12.setText(staff_id)
        self.textEdit_16.setText(type)
        self.textEdit_13.setText((str(start_dt) + ' (' + str(start_len) + ')').replace('all', 'all day'))
        if end_dt == 'None':
            self.textEdit_14.setText('(Single day leave)')
        else:
            self.textEdit_14.setText((str(end_dt) + ' (' + str(end_len) + ')').replace('all', 'all day'))
        self.textEdit_15.setText(during)
        self.textEdit_2.setText(remarks)

    def show_ot_panel(self):
        self.textEdit_10.setText('')
        self.textEdit_9.setText('')
        self.textEdit_6.setText('')
        self.textEdit_7.setText('')
        self.textEdit_8.setText('')
        self.textEdit_5.setText('')
        self.textEdit.setText('')

        index = self.tableWidget_2.currentRow()
        try:
            staff_name = self.tableWidget_2.item(index, 1).text()
            staff_id = self.tableWidget_2.item(index, 2).text()
            ot_dt = self.tableWidget_2.item(index, 4).text()
            start_tm = self.tableWidget_2.item(index, 5).text()
            end_tm = self.tableWidget_2.item(index, 6).text()
            during = self.tableWidget_2.item(index, 7).text()
            remarks = self.tableWidget_2.item(index, 8).text()
        except:
            return

        self.textEdit_10.setText(staff_name)
        self.textEdit_9.setText(staff_id)
        self.textEdit_6.setText(ot_dt)
        self.textEdit_7.setText(start_tm)
        self.textEdit_8.setText(end_tm)
        self.textEdit_5.setText(during)
        self.textEdit.setText(remarks)

    def show_late_panel(self):
        self.textEdit_17.setText('')
        self.textEdit_21.setText('')
        self.textEdit_22.setText('')
        self.textEdit_18.setText('')

        index = self.tableWidget_4.currentRow()
        try:
            staff_name = self.tableWidget_4.item(index, 1).text()
            staff_id = self.tableWidget_4.item(index, 2).text()
            late_dt = self.tableWidget_4.item(index, 4).text()
            clockin_tm = self.tableWidget_4.item(index, 5).text()
            remarks = self.tableWidget_4.item(index, 6).text()
        except:
            return

        self.textEdit_17.setText(staff_name)
        self.textEdit_21.setText(staff_id)
        self.textEdit_22.setText(late_dt)
        self.textEdit_18.setText(clockin_tm)
        self.textEdit_3.setText(remarks)

    def show_forget_panel(self):
        self.textEdit_39.setText('')
        self.textEdit_44.setText('')
        self.textEdit_45.setText('')
        self.textEdit_41.setText('')
        self.textEdit_42.setText('')
        self.textEdit_43.setText('')
        self.textEdit_46.setText('')
        self.textEdit_48.setText('')
        self.textEdit_47.setText('')
        self.textEdit_40.setText('')

        index = self.tableWidget_8.currentRow()
        try:
            staff_name = self.tableWidget_8.item(index, 1).text()
            staff_id = self.tableWidget_8.item(index, 2).text()
            clock_date = self.tableWidget_8.item(index, 4).text()
            clock_in = self.tableWidget_8.item(index, 5).text()
            clock_out = self.tableWidget_8.item(index, 6).text()
            out1=self.tableWidget_8.item(index, 7).text()
            in1=self.tableWidget_8.item(index, 8).text()
            out2=self.tableWidget_8.item(index, 9).text()
            in2=self.tableWidget_8.item(index, 10).text()
            remarks=self.tableWidget_8.item(index, 11).text()
        except:
            return

        self.textEdit_39.setText(staff_name)
        self.textEdit_44.setText(staff_id)
        self.textEdit_45.setText(clock_date)
        self.textEdit_41.setText(clock_in)
        self.textEdit_42.setText(clock_out)
        self.textEdit_43.setText(out1)
        self.textEdit_46.setText(in1)
        self.textEdit_48.setText(out2)
        self.textEdit_47.setText(in2)
        self.textEdit_40.setText(remarks)
#------------------------------------------------------
    def leave_accepted(self):
        index = self.tableWidget_3.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_3.item(index, 0).text()
        staff_id_ = self.tableWidget_3.item(index, 2).text()  #updated for 1.1
        type_ = self.tableWidget_3.item(index, 3).text() #updated for 1.1
        during_ = float(self.tableWidget_3.item(index, 9).text()) #updated for 1.1

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:
                if res[0][0] == res[0][1]:
                    if res[0][1] == res[0][2]:
                        current_to = 9999
                        current_po = 'HR'
                    else:
                        current_to = res[0][2]
                        current_po = 'MD'
                else:
                    current_to = res[0][1]
                    current_po = 'DM'
            else:
                current_to = res[0][2]
                current_po = 'MD'
        elif current_po == 'DM':
            if res[0][2] == res[0][1]:
                current_to = 9999
                current_po = 'HR'
            else:
                current_to = res[0][2]
                current_po = 'MD'
        else:
            current_to = 9999
            current_po = 'HR'

        if current_po == 'DM':
            SQL = """UPDATE leave_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))
        elif current_po == 'MD':
            SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))
        else:
            SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, MD=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

        DB.commit()

        self.cursor_approve.close()
        self.show_leave_contents()
        self.show_leave_panel()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2=query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='leave')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='leave')

            if type_== 'Annual leave':
                self.cursor_calc = DB.cursor()
                SQL="""SELECT AN_DAYS FROM akt_staff WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res=self.cursor_calc.fetchall()
                if res==():
                    pass
                else:
                    an_days=res[0][0]
                    an_days-=during_
                    SQL="""UPDATE akt_staff SET AN_DAYS=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (an_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
            elif type_=='Sick leave':
                self.cursor_calc = DB.cursor()
                SQL = """SELECT SICK_DAYS FROM akt_staff WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res = self.cursor_calc.fetchall()
                if res == ():
                    pass
                else:
                    sick_days = res[0][0]
                    sick_days -= during_
                    SQL = """UPDATE akt_staff SET SICK_DAYS=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (sick_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
            elif type_=='Hometown':
                self.cursor_calc = DB.cursor()
                SQL = """SELECT HOME_TOWN FROM akt_staff WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res = self.cursor_calc.fetchall()
                if res == ():
                    pass
                else:
                    home_days = res[0][0]
                    home_days -= 1
                    SQL = """UPDATE akt_staff SET HOME_TOWN=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (home_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
        # ===========================================================


        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def leave_declined(self):
        index = self.tableWidget_3.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_3.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE leave_request SET LEADER=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))
        elif current_po == 'DM':
            SQL = """UPDATE leave_request SET DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))
        else:
            SQL = """UPDATE leave_request SET MD=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_leave_contents()
        self.show_leave_panel()

        # Mail sending 2==============================================
        info_lst2=query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='leave')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def ot_accepted(self):
        index = self.tableWidget_2.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_2.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:
                if res[0][0] == res[0][1]:
                    if res[0][1] == res[0][2]:
                        current_to = 9999
                        current_po = 'HR'
                    else:
                        current_to = res[0][2]
                        current_po = 'MD'
                else:
                    current_to = res[0][1]
                    current_po = 'DM'
            else:
                current_to = res[0][2]
                current_po = 'MD'
        elif current_po == 'DM':
            if res[0][2] == res[0][1]:
                current_to = 9999
                current_po = 'HR'
            else:
                current_to = res[0][2]
                current_po = 'MD'
        else:
            current_to = 9999
            current_po = 'HR'

        if current_po == 'DM':
            SQL = """UPDATE ot_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))
        elif current_po == 'MD':
            SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))
        else:
            SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, MD=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_ot_contents()
        self.show_ot_panel()
        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2= query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='ot')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='ot')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def ot_declined(self):
        index = self.tableWidget_2.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_2.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE ot_request SET LEADER=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))
        elif current_po == 'DM':
            SQL = """UPDATE ot_request SET DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))
        else:
            SQL = """UPDATE ot_request SET MD=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'HR', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_ot_contents()
        self.show_ot_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='ot')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def late_accepted(self):
        index = self.tableWidget_4.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_4.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:
                if res[0][0] == res[0][1]:
                    current_to = 9999
                    current_po = 'HR'
                else:
                    current_to = res[0][1]
                    current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'
        else:
            current_to = 9999
            current_po = 'HR'

        if current_po == 'DM':
            SQL = """UPDATE apply_late SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))

        else:
            SQL = """UPDATE apply_late SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_late_contents()
        self.show_late_panel()
        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2 = query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='late')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='late')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def late_declined(self):
        index = self.tableWidget_4.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_4.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE apply_late SET LEADER=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))
        else:
            SQL = """UPDATE apply_late SET DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_late_contents()
        self.show_late_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='late')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def forget_accepted(self):
        index = self.tableWidget_8.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_8.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:
                if res[0][0] == res[0][1]:
                    current_to = 9999
                    current_po = 'HR'
                else:
                    current_to = res[0][1]
                    current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'
        else:
            current_to = 9999
            current_po = 'HR'

        if current_po == 'DM':
            SQL = """UPDATE forget_record SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))

        else:
            SQL = """UPDATE forget_record SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_forget_contents()
        self.show_forget_panel()

        if current_to == 9999:
            SQL="""SELECT USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2 FROM forget_record WHERE SERIAL=%s"""
            self.cursor_migrant=DB.cursor()
            self.cursor_migrant.execute(SQL, (request_id))
            data=self.cursor_migrant.fetchall()
            user_id=data[0][0]
            clockin=data[0][1]
            clockout=data[0][2]
            out1=data[0][3]
            in1=data[0][4]
            out2=data[0][5]
            in2=data[0][6]

            try:
                serial_timecard=str(user_id)+clockin.strftime('%Y%m%d')
            except AttributeError:
                serial_timecard=str(user_id)+clockout.strftime('%Y%m%d')

            if clockin==None:
                SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data = self.cursor_migrant.fetchall()
                if data != ():
                    SQL = """UPDATE time_card SET CLOCK_OUT=%s WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (clockout, serial_timecard))
                    DB.commit()
                else:
                    SQL="""INSERT INTO time_card (SERIAL,USER_ID,CLOCK_OUT) VALUES (%s,%s,%s)"""
                    self.cursor_migrant.execute(SQL,(serial_timecard,user_id,clockout))
                    DB.commit()

            elif clockout==None:
                SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data = self.cursor_migrant.fetchall()
                if data != ():
                    SQL = """UPDATE time_card SET CLOCK_IN=%s WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (clockin, serial_timecard))
                    DB.commit()
                else:
                    SQL = """INSERT INTO time_card (SERIAL,USER_ID,CLOCK_IN) VALUES (%s,%s,%s)"""
                    self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin))
                    DB.commit()

            else:
                SQL="""SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data=self.cursor_migrant.fetchall()
                if data!=():
                    SQL="""DELETE FROM time_card WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (serial_timecard))
                    DB.commit()

                SQL="""INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
                self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin, clockout, out1, in1, out2, in2))
                DB.commit()

            self.cursor_migrant.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2 = query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='forget')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='forget')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def forget_declined(self):
        index = self.tableWidget_8.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_8.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE forget_record SET LEADER=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))
        else:
            SQL = """UPDATE forget_record SET DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_forget_contents()
        self.show_forget_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='forget')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def quit(self):
        ApprovePanel.close()

    def closeEvent(self, event):
        self.setupUi(self)
        self.binding_btn()
        MainWindow.show()

class ApplyLateClockIn(QDialog, Ui_ApplyLateClockIn):
    def __init__(self):
        super(ApplyLateClockIn, self).__init__()
        self.setupUi(self)

        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.timeEdit_2.setTime(QTime.fromString('09:00', 'hh:mm'))
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.pushButton_3.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.apply_late)
        self.dateEdit.dateChanged.connect(self.show_tableData)
        self.comboBox_3.currentTextChanged.connect(self.show_tableData)
        self.pushButton_2.clicked.connect(self.withdraw_apply)

    def initialize(self):
        self.show_tableData()

    def show_tableData(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM apply_late WHERE LATE_DT>=%s AND LATE_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            table_col=0
            for j in [0, 3, 4, 5, 6, 7, 8, 9]:
                value=str(res[i][j])
                if j in [7, 8, 9]:
                    if str(res[i][j])=="None":
                        value="Unconfirmed"
                    elif str(res[i][j])=="OK":
                        value="OK"
                    else:
                        value="Declined"

                self.tableWidget.setItem(i, table_col, QTableWidgetItem(value))
                table_col+=1

        for i in range(8):
            if i==4:
                continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()

    def apply_late(self):
        late_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        late_dt_compare=late_dt.strftime('%Y%m%d')
        now_dt_compare=TimeCard.t.strftime('%Y%m%d')
        if late_dt_compare<=now_dt_compare:
            QMessageBox.critical(self, 'Warning','You can only apply for late clock-in 1 day beforehand.')
            return

        user_id = ID
        user_name = MainWindow.name
        apply_dttm = TimeCard.t
        clockin_tm = datetime.datetime.strptime(self.timeEdit_2.time().toString('hh:mm'), '%H:%M')
        remarks = self.textEdit.toPlainText()

        msm = f'Apply late clock-in?\n' \
              f'Late date: {late_dt.strftime("%d/%m/%Y")}\n' \
              f'Clock-in time: {clockin_tm.strftime("%H:%M:%S")}\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor_apply=DB.cursor()
        SQL="""SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor_apply.execute(SQL, (user_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_apply=DB.cursor()
            self.cursor_apply.execute(SQL, (user_id))

        res=self.cursor_apply.fetchall()

        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'

        if current_to!=9999:
            sql = """INSERT INTO apply_late (USER_ID, USER_NAME, APPLY_DTTM, LATE_DT, CLOCKIN_TM, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
                (%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (user_id, user_name, apply_dttm, late_dt, clockin_tm, remarks, current_to, current_po))
            DB.commit()
        else:
            sql = """INSERT INTO apply_late (USER_ID, USER_NAME, APPLY_DTTM, LATE_DT, CLOCKIN_TM, REMARKS,LEADER, DM, HR, CURRENT_TO, CURRENT_PO) VALUES
                            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
            user_id, user_name, apply_dttm, late_dt, clockin_tm, remarks, 'OK','OK','OK',current_to, current_po))
            DB.commit()

        self.cursor_apply.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='late')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to==9999:
            info_lst = query_email(id=ID)
            if info_lst == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst[1],
                                              receiver_name=info_lst[0],
                                              mode='late')
        # ===========================================================

        QMessageBox.information(self, 'Info',
                                'Late clock-in request has been sent successfully! Please wait for the approvements.')
        self.show_tableData()

    def withdraw_apply(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 7).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.cursor.close()
        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.show_tableData()

    def quit(self):
        ApplyLateClockIn.close()

    def closeEvent(self, event):
        TimeCard.show()

class ForgetRecord(QDialog, Ui_ForgetRecord):
    def __init__(self):
        super(ForgetRecord, self).__init__()
        self.setupUi(self)

        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)
        self.timeEdit_2.setTime(QTime.fromString('08:30', 'hh:mm'))
        self.timeEdit_3.setTime(QTime.fromString('17:45', 'hh:mm'))
        self.timeEdit_4.setTime(QTime.fromString('12:00', 'hh:mm'))
        self.timeEdit_5.setTime(QTime.fromString('13:00', 'hh:mm'))
        self.timeEdit_6.setTime(QTime.fromString('14:00', 'hh:mm'))
        self.timeEdit_7.setTime(QTime.fromString('15:00', 'hh:mm'))
        self.checkBox_2.setEnabled(False)
        self.timeEdit_4.setEnabled(False)
        self.timeEdit_5.setEnabled(False)
        self.timeEdit_6.setEnabled(False)
        self.timeEdit_7.setEnabled(False)
        self.checkBox.setChecked(False)
        self.label_31.setText('')

        self.timeEdit_2.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_3.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_4.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_5.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_6.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_7.timeChanged.connect(self.time_stamp_check)
        self.checkBox.stateChanged.connect(self.time_stamp_check)
        self.checkBox_2.stateChanged.connect(self.time_stamp_check)
        self.checkBox_3.stateChanged.connect(self.time_stamp_check)
        self.checkBox_4.stateChanged.connect(self.time_stamp_check)
        self.checkBox_3.stateChanged.connect(self.singleTime_or_not)
        self.checkBox_4.stateChanged.connect(self.singleTime_or_not)

        self.checkBox_3.stateChanged.connect(self.uncheck_clock_in)
        self.checkBox_4.stateChanged.connect(self.uncheck_clock_out)

        self.checkBox.stateChanged.connect(self.if_in_out_1)
        self.checkBox_2.stateChanged.connect(self.if_in_out_2)
        self.pushButton_3.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.submit)
        self.dateEdit.dateChanged.connect(self.show_tableData)
        self.comboBox_3.currentTextChanged.connect(self.show_tableData)
        self.pushButton_2.clicked.connect(self.withdraw_apply)

    def test(self):
        print(self.timeEdit_2.text())
        print(self.timeEdit_2.text().replace(':',''))
        print(type(self.timeEdit_2.text()))

    def initialize(self):
        self.show_tableData()

    def show_tableData(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM forget_record WHERE CLOCK_DT>=%s AND CLOCK_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                value = str(res[i][j])
                if j in [12, 13, 14]:
                    if str(res[i][j]) == "None":
                        value = "Unconfirmed"
                    elif str(res[i][j]) == "OK":
                        value = "OK"
                    else:
                        value = "Declined"

                if j in [5,6,7,8,9,10]:
                    if res[i][j]!=None:
                        value=res[i][j].strftime("%H:%M:%S")

                self.tableWidget.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(13):
            if i == 9:
                continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()

    def submit(self):
        if self.label_31.text()!='':
            QMessageBox.critical(self, 'Error', 'Error: Please input the correct time range!')
            return

        forgot_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        forgot_dt_compare = forgot_dt.strftime('%Y%m%d')
        now_dt_compare = TimeCard.t.strftime('%Y%m%d')
        if forgot_dt_compare > now_dt_compare:
            QMessageBox.critical(self, 'Error', 'Error: You can not submit forget-time-record report for the future\'s date!')
            return

        user_id = ID
        user_name = MainWindow.name
        request_dttm = TimeCard.t
        clock_dt=forgot_dt
        clockin = datetime.datetime.strptime(self.dateEdit_3.date().toString('yyyy/MM/dd/')+self.timeEdit_2.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        clockout = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_3.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out1=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_4.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in1=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_5.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out2=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_6.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in2=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_7.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        remarks = self.textEdit.toPlainText()

        if not self.checkBox.isChecked():
            out1=None
            in1=None

        if not self.checkBox_2.isChecked():
            out2=None
            in2=None

        if not self.checkBox_3.isChecked():
            clockin=None

        if not self.checkBox_4.isChecked():
            clockout=None

        time_func = lambda y: None if y == None else y.strftime("%H:%M:%S")
        msm = f'Submit request?\n' \
              f'Date of clock in/out forgotten: {clock_dt.strftime("%d/%m/%Y")}\n' \
              f'Clock-in: {time_func(clockin)}\n' \
              f'Clock-out: {time_func(clockout)}\n' \
              f'Out-1: {time_func(out1)}\n' \
              f'In-1: {time_func(in1)}\n' \
              f'Out-2: {time_func(out2)}\n' \
              f'In-2: {time_func(in2)}\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor_apply = DB.cursor()
        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor_apply.execute(SQL, (user_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_apply = DB.cursor()
            self.cursor_apply.execute(SQL, (user_id))

        res = self.cursor_apply.fetchall()

        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'

        if current_to!=9999:
            sql = """INSERT INTO forget_record (USER_ID, USER_NAME, REQUEST_DTTM, CLOCK_DT, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
                        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
            user_id, user_name, request_dttm, clock_dt, clockin, clockout, out1, in1, out2, in2, remarks, current_to, current_po))
            DB.commit()
        else:
            sql = """INSERT INTO forget_record (USER_ID, USER_NAME, REQUEST_DTTM, CLOCK_DT, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2, REMARKS,LEADER, DM, HR, CURRENT_TO, CURRENT_PO) VALUES
                                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
                user_id, user_name, request_dttm, clock_dt, clockin, clockout, out1, in1, out2, in2, remarks, 'OK','OK','OK',
                current_to, current_po))
            DB.commit()

        self.cursor_apply.close()

        if current_to == 9999:
            self.cursor_migrant = DB.cursor()
            serial_timecard = str(user_id) + clockin.strftime('%Y%m%d')

            SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
            self.cursor_migrant.execute(SQL, (serial_timecard))
            data = self.cursor_migrant.fetchall()
            if data != ():
                SQL = """DELETE FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                DB.commit()

            SQL = """INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin, clockout, out1, in1, out2, in2))
            DB.commit()

            self.cursor_migrant.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='forget')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            info_lst = query_email(id=ID)
            if info_lst == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst[1],
                                              receiver_name=info_lst[0],
                                              mode='forget')
        # ===========================================================

        QMessageBox.information(self, 'Info',
                                'Clock in/out forgotten report has been sent successfully! Please wait for the approvements.')
        self.show_tableData()

    def withdraw_apply(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 12).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.cursor.close()
        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.show_tableData()

    def uncheck_clock_in(self):
        if not self.checkBox_3.isChecked():
            if not self.checkBox_4.isChecked():
                self.checkBox_4.setChecked(True)

    def uncheck_clock_out(self):
        if not self.checkBox_4.isChecked():
            if not self.checkBox_3.isChecked():
                self.checkBox_3.setChecked(True)

    def singleTime_or_not(self):
        if self.checkBox_3.isChecked()==False or self.checkBox_4.isChecked()==False:
            self.checkBox.setChecked(False)
            self.checkBox.setEnabled(False)
        else:
            self.checkBox.setEnabled(True)

        if self.checkBox_3.isChecked() == False:
            self.timeEdit_2.setEnabled(False)
        else:
            self.timeEdit_2.setEnabled(True)

        if self.checkBox_4.isChecked() == False:
            self.timeEdit_3.setEnabled(False)
        else:
            self.timeEdit_3.setEnabled(True)

    def time_stamp_check(self):
        if self.checkBox_3.isChecked() == False or self.checkBox_4.isChecked() == False:
            self.label_31.setText('')
            return

        in_=int(self.timeEdit_2.text().replace(':',''))
        out_1=int(self.timeEdit_4.text().replace(':',''))
        in_1 = int(self.timeEdit_5.text().replace(':', ''))
        out_2 = int(self.timeEdit_6.text().replace(':', ''))
        in_2 = int(self.timeEdit_7.text().replace(':', ''))
        out_=int(self.timeEdit_3.text().replace(':',''))

        if self.checkBox.isChecked()==False:
            if not in_<out_:
                self.label_31.setText('Please input the correct time range!')
            else:
                self.label_31.setText('')

        else:
            if self.checkBox_2.isChecked()==False:
                if not in_<out_1<in_1<out_:
                    self.label_31.setText('Please input the correct time range!')
                else:
                    self.label_31.setText('')
            else:
                if not in_<out_1<in_1<out_2<in_2<out_:
                    self.label_31.setText('Please input the correct time range!')
                else:
                    self.label_31.setText('')

    def if_in_out_1(self):
        if self.checkBox.isChecked():
            self.timeEdit_4.setEnabled(True)
            self.timeEdit_5.setEnabled(True)
            self.checkBox_2.setEnabled(True)
        else:
            self.timeEdit_4.setEnabled(False)
            self.timeEdit_5.setEnabled(False)
            self.checkBox_2.setChecked(False)
            self.checkBox_2.setEnabled(False)

    def if_in_out_2(self):
        if self.checkBox_2.isChecked():
            self.timeEdit_6.setEnabled(True)
            self.timeEdit_7.setEnabled(True)
        else:
            self.timeEdit_6.setEnabled(False)
            self.timeEdit_7.setEnabled(False)

    def quit(self):
        ForgetRecord.close()

    def closeEvent(self, event):
        TimeCard.show()

class AdminMain(QMainWindow, Ui_AdminMain):
    def __init__(self):
        super(AdminMain, self).__init__()
        self.setupUi(self)

        self.pushButton_7.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.to_staff_manage)
        self.pushButton_2.clicked.connect(self.to_team_stru)
        self.pushButton_3.clicked.connect(self.to_login_pass)
        self.pushButton_9.clicked.connect(self.to_calendar_setting)

    def to_calendar_setting(self):
        CalendarSetting.show()
        CalendarSetting.show_on_table()
        self.setVisible(False)

    def to_login_pass(self):
        LoginPass.show()
        LoginPass.show_on_table()
        self.setVisible(False)

    def to_staff_manage(self):
        StaffManage.show()
        StaffManage.show_on_table()
        self.setVisible(False)

    def to_team_stru(self):
        TeamStructure.show()
        TeamStructure.show_on_table()
        self.setVisible(False)

    def quit(self):
        self.pushButton_7.setAttribute(Qt.WA_UnderMouse, False)
        AdminMain.destroy()
        MainWindow.setVisible(True)

    def closeEvent(self, event):
        self.quit()

class StaffManage(QMainWindow, Ui_StaffManage):
    def __init__(self):
        super(StaffManage, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        # print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM akt_staff"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(9):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb=xl.load_workbook(a[0])
        ws=wb.active
        if ws.cell(row=1, column=1).value.strip()!='ID'\
                or ws.cell(row=1, column=2).value.strip()!='NAME'\
                or ws.cell(row=1, column=3).value.strip()!='POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'DIVISION' \
                or ws.cell(row=1, column=5).value.strip() != 'EMAIL' \
                or ws.cell(row=1, column=6).value.strip() != 'ANNUAL LEAVE' \
                or ws.cell(row=1, column=7).value.strip() != 'SICK LEAVE' \
                or ws.cell(row=1, column=8).value.strip() != 'HOMETOWN TRAVEL' \
                or ws.cell(row=1, column=9).value.strip() != 'REMARKS':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data=[]
        for i in range(2, ws.max_row+1):
            line=[]
            for j in range(1, ws.max_column+1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL="""DELETE FROM akt_staff WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO akt_staff (ID, NAME, POSITION, DIVISION, EMAIL, AN_DAYS, SICK_DAYS, HOME_TOWN, REMARK) VALUES (%s, %s,%s,%s,%s,%s,%s,%s,%s)"""
            try:
                cur.execute(SQL, (
                each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7],each[8]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur=DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Staff management database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Staff Information',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(9):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'DIVISION', 'EMAIL', 'ANNUAL LEAVE', 'SICK LEAVE', 'HOMETOWN TRAVEL', 'REMARKS']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        StaffManage.destroy()

    def closeEvent(self, event):
        self.quit()

class TeamStructure(QMainWindow, Ui_TeamStructure):
    def __init__(self):
        super(TeamStructure, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        # print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM team_stru"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(10):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'ID' \
                or ws.cell(row=1, column=2).value.strip() != 'NAME' \
                or ws.cell(row=1, column=3).value.strip() != 'POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'DIVISION' \
                or ws.cell(row=1, column=5).value.strip() != 'LEADER ID' \
                or ws.cell(row=1, column=6).value.strip() != 'LEADER NAME' \
                or ws.cell(row=1, column=7).value.strip() != 'DM ID' \
                or ws.cell(row=1, column=8).value.strip() != 'DM NAME' \
                or ws.cell(row=1, column=9).value.strip() != 'MD ID'\
                or ws.cell(row=1, column=10).value.strip() != 'MD NAME':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column+1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM team_stru WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO team_stru (ID, NAME, POSITION, DIVISION, LEADER_ID, LEADER_NAME, DM_ID, DM_NAME, MD_ID, MD_NAME) VALUES (%s, %s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Team structure database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Team Structure',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(10):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'DIVISION', 'LEADER ID', 'LEADER NAME', 'DM ID', 'DM NAME', 'MD ID', 'MD NAME',]
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        TeamStructure.destroy()

    def closeEvent(self, event):
        self.quit()

class LoginPass(QMainWindow, Ui_LoginPass):
    def __init__(self):
        super(LoginPass, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        self.cursor = DB.cursor()
        sql = """SELECT * FROM login_pass"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(5):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'ID' \
                or ws.cell(row=1, column=2).value.strip() != 'NAME' \
                or ws.cell(row=1, column=3).value.strip() != 'POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'PASSWORD' \
                or ws.cell(row=1, column=5).value.strip() != 'PRIORITY':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column + 1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM login_pass WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO login_pass (ID, NAME, POSITION, PASSWORD, PRIORITY) VALUES (%s, %s,%s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Staff account database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Staff Accounts',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(5):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'PASSWORD', 'PRIORITY']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        LoginPass.destroy()

    def closeEvent(self, event):
        self.quit()

class CalendarSetting(QMainWindow, Ui_CalendarSetting):
    def __init__(self):
        super(CalendarSetting, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        self.cursor = DB.cursor()
        sql = """SELECT * FROM calendar"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(4):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'DATE' \
                or ws.cell(row=1, column=2).value.strip() != 'WEEKDAY' \
                or ws.cell(row=1, column=3).value.strip() != 'IF WORK' \
                or ws.cell(row=1, column=4).value.strip() != 'REMARKS':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column + 1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM calendar WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO calendar (DATE, WEEKDAY, IF_WORK, REMARKS) VALUES (%s, %s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Calendar database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Calendar',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(4):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['DATE', 'WEEKDAY', 'IF_WORK', 'REMARKS']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        CalendarSetting.destroy()

    def closeEvent(self, event):
        self.quit()

def calculate_worktime(data_line):
    day_lag = data_line[8]

    clock_out = data_line[3]
    if clock_out == None:
        clock_out = '-'

    clock_in = data_line[2]
    try:
        out1 = data_line[4]
    except:
        out1 = '-'
    if out1 == None:
        out1 = '-'
    try:
        in1 = data_line[5]
    except:
        in1 = '-'
    if in1 == None:
        in1 = '-'
    try:
        out2 = data_line[6]
    except:
        out2 = '-'
    if out2 == None:
        out2 = '-'
    try:
        in2 = data_line[7]
    except:
        in2 = '-'
    if in2 == None:
        in2 = '-'

    if clock_out == '-':
        work_time = '-'
        over_time = '-'
    # ---------------------计算实际作业时间
    else:
        if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
            # 1
            if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                work_time = (clock_out - clock_in).seconds
            # 2
            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 3
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
                work_time -= 3600
            # 4
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
                work_time -= 3600
            # 5
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 4500

        elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1230'):
            # 6
            if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = 0
            # 7
            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
            # 8
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230',
                    '%Y%m%d%H%M')).seconds
            # 9
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1500'):
            # 10
            if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
            # 11
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 12
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1515'):
            # 13
            if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = 0
            # 14
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
        else:
            work_time = (clock_out - clock_in).seconds

        # ------------------计算离岗累计时间
        space_time1 = 0
        space_time2 = 0
        if out1 != '-' and in1 != '-':
            # 1
            if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time1 = (in1 - out1).seconds
            # 2
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 3
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds - 3600
            # 4
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds - 3600
            # 5
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 3600 - 900
            # 6
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = 0
            # 7
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds
            # 11
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 12
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 900
            # 13
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = 0
            # 14
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time1 = (in1 - out1).seconds

        # -----------------out2 in2
        if out2 != '-' and in2 != '-':
            # 1
            if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time2 = (in2 - out2).seconds
            # 2
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 3
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds - 3600
            # 4
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds - 3600
            # 5
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 3600 - 900
            # 6
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = 0
            # 7
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds
            # 11
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 12
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 900
            # 13
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = 0
            # 14
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time2 = (in2 - out2).seconds

        work_time -= space_time1
        work_time -= space_time2
        # --------------------------如过午夜加上一天
        if day_lag == 1:
            work_time += 86400

        over_time = work_time - 28800
        if over_time < 0:
            over_time = 0

    if work_time != '-':
        work_time = round(int(work_time) / 3600, 2)
    if over_time != '-':
        over_time = round(int(over_time) / 3600, 2)

    qdt = clock_in
    dt_for_query = datetime.datetime.strptime(qdt.strftime('%d/%m/%Y'), '%d/%m/%Y')
    sql = """SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
    cursor_queryhours = DB.cursor()
    try:
        cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
    except:
        reconnect_DB(MainWindow)
        cursor_queryhours = DB.cursor()
        cursor_queryhours.execute(sql, (ID, dt_for_query), '9999')
    res = cursor_queryhours.fetchall()
    cursor_queryhours.close()
    if res == ():
        approved_ot='0'
    else:
        approved_ot = str(res[0][0])

    return [work_time, over_time, approved_ot]

def query_email(id):
    sql = """SELECT NAME, EMAIL FROM akt_staff WHERE ID=%s"""
    cursor_query=DB.cursor()
    try:
        cursor_query.execute(sql, (id))
    except pymysql.err.OperationalError:
        reconnect_DB(MainWindow)
        cursor_query = DB.cursor()
        cursor_query.execute(sql, (id))
    res=cursor_query.fetchall()
    cursor_query.close()
    if res==():
        return -1
    name=res[0][0]
    email=res[0][1]
    return [name, email]

def reconnect_DB(form):
    global DB
    try:
        DB.close()
    except:
        pass
    try:
        DB = pymysql.connect(host='210.1.31.3',
                             user='hr',
                             port=3306,
                             passwd='gwP6xTsA',
                             db='TEST_AKT1')
    except pymysql.err.OperationalError:
        QMessageBox.critical(form, 'Network Error', 'Can not connect to the server, please check your network!')

if __name__ == '__main__':

    DB = None
    ID = -1

    app = QApplication(sys.argv)

    loginWindow = loginWindow()
    loginWindow.show()

    MainWindow = MainWindow()
    PassWindow = PassWindow()
    TimeCard = TimeCard()
    AskForLeave = AskForLeave()
    OTApplication = OTApplication()
    BookMeetingRoom = BookMeetingRoom()
    ApprovePanel = ApprovePanel()
    ApplyLateClockIn=ApplyLateClockIn()
    ForgetRecord=ForgetRecord()

    mailsender=MailSender()

    AdminMain = AdminMain()
    StaffManage = StaffManage()
    TeamStructure = TeamStructure()
    LoginPass = LoginPass()
    CalendarSetting = CalendarSetting()

    sys.exit(app.exec_())
