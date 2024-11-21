import pymysql
import openpyxl as xl

wb = xl.load_workbook('./db_excel/team_structure.xlsx')
ws = wb.active
data = []
for i in range(2, ws.max_row + 1):
    line = []
    for j in range(1, ws.max_column + 1):
        line.append(ws.cell(row=i, column=j).value)

    data.append(line)

for each in data:
    print(each)


db=pymysql.connect(host='210.1.31.3',
                   user='hr',
                   port=3306,
                   passwd='gwP6xTsA',
                   db='akaganeHR')

cursor=db.cursor()

for each in data:
    #if each[4]==None:
    sql = f"""INSERT INTO team_stru (ID, NAME, POSITION, DIVISION, LEADER_ID,LEADER_NAME,DM_ID,DM_NAME,MD_ID,MD_NAME) 
        VALUES ({int(each[0])}, '{str(each[1])}', '{str(each[2])}', 
        '{str(each[3])}', '{str(each[4])}','{str(each[5])}','{str(each[6])}','{str(each[7])}','{str(each[8])}','{str(each[9])}')"""
    #else:
    #    sql = f"""INSERT INTO login_pass (ID, NAME, POSITION, PASSWORD, PRIORITY)
    #            VALUES ({int(each[0])}, '{str(each[1])}', '{str(each[2])}',
    #            '{str(each[3])}', '{str(each[4])}')"""
    try:
        # 执行sql语句
        cursor.execute(sql)
        # 提交到数据库执行
        db.commit()
        print(cursor.fetchall())
    except:
        # 如果发生错误则回滚
        db.rollback()
        print(1)

cursor.close()
# 关闭数据库连接
db.close()