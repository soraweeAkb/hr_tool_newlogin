
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb=xl.Workbook()
ws=wb.active

ws['B1'].value='AKAGANE (Thailand) Co.,Ltd.'
ws['B1'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=16)
ws['B1'].alignment = Alignment(horizontal='left',vertical='center',wrap_text=False)

FontOBJ=Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11)
AlignOBJ=Alignment(horizontal='left', vertical='center',wrap_text=False)

ws['B2'].value='Staff Name'
ws['B2'].font=FontOBJ
ws['B2'].alignment=AlignOBJ

ws['B3'].value='Staff Code'
ws['B3'].font=FontOBJ
ws['B3'].alignment=AlignOBJ

ws['B4'].value='Period'
ws['B4'].font=FontOBJ
ws['B4'].alignment=AlignOBJ

ws['C2'].value='Kanniga Ubonsai'
ws['C2'].font=FontOBJ
ws['C2'].alignment=AlignOBJ

ws['C3'].value='1001'
ws['C3'].font=FontOBJ
ws['C3'].alignment=AlignOBJ

ws['C4'].value="16 Nov'20 to 15 Dec'20"
ws['C4'].font=FontOBJ
ws['C4'].alignment=AlignOBJ

ws['B6'].value='Official Time'
ws['B6'].font=FontOBJ
ws['B6'].alignment=AlignOBJ

#------------------------------Inside table frame
AlignOBJ=Alignment(horizontal='center', vertical='center',wrap_text=False)

ws.cell(row=7,column=2).value='Date'
ws.cell(row=7,column=3).value='Weekday'
ws.cell(row=7,column=4).value='Check In'
ws.cell(row=7,column=5).value='Start'
ws.cell(row=7,column=6).value='Out 1'
ws.cell(row=7,column=7).value='In 1'
ws.cell(row=7,column=8).value='Out 2'
ws.cell(row=7,column=9).value='In 2'
ws.cell(row=7,column=10).value='Out Hour'
ws.cell(row=7,column=11).value='Lunch'
ws.cell(row=7,column=12).value='Lunch'
ws.cell(row=7,column=13).value='Tea Time'
ws.cell(row=7,column=14).value='Return Time'
ws.cell(row=7,column=15).value='Check Out'
ws.cell(row=7,column=16).value='Worked Hour'
ws.cell(row=7,column=17).value='Excl. Break'
ws.cell(row=7,column=18).value='Base Hour'
ws.cell(row=7,column=19).value='OT Hour'
ws.cell(row=7,column=20).value='Dinner Count'
ws.cell(row=7,column=21).value='Lunch Count'

for j in range(2, 22):
    ws.cell(row=7, column=j).font=FontOBJ
    ws.cell(row=7, column=j).alignment=AlignOBJ

AlignOBJ=Alignment(horizontal='center', vertical='center',wrap_text=True)
for j in [14,20,21]:
    ws.cell(row=7, column=j).alignment = AlignOBJ

FontOBJ=Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
for j in [11, 12, 13, 14, 18]:
    ws.cell(row=7, column=j).font=FontOBJ

ws.row_dimensions[1].height=21.75
ws.row_dimensions[2].height=21.75
ws.row_dimensions[3].height=21.75
ws.row_dimensions[4].height=21.75
ws.row_dimensions.group(5,hidden=True)
ws.row_dimensions[6].height=14.25

ws.row_dimensions[7].height=26.25
ws.row_dimensions[45].height=39

ws.column_dimensions['A'].width=1.63 #0.92
ws.column_dimensions['B'].width=13.57+0.71
ws.column_dimensions['C'].width=9.43+0.71
ws.column_dimensions['D'].width=8.7+0.71
ws.column_dimensions['E'].width=8.43+0.71
ws.column_dimensions['F'].width=9.57+0.71
ws.column_dimensions['G'].width=9.57+0.71
ws.column_dimensions['H'].width=9.57+0.71
ws.column_dimensions['I'].width=9.57+0.71
ws.column_dimensions['J'].width=9.57+0.71
ws.column_dimensions['K'].width=9.57+0.71
ws.column_dimensions['L'].width=9.57+0.71
ws.column_dimensions['M'].width=9.57+0.71
ws.column_dimensions['N'].width=9.57+0.71
ws.column_dimensions['O'].width=10.71+0.71
ws.column_dimensions['P'].width=14.29+0.71
ws.column_dimensions['Q'].width=15.14+0.71
ws.column_dimensions['R'].width=12.43+0.71
ws.column_dimensions['S'].width=10.29+0.71
ws.column_dimensions['T'].width=8.14+0.71
ws.column_dimensions['U'].width=7.71+0.71
ws.column_dimensions['V'].width=10.43+0.71

FontOBJ=Font(name=u'ＭＳ Ｐゴシック', bold=False, italic=False, size=11)

for i in range(8, 42):
    for j in [2,3,4,5,6,7,8,9,10,15,16,17,19,20,21]:
        ws.cell(row=i, column=j).font=FontOBJ

FontOBJ=Font(name=u'ＭＳ Ｐゴシック', bold=False, italic=False, size=11, color="FF0000")
for i in range(8,42):
    for j in [11,12,13,14,18]:
        ws.cell(row=i, column=j).font = FontOBJ

FontOBJ=Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11)
for i in [42,45]:
    for j in range(2,23):
        ws.cell(row=i, column=j).font=FontOBJ

border = Border(left=Side(border_style='medium',color='000000'),
                right=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'))

for i in [7]:
    for j in range(2, 22):
        ws.cell(row=i, column=j).border=border

border = Border(left=Side(border_style='medium',color='000000'),
                right=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='thin',color='000000'))

for i in [8]:
    for j in range(2,22):
        ws.cell(row=i,column=j).border=border

border = Border(left=Side(border_style='medium',color='000000'),
                right=Side(border_style='medium',color='000000'),
                top=Side(border_style='thin',color='000000'),
                bottom=Side(border_style='thin',color='000000'))

for i in range(9,42):
    for j in range(2,22):
        ws.cell(row=i, column=j).border = border

border = Border(left=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'))

ws.cell(row=42, column=2).border=border

border = Border(top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'))

for i in [42]:
    for j in range(3,16):
        ws.cell(row=i, column=j).border=border

border = Border(right=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'))

ws.cell(row=42,column=16).border=border

border = Border(left=Side(border_style='medium',color='000000'),
                right=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'))

for i in [42]:
    for j in range(17,22):
        ws.cell(row=i, column=j).border=border

ws['I42'].value='Total'
ws['R43'].value='OT rate'
ws['R44'].value='Deduct rate'
ws['H45'].value='Staff Signature'
ws['P45'].value='Approved Signature'

fill = PatternFill("solid", fgColor='FFFF99')
for i in range(8,42):
    for j in [4,15]:
        ws.cell(row=i, column=j).fill=fill

fill = PatternFill("solid", fgColor='3FF626')
for i in range(8,42):
    for j in [5,10,16,17,19]:
        ws.cell(row=i, column=j).fill=fill

fill = PatternFill("solid", fgColor='FFFF00')
for i in range(8,42):
    for j in [6,7,8,9]:
        ws.cell(row=i, column=j).fill=fill



wb.save('111.xlsx')
wb.close()

