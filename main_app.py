from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QMessageBox, QTableWidgetItem, QDialog, QFileDialog
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QDate, QTime, Qt, QThread, pyqtSignal

import sys
APP = QApplication(sys.argv)

DESKTOP = QApplication.desktop()

WIDTH = DESKTOP.width()

if WIDTH <= 1366:
    from ui.login import Ui_loginWindow
    from ui.mainwindow import Ui_MainWindow
    from ui.password import Ui_PassWindow
    from ui.timecard import Ui_TimeCard
    from ui.askforleave import Ui_AskForLeave
    from ui.otapplication import Ui_OTApplication
    from ui.bookmeetingroom import Ui_BookMeetingRoom
    from ui.approve_panel import Ui_ApprovePanel
    from ui.applylateclockin import Ui_ApplyLateClockIn
    from ui.forgetrecord import Ui_ForgetRecord

    from ui.main_admin import Ui_AdminMain
    from ui.staff_manage import Ui_StaffManage
    from ui.team_structure import Ui_TeamStructure
    from ui.login_pass import Ui_LoginPass
    from ui.calendar import Ui_CalendarSetting
    from ui.ot_sheet import Ui_OT_Sheet
    from ui.monitor import Ui_Monitor
    from ui.tip_window import Ui_Tipwindow

    W_LEAVE=831
    H_LEAVE=550
    W_OT=858
    H_OT=526

    W_APPROVE=1065
    H_APPROVE=672
else:
    from ui_highDPI.login import Ui_loginWindow
    from ui_highDPI.mainwindow import Ui_MainWindow
    from ui_highDPI.password import Ui_PassWindow
    from ui_highDPI.timecard import Ui_TimeCard
    from ui_highDPI.askforleave import Ui_AskForLeave
    from ui_highDPI.otapplication import Ui_OTApplication
    from ui_highDPI.bookmeetingroom import Ui_BookMeetingRoom
    from ui_highDPI.approve_panel import Ui_ApprovePanel
    from ui_highDPI.applylateclockin import Ui_ApplyLateClockIn
    from ui_highDPI.forgetrecord import Ui_ForgetRecord

    from ui_highDPI.main_admin import Ui_AdminMain
    from ui_highDPI.staff_manage import Ui_StaffManage
    from ui_highDPI.team_structure import Ui_TeamStructure
    from ui_highDPI.login_pass import Ui_LoginPass
    from ui_highDPI.calendar import Ui_CalendarSetting
    from ui_highDPI.ot_sheet import Ui_OT_Sheet
    from ui_highDPI.monitor import Ui_Monitor
    from ui_highDPI.tip_window import Ui_Tipwindow

    W_LEAVE = 1067
    H_LEAVE = 722
    W_OT = 1067
    H_OT = 701

    W_APPROVE = 1296
    H_APPROVE = 776

from modules.NTP_time import NTP_DateTime
from modules.Mail_Sender import MailSender
from modules.Functions import set_format

import datetime
import time
import csv
import os
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import pymysql
import ftplib
import subprocess

class Tipwindow(QWidget, Ui_Tipwindow):
    def __init__(self):
        super(Tipwindow, self).__init__()
        self.setupUi(self)

class Login_Thread(QThread):
    tipwindow_show = pyqtSignal()
    tipwindow_close = pyqtSignal()
    message_box = pyqtSignal(str, str, str)
    update_tiplabel = pyqtSignal(str)
    send_id = pyqtSignal(int)
    to_main_win = pyqtSignal(str)

    def __init__(self, lineEdit, lineEdit_2, comboBox):
        super(Login_Thread, self).__init__()
        self.lineEdit = lineEdit
        self.lineEdit_2 = lineEdit_2
        self.comboBox = comboBox

    def run(self):
        self.tipwindow_show.emit()
        self.update_tiplabel.emit('loading...')
        global DB
        try:
            DB.close()
        except:
            pass
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='akaganeHR')
        except pymysql.err.OperationalError:
            #QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            self.message_box.emit('critical', 'Network Error', 'Can not connect to the server, please check your network!')
            self.tipwindow_close.emit()
            return

        self.cursor_version = DB.cursor()
        SQL = """SELECT VERSION FROM version_control WHERE ID=%s"""
        self.cursor_version.execute(SQL, (1))
        results=self.cursor_version.fetchall()
        self.cursor_version.close()
        if results == ():
            #QMessageBox.information(self, 'Info' , 'Sorry, the system is under maintenance, please try it later...')
            self.message_box.emit('information', 'Info' , 'Sorry, the system is under maintenance, please try it later...')
            self.tipwindow_close.emit()
            return
        version=results[0][0]
        if version==None:
            #QMessageBox.information(self, 'Info', 'Sorry, the system and database is under maintenance, please try it later...')
            self.message_box.emit('information', 'Info', 'Sorry, the system and database is under maintenance, please try it later...')
            self.tipwindow_close.emit()
            return

        current_version = CURRENT_VER
        if current_version < version:
            #QMessageBox.information(self, 'Version Too Old', f'Sorry, the version you are using is too old, please update to version {version} first!')
            self.message_box.emit('information', 'Version Too Old', f'Sorry, the version you are using is too old, please update to version {version} first!')
            self.tipwindow_close.emit()
            return

        self.cursor = DB.cursor()

        self.id = int(str(self.lineEdit.text()).strip())
        #global ID
        #ID = self.id
        self.send_id.emit(self.id)
        self.password = self.lineEdit_2.text()
        if not str(self.id).isdigit():
            #QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            self.message_box.emit('warning', 'Warning', 'Wrong user ID!')
            self.cursor.close()
            self.tipwindow_close.emit()
            return

        sql = """SELECT PASSWORD, PRIORITY FROM login_pass WHERE ID=%s"""
        self.cursor.execute(sql, (self.id))
        pass_check = self.cursor.fetchall()
        if pass_check == ():
            #QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            self.message_box.emit('warning', 'Warning', 'Wrong user ID!')
            # print('Wrong user id!')
            self.cursor.close()
            self.tipwindow_close.emit()

            return
        if self.password == pass_check[0][0]:
            if self.comboBox.currentText() == 'Administrator':
                if pass_check[0][1] == 'admin':
                    #self.to_main_win(mode='admin')
                    self.to_main_win.emit('admin')
                else:
                    #QMessageBox.warning(self, 'warning',
                    #                    'Sorry, you do not have the administration authority, please select the "Normal" mode.')
                    self.message_box.emit('warning', 'Warning',
                                        'Sorry, you do not have the administration authority, please select the "Normal" mode.')

            elif self.comboBox.currentText() == 'HR Approver':
                if pass_check[0][1] == 'admin':
                    #self.to_main_win(mode='hr_approver')
                    self.to_main_win.emit('hr_approver')
                else:
                    #QMessageBox.warning(self, 'warning',
                    #                    'Sorry, you do not have the HR-approving authority, please select the "Normal" mode.')
                    self.message_box.emit('warning', 'Warning',
                                        'Sorry, you do not have the HR-approving authority, please select the "Normal" mode.')
            else:
                #self.to_main_win(mode='normal')
                self.to_main_win.emit('normal')

        else:
            #QMessageBox.warning(self, 'Warning', 'Wrong password!')
            self.message_box.emit('warning', 'Warning', 'Wrong password!')
            # print('Wrong password!')

        self.cursor.close()
        self.tipwindow_close.emit()

class loginWindow(QMainWindow, Ui_loginWindow):
    def __init__(self):
        super(loginWindow, self).__init__()
        self.setupUi(self)

        self.label_7.setText('Developed in 2020  Ver.3.1')
        self.label_2.setText('HR Information System V3.1')

        self.id = ''
        self.pushButton.clicked.connect(self.login)
        self.pushButton_2.clicked.connect(self.quit)
        #self.pushButton_2.clicked.connect(self.temp)
        self.pushButton_3.clicked.connect(self.upgrade_)

        self.init_default()

    def init_default(self):
        info_line = []
        try:
            with open('info.csv', 'r') as file:
                reader = csv.reader(file)
                for line in reader:
                    info_line.append(line)
            file.close()
        except:
            pass

        if info_line!=[]:
            if info_line[0][3]=='1':
                self.lineEdit.setText(str(info_line[0][0]))
                self.lineEdit_2.setText(str(info_line[0][1]))
                self.comboBox.setCurrentText(str(info_line[0][2]))
                self.checkBox.setChecked(True)

    def temp(self):
        self.pushButton_2.setAttribute(Qt.WA_UnderMouse, False)
        global DB
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='akaganeHR')
        except pymysql.err.OperationalError:
            QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            return

        self.cursor = DB.cursor()

        self.id = '3001'
        self.password = 'akt1765323'
        sql = """SELECT PASSWORD, PRIORITY FROM login_pass WHERE ID=%s"""
        self.cursor.execute(sql, (self.id))
        pass_check = self.cursor.fetchall()
        if pass_check == ():
            QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            # print('Wrong user id!')
            self.cursor.close()

            return
        if self.password == pass_check[0][0]:
            if self.comboBox.currentText() == 'Administrator':
                if pass_check[0][1] == 'admin':
                    self.to_main_win(mode='admin')
                else:
                    QMessageBox.warning(self, 'warning',
                                        'Sorry, you do not have the administration authority, please select the "Normal" mode.')

            else:
                self.to_main_win(mode='normal')

        else:
            QMessageBox.warning(self, 'Warning', 'Wrong password!')
            # print('Wrong password!')

        self.cursor.close()

    def upgrade_(self):
        self.pushButton_3.setAttribute(Qt.WA_UnderMouse, False)

        global DB
        try:
            DB.close()
        except:
            pass
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='akaganeHR')
        except pymysql.err.OperationalError:
            QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            return

        self.cursor_version = DB.cursor()
        SQL = """SELECT VERSION, CURRENT_VER FROM version_control WHERE ID=%s"""
        self.cursor_version.execute(SQL, (1))
        results = self.cursor_version.fetchall()
        self.cursor_version.close()
        if results == ():
            QMessageBox.information(self, 'Info', 'Sorry, the system is under maintenance, please try it later...')
            return
        version = results[0][0]
        if version == None:
            QMessageBox.information(self, 'Info',
                                    'Sorry, the system and database is under maintenance, please try it later...')
            return

        current_version = CURRENT_VER
        OFFICIAL_CURRENT_VER=results[0][1]
        if float(current_version) == float(OFFICIAL_CURRENT_VER):
            QMessageBox.information(self, 'Info',
                                    'No need to upgrade for now, the system version is already the newest!')
            return

        a=QMessageBox.question(self, 'Inquiry',
                             f'Your current version is V{current_version}, do you want to update to the newest version?')
        if a==QMessageBox.No:
            DB.close()
            return

        self.upgrade = UPGRADE()
        self.upgrade.finish_box.connect(self.finish_msgbox)
        self.upgrade.update_label.connect(Monitor.update_text)
        self.upgrade.update_progress.connect(Monitor.update_progressbar)
        self.upgrade.monitor_close.connect(Monitor.monitor_close_login)
        self.upgrade.monitor_open.connect(self.monitor_show)

        self.upgrade.start()
        self.monitor_show()

    def receive_id(self, id):
        self.id = str(id)
        global ID
        ID = self.id

    def message_box(self, type, title, text):
        if type == 'critical':
            QMessageBox.critical(self, title, text)
        elif type == 'information':
            QMessageBox.information(self, title, text)
        elif type == 'warning':
            QMessageBox.warning(self, title, text)

    def update_tiplabel(self, text):
        Tipwindow.label.setText(text)

    def tipwindow_show(self):
        Tipwindow.show()
        self.setEnabled(False)

    def tipwindow_close(self):
        Tipwindow.destroy()
        self.setEnabled(True)

    def monitor_show(self):
        Monitor.show()
        Monitor.initializing()
        self.hide()

    def finish_msgbox(self, title, text):
        QMessageBox.information(self, title, text)

    def login(self):
        self.pushButton.setAttribute(Qt.WA_UnderMouse, False)

        self.login_thread=Login_Thread(lineEdit=self.lineEdit, lineEdit_2=self.lineEdit_2, comboBox=self.comboBox)
        self.login_thread.tipwindow_show.connect(self.tipwindow_show)
        self.login_thread.tipwindow_close.connect(self.tipwindow_close)
        self.login_thread.message_box.connect(self.message_box)
        self.login_thread.update_tiplabel.connect(self.update_tiplabel)
        self.login_thread.send_id.connect(self.receive_id)
        self.login_thread.to_main_win.connect(self.to_main_win)


        self.login_thread.start()

        """global DB
        try:
            DB.close()
        except:
            pass
        try:
            DB = pymysql.connect(host='210.1.31.3',
                                 user='hr',
                                 port=3306,
                                 passwd='gwP6xTsA',
                                 db='akaganeHR')
        except pymysql.err.OperationalError:
            QMessageBox.critical(self, 'Network Error', 'Can not connect to the server, please check your network!')
            return

        self.cursor_version = DB.cursor()"""
        #SQL = """SELECT VERSION FROM version_control WHERE ID=%s"""
        """self.cursor_version.execute(SQL, (1))
        results=self.cursor_version.fetchall()
        self.cursor_version.close()
        if results == ():
            QMessageBox.information(self, 'Info' , 'Sorry, the system is under maintenance, please try it later...')
            return
        version=results[0][0]
        if version==None:
            QMessageBox.information(self, 'Info', 'Sorry, the system and database is under maintenance, please try it later...')
            return

        current_version = CURRENT_VER
        if current_version < version:
            QMessageBox.information(self, 'Version Too Old', f'Sorry, the version you are using is too old, please update to version {version} first!')
            return

        self.cursor = DB.cursor()

        self.id = str(self.lineEdit.text()).strip()
        self.password = self.lineEdit_2.text()
        if not self.id.isdigit():
            QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            self.cursor.close()

            return"""

        #sql = """SELECT PASSWORD, PRIORITY FROM login_pass WHERE ID=%s"""
        """self.cursor.execute(sql, (self.id))
        pass_check = self.cursor.fetchall()
        if pass_check == ():
            QMessageBox.warning(self, 'Warning', 'Wrong user ID!')
            # print('Wrong user id!')
            self.cursor.close()

            return
        if self.password == pass_check[0][0]:
            if self.comboBox.currentText() == 'Administrator':
                if pass_check[0][1] == 'admin':
                    self.to_main_win(mode='admin')
                else:
                    QMessageBox.warning(self, 'warning',
                                        'Sorry, you do not have the administration authority, please select the "Normal" mode.')

            elif self.comboBox.currentText() == 'HR Approver':
                if pass_check[0][1] == 'admin':
                    self.to_main_win(mode='hr_approver')
                else:
                    QMessageBox.warning(self, 'warning',
                                        'Sorry, you do not have the HR-approving authority, please select the "Normal" mode.')

            else:
                self.to_main_win(mode='normal')

        else:
            QMessageBox.warning(self, 'Warning', 'Wrong password!')
            # print('Wrong password!')

        self.cursor.close()"""

    def to_main_win(self, mode):
        global ID
        ID = self.id

        global HR_MODE
        if mode == 'hr_approver':
            HR_MODE=1
        else:
            HR_MODE=0


        global MainWindow
        MainWindow.show()
        # print(mode!='admin')
        if mode != 'admin':
            MainWindow.pushButton_5.setEnabled(False)
            MainWindow.pushButton_5.setStyleSheet("QPushButton\n"
                                                  "{\n"
                                                  "    /*字体为微软雅黑*/\n"
                                                  "    font-family:Microsoft Yahei;\n"
                                                  "    /*字体大小为20点*/\n"
                                                  "    font-size:15pt;\n"
                                                  "    /*字体颜色为白色*/    \n"
                                                  "    color:white;\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(133, 173, 173);\n"
                                                  "    /*边框圆角半径为8像素*/ \n"
                                                  "    border-radius:10px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮停留态*/\n"
                                                  "QPushButton:hover\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(58, 0, 175);\n"
                                                  "    padding-left:-3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:-3px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮按下态*/\n"
                                                  "QPushButton:pressed\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:rgb(57, 0, 122);\n"
                                                  "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                  "    padding-left:3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:3px;\n"
                                                  "}")
        else:
            MainWindow.pushButton_5.setEnabled(True)
            MainWindow.pushButton_5.setStyleSheet("QPushButton\n"
                                                  "{\n"
                                                  "    /*字体为微软雅黑*/\n"
                                                  "    font-family:Microsoft Yahei;\n"
                                                  "    /*字体大小为20点*/\n"
                                                  "    font-size:15pt;\n"
                                                  "    /*字体颜色为白色*/    \n"
                                                  "    color:white;\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:#C1272D;\n"
                                                  "    /*边框圆角半径为8像素*/ \n"
                                                  "    border-radius:10px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮停留态*/\n"
                                                  "QPushButton:hover\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:#9e1e24;\n"
                                                  "    padding-left:-3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:-3px;\n"
                                                  "}\n"
                                                  "\n"
                                                  "/*按钮按下态*/\n"
                                                  "QPushButton:pressed\n"
                                                  "{\n"
                                                  "    /*背景颜色*/  \n"
                                                  "    background-color:#CD5257;\n"
                                                  "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                  "    padding-left:3px;\n"
                                                  "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                  "    padding-top:3px;\n"
                                                  "}")
        MainWindow.init_db()

        global loginWindow
        loginWindow.destroy()
        TimeCard.initial()

        if self.checkBox.isChecked():
            with open('info.csv', 'w', newline='') as file:
                writer=csv.writer(file)
                writer.writerow([self.lineEdit.text(), self.lineEdit_2.text(),self.comboBox.currentText(), '1'])
            file.close()
        else:
            try:
                os.remove('info.csv')
            except:
                pass

    def quit(self):
        if DB != None:
            DB.close()
        sys.exit()

    def closeEvent(self, event):
        if DB != None:
            DB.close()
        sys.exit()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.pushButton_7.clicked.connect(self.quit)
        self.pushButton_6.clicked.connect(self.to_login)
        self.pushButton_8.clicked.connect(self.to_passchange)
        self.pushButton.clicked.connect(self.to_timecard)
        self.pushButton_2.clicked.connect(self.to_askforleave)
        self.pushButton_3.clicked.connect(self.to_otapplication)
        self.pushButton_4.clicked.connect(self.to_bookmeetingroom)
        self.pushButton_9.clicked.connect(self.to_approvepanel)
        self.pushButton_5.clicked.connect(self.to_main_admin)

    def to_main_admin(self):
        self.pushButton_5.setAttribute(Qt.WA_UnderMouse, False)
        AdminMain.show()
        MainWindow.setVisible(False)

    def to_approvepanel(self):
        self.pushButton_9.setAttribute(Qt.WA_UnderMouse,False)
        ApprovePanel.show()
        ApprovePanel.initializing()
        MainWindow.destroy()

    def to_bookmeetingroom(self):
        self.pushButton_4.setAttribute(Qt.WA_UnderMouse, False)
        BookMeetingRoom.show()
        BookMeetingRoom.initializing()
        BookMeetingRoom.calendarWidget.setFocus()
        MainWindow.destroy()

    def to_otapplication(self):
        self.pushButton_3.setAttribute(Qt.WA_UnderMouse, False)
        OTApplication.show()
        OTApplication.initializing()
        MainWindow.destroy()

    def to_askforleave(self):
        self.pushButton_2.setAttribute(Qt.WA_UnderMouse, False)
        AskForLeave.show()
        AskForLeave.initializing()
        MainWindow.destroy()

    def to_timecard(self):
        self.pushButton.setAttribute(Qt.WA_UnderMouse, False)
        MainWindow.destroy()
        TimeCard.show()
        TimeCard.startTimer()
        TimeCard.calendarWidget.setFocus()

    def to_passchange(self):
        self.pushButton_8.setAttribute(Qt.WA_UnderMouse, False)
        PassWindow.show()
        PassWindow.initializing()
        MainWindow.setEnabled(False)

    def to_login(self):
        self.pushButton_6.setAttribute(Qt.WA_UnderMouse, False)
        MainWindow.destroy()
        TimeCard.timerStop()
        loginWindow.show()

    def quit(self):
        a = QMessageBox.question(self, 'Query', 'Are you sure to quit the system?', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            return

        # print(12345)
        if DB != None:
            DB.close()
        sys.exit()

    def init_db(self):
        self.cursor = DB.cursor()
        sql = """SELECT NAME, POSITION FROM team_stru WHERE ID=%s"""
        self.cursor.execute(sql, (ID))
        res = self.cursor.fetchall()
        self.name = res[0][0]
        self.label.setText(f'Hi, {self.name}, welcome!')
        self.position = res[0][1]

        if HR_MODE:
            self.label_4.setText('(HR staff only)')
            self.pushButton_9.setText('HR Approvement')
            self.pushButton_9.setEnabled(True)
            self.pushButton_9.setStyleSheet("QPushButton\n"
                                            "{\n"
                                            "    /*字体为微软雅黑*/\n"
                                            "    font-family:Microsoft Yahei;\n"
                                            "    /*字体大小为20点*/\n"
                                            "    font-size:15pt;\n"
                                            "    /*字体颜色为白色*/    \n"
                                            "    color:white;\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:#BF3919;\n"
                                            "    /*边框圆角半径为8像素*/ \n"
                                            "    border-radius:10px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮停留态*/\n"
                                            "QPushButton:hover\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:#9C3014;\n"
                                            "    padding-left:-3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:-3px;\n"
                                            "}\n"
                                            "\n"
                                            "/*按钮按下态*/\n"
                                            "QPushButton:pressed\n"
                                            "{\n"
                                            "    /*背景颜色*/  \n"
                                            "    background-color:#8D2B12;\n"
                                            "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                            "    padding-left:3px;\n"
                                            "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                            "    padding-top:3px;\n"
                                            "}")
        else:
            self.label_4.setText('(Group leader/MD/DM only)')
            if self.position.strip() != 'Leader' and self.position.strip() != 'DM' and self.position.strip() != 'MD':
                # print(111)
                self.pushButton_9.setEnabled(False)
                self.pushButton_9.setStyleSheet("QPushButton\n"
                                                "{\n"
                                                "    /*字体为微软雅黑*/\n"
                                                "    font-family:Microsoft Yahei;\n"
                                                "    /*字体大小为20点*/\n"
                                                "    font-size:15pt;\n"
                                                "    /*字体颜色为白色*/    \n"
                                                "    color:white;\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:rgb(133, 173, 173);\n"
                                                "    /*边框圆角半径为8像素*/ \n"
                                                "    border-radius:10px;\n"
                                                "}\n"
                                                "\n"
                                                "/*按钮停留态*/\n"
                                                "QPushButton:hover\n"
                                                "{\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:rgb(58, 0, 175);\n"
                                                "    padding-left:-3px;\n"
                                                "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                "    padding-top:-3px;\n"
                                                "}\n"
                                                "\n"
                                                "/*按钮按下态*/\n"
                                                "QPushButton:pressed\n"
                                                "{\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:rgb(57, 0, 122);\n"
                                                "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                "    padding-left:3px;\n"
                                                "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                "    padding-top:3px;\n"
                                                "}")
            else:
                self.pushButton_9.setEnabled(True)
                self.pushButton_9.setStyleSheet("QPushButton\n"
                                                "{\n"
                                                "    /*字体为微软雅黑*/\n"
                                                "    font-family:Microsoft Yahei;\n"
                                                "    /*字体大小为20点*/\n"
                                                "    font-size:15pt;\n"
                                                "    /*字体颜色为白色*/    \n"
                                                "    color:white;\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:#C1272D;\n"
                                                "    /*边框圆角半径为8像素*/ \n"
                                                "    border-radius:10px;\n"
                                                "}\n"
                                                "\n"
                                                "/*按钮停留态*/\n"
                                                "QPushButton:hover\n"
                                                "{\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:#9e1e24;\n"
                                                "    padding-left:-3px;\n"
                                                "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                "    padding-top:-3px;\n"
                                                "}\n"
                                                "\n"
                                                "/*按钮按下态*/\n"
                                                "QPushButton:pressed\n"
                                                "{\n"
                                                "    /*背景颜色*/  \n"
                                                "    background-color:#CD5257;\n"
                                                "    /*左内边距为3像素，让按下时字向右移动3像素*/  \n"
                                                "    padding-left:3px;\n"
                                                "    /*上内边距为3像素，让按下时字向下移动3像素*/  \n"
                                                "    padding-top:3px;\n"
                                                "}")

        self.cursor.close()

    def closeEvent(self, event):
        a = QMessageBox.question(self, 'Query', 'Are you sure to quit the system?', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            event.ignore()
            return

        mailsender.yag_server.close()

        self.cursor.close()
        if DB != None:
            DB.close()
        sys.exit()

class NTP_Get_DateTime(QThread):
    tipwindow_show = pyqtSignal()
    tipwindow_close = pyqtSignal()
    update_tiplabel = pyqtSignal(str)
    time_signal = pyqtSignal(datetime.datetime)

    def __init__(self):
        super(NTP_Get_DateTime, self).__init__()
        self.NTP = NTP_DateTime()

    def run(self):
        self.update_tiplabel.emit('Synchronizing NTP time...')
        self.tipwindow_show.emit()
        t0 = self.NTP.get_datetime()
        self.time_signal.emit(t0)
        self.tipwindow_close.emit()

class TimeCard(QMainWindow, Ui_TimeCard):

    def __init__(self):
        super(TimeCard, self).__init__()
        self.setupUi(self)

        self.textEdit.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_2.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_3.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_4.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_5.setFocusPolicy(QtCore.Qt.NoFocus)
        self.textEdit_6.setEnabled(False)
        self.textEdit_7.setEnabled(False)
        self.textEdit_8.setEnabled(False)
        self.textEdit_9.setEnabled(False)

        self.NTP = NTP_DateTime()
        self.pushButton_6.clicked.connect(self.quit)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.showtime)

        self.pushButton.clicked.connect(self.clock_in)
        self.pushButton_2.clicked.connect(self.clock_out)
        self.pushButton_5.clicked.connect(self.cancel_clock_out)
        self.calendarWidget.clicked.connect(self.select_date)
        #self.calendarWidget.clicked.connect(self.startTimer)
        self.pushButton_7.clicked.connect(self.to_applylateclockin)
        self.pushButton_3.clicked.connect(self.to_forgetrecord)
        self.pushButton_4.clicked.connect(self.export_excel)

        self.ntp_thread = NTP_Get_DateTime()
        self.ntp_thread.tipwindow_show.connect(self.tipwindow_show)
        self.ntp_thread.tipwindow_close.connect(self.tipwindow_close)
        self.ntp_thread.update_tiplabel.connect(self.update_tiplabel)
        self.ntp_thread.time_signal.connect(self.get_t0)
        self.ntp_thread.finished.connect(self.ntp_thread_finished)

        self.t0 = datetime.datetime.now()
        self.t = datetime.datetime.now()

    def export_excel(self):
        self.calendarWidget.setFocus()
        yearmonth = str(self.calendarWidget.yearShown()) + "%02d" % self.calendarWidget.monthShown()

        self.cursor_exc=DB.cursor()
        SQL="""SELECT * FROM time_card WHERE SERIAL>%s and SERIAL<%s"""
        try:
            self.cursor_exc.execute(SQL, (int(str(ID)+yearmonth+'00'), int(str(ID)+yearmonth+'32') ))
        except:
            reconnect_DB(self)
            self.cursor_exc = DB.cursor()
            self.cursor_exc.execute(SQL, (int(str(ID) + yearmonth + '00'), int(str(ID) + yearmonth + '32')))

        res=self.cursor_exc.fetchall()
        self.cursor_exc.close()
        #print(res)
        if res==():
            QMessageBox.information(self, 'Empty Record', f'The time card record in the year-month you selected({yearmonth}) is empty!')
            return

        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s timecard({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        #print(a[0])
        wb=xl.Workbook()
        ws=wb.active
        headers=['Series No', 'User ID', 'Clock In', 'Clock Out', 'Out 1', 'In 1', 'Out 2', 'In 2', 'Day Lag', 'Actual Work Time', 'Over Time', 'Approved OT']
        ws.append(headers)
        for each_line in res:
            line_head=each_line
            data_set=calculate_worktime(data_line=each_line)
            worktime=data_set[0]
            overtime=data_set[1]
            approved_ot=data_set[2]
            line_end=[worktime, overtime, approved_ot]
            line_full=list(line_head)+line_end
            ws.append(line_full)
            #print(line_full)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def to_forgetrecord(self):
        self.pushButton_3.setAttribute(Qt.WA_UnderMouse, False)
        ForgetRecord.show()
        ForgetRecord.initialize()
        TimeCard.destroy()

    def to_applylateclockin(self):
        self.pushButton_7.setAttribute(Qt.WA_UnderMouse, False)
        ApplyLateClockIn.show()
        ApplyLateClockIn.initialize()
        TimeCard.destroy()

    def select_date(self):
        qdt = self.calendarWidget.selectedDate()
        dt_for_query=datetime.datetime.strptime(qdt.toString('dd/MM/yyyy'),'%d/%m/%Y')
        sql="""SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
        self.cursor_queryhours=DB.cursor()
        try:
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        except:
            reconnect_DB(self)
            self.cursor_queryhours=DB.cursor()
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        res=self.cursor_queryhours.fetchall()
        self.cursor_queryhours.close()
        if res==():
            self.textEdit_5.setText('0')
        else:
            ot_hours=str(res[0][0])
            self.textEdit_5.setText(ot_hours)

        dt = qdt.toString('yyyyMMdd')
        user_id = str(ID)
        serial = user_id + dt
        sql = """SELECT CLOCK_IN, CLOCK_OUT, DAY_LAG, OUT_1, IN_1, OUT_2, IN_2 FROM time_card WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except:
            self.cursor = DB.cursor()
            try:
                self.cursor.execute(sql, (serial))
            except:
                reconnect_DB(self)
                self.cursor=DB.cursor()
                self.cursor.execute(sql, (serial))

        result = self.cursor.fetchall()
        if result == ():
            self.textEdit.setText('-')
            self.textEdit_2.setText('-')
            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')

            self.textEdit_6.setText('-')
            self.textEdit_7.setText('-')
            self.textEdit_8.setText('-')
            self.textEdit_9.setText('-')
            return

        day_lag = result[0][2]

        clock_out = result[0][1]

        if clock_out == None:
            clock_out = '-'

        clock_in = result[0][0]

        if clock_in == None:
            clock_in = '-'

        try:
            out1 = result[0][3]
        except:
            out1 = '-'
        if out1 == None:
            out1 = '-'
        try:
            in1 = result[0][4]
        except:
            in1 = '-'
        if in1 == None:
            in1 = '-'
        try:
            out2 = result[0][5]
        except:
            out2 = '-'
        if out2 == None:
            out2 = '-'
        try:
            in2 = result[0][6]
        except:
            in2 = '-'
        if in2 == None:
            in2 = '-'

        if clock_out == '-' or clock_in=='-':
            work_time = '-'
            over_time = '-'
        # ---------------------计算实际作业时间
        else:
            if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                #1
                if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    work_time = (clock_out - clock_in).seconds
                #2
                elif int(clock_in.strftime('%Y%m%d') + '1130')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                #3
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 3600
                #4
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                    work_time -= 3600
                #5
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 4500

            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                #6
                if int(clock_in.strftime('%Y%m%d') + '1130')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230'):
                    work_time= 0
                #7
                elif int(clock_in.strftime('%Y%m%d') + '1230')<int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                              '%Y%m%d%H%M')).seconds
                #8
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time=(datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                              '%Y%m%d%H%M')).seconds
                #9
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                    work_time-=900

            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                #10
                if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                #11
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - clock_in).seconds
                #12
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time-=900

            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                #13
                if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                    work_time=0
                #14
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                        '%Y%m%d%H%M')).seconds
            else:
                work_time = (clock_out - clock_in).seconds

            # ------------------计算离岗累计时间
            space_time1 = 0
            space_time2 = 0
            if out1 != '-' and in1 != '-':
                # 1
                if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time1 = (in1 - out1).seconds
                # 2
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 3
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds - 3600
                # 4
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds - 3600
                # 5
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 3600 - 900
                # 6
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = 0
                # 7
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds
                # 11
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 12
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 900
                # 13
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = 0
                # 14
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time1 = (in1 - out1).seconds

            # -----------------out2 in2
            if out2 != '-' and in2 != '-':
                # 1
                if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time2 = (in2 - out2).seconds
                # 2
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 3
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds - 3600
                # 4
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds - 3600
                # 5
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 3600 - 900
                # 6
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = 0
                # 7
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds
                # 11
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 12
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 900
                # 13
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = 0
                # 14
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time2 = (in2 - out2).seconds

            work_time -= space_time1
            work_time -= space_time2
            # --------------------------如过午夜加上一天
            #if day_lag == 1:
                #work_time += 86400   删除逻辑错误代码

            over_time = work_time - 28800
            if over_time < 0:
                over_time = 0

        if work_time != '-':
            work_time = round(int(work_time) / 3600, 2)
        if over_time != '-':
            over_time = round(int(over_time) / 3600, 2)
        if clock_in != '-':
            self.textEdit.setText(clock_in.strftime('%H:%M:%S'))
        else:
            self.textEdit.setText(clock_in)
        if clock_out != '-':
            self.textEdit_2.setText(clock_out.strftime('%H:%M:%S'))
        else:
            self.textEdit_2.setText(clock_out)
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))
        if out1 != '-':
            self.textEdit_6.setText(out1.strftime('%H:%M:%S'))
        else:
            self.textEdit_6.setText(out1)
        if in1 != '-':
            self.textEdit_7.setText(in1.strftime('%H:%M:%S'))
        else:
            self.textEdit_7.setText(in1)

        if out2 != '-':
            self.textEdit_8.setText(out2.strftime('%H:%M:%S'))
        else:
            self.textEdit_8.setText(out2)
        if in2 != '-':
            self.textEdit_9.setText(in2.strftime('%H:%M:%S'))
        else:
            self.textEdit_9.setText(in2)

    def refresh_panel(self):
        dt = self.t.strftime('%Y%m%d')
        dt_for_query = datetime.datetime.strptime(self.t.strftime('%d/%m/%Y'),'%d/%m/%Y')
        sql = """SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
        self.cursor_queryhours = DB.cursor()
        try:
            self.cursor_queryhours.execute(sql, (ID, dt_for_query, '9999'))
        except:
            reconnect_DB(self)
            self.cursor_queryhours = DB.cursor()
            self.cursor_queryhours.execute(sql, (ID, dt_for_query))
        res = self.cursor_queryhours.fetchall()
        self.cursor_queryhours.close()
        if res == ():
            self.textEdit_5.setText('0')
        else:
            ot_hours = str(res[0][0])
            self.textEdit_5.setText(ot_hours)

        user_id = str(ID)
        serial = user_id + dt
        sql = """SELECT CLOCK_IN, CLOCK_OUT, DAY_LAG, OUT_1, IN_1, OUT_2, IN_2 FROM time_card WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except:
            self.cursor = DB.cursor()
            try:
                self.cursor.execute(sql, (serial))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor=DB.cursor()
                self.cursor.execute(sql, (serial))

        result = self.cursor.fetchall()
        if result == ():
            self.textEdit.setText('-')
            self.textEdit_2.setText('-')
            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')

            self.textEdit_6.setText('-')
            self.textEdit_7.setText('-')
            self.textEdit_8.setText('-')
            self.textEdit_9.setText('-')
            return

        day_lag = result[0][2]

        clock_out = result[0][1]
        if clock_out == None:
            clock_out = '-'

        clock_in = result[0][0]
        try:
            out1 = result[0][3]
        except:
            out1 = '-'
        if out1 == None:
            out1 = '-'
        try:
            in1 = result[0][4]
        except:
            in1='-'
        if in1 == None:
            in1 = '-'
        try:
            out2 = result[0][5]
        except:
            out2='-'
        if out2 == None:
            out2 = '-'
        try:
            in2 = result[0][6]
        except:
            in2='-'
        if in2 == None:
            in2 = '-'

        if clock_out == '-':
            work_time = '-'
            over_time = '-'
        else:
            if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                # 1
                if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    work_time = (clock_out - clock_in).seconds
                # 2
                elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 3
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 3600
                # 4
                elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                    work_time -= 3600
                # 5
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 4500

            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                # 6
                if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1230'):
                    work_time = 0
                # 7
                elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                # 8
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230',
                        '%Y%m%d%H%M')).seconds
                # 9
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                        '%Y%m%d%H%M')).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                # 10
                if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1500'):
                    work_time = (clock_out - clock_in).seconds
                # 11
                elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                            '%Y%m%d%H%M') - clock_in).seconds
                # 12
                else:
                    work_time = (clock_out - clock_in).seconds
                    work_time -= 900

            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                # 13
                if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                        clock_in.strftime('%Y%m%d') + '1515'):
                    work_time = 0
                # 14
                else:
                    work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                        '%Y%m%d%H%M')).seconds
            else:
                work_time = (clock_out - clock_in).seconds

            # ------------------计算离岗累计时间
            space_time1 = 0
            space_time2 = 0
            if out1 != '-' and in1 != '-':
                # 1
                if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time1 = (in1 - out1).seconds
                # 2
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 3
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds - 3600
                # 4
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds - 3600
                # 5
                elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 3600 - 900
                # 6
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time1 = 0
                # 7
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time1 = (in1 - out1).seconds
                # 11
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out1).seconds
                # 12
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - out1).seconds - 900
                # 13
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time1 = 0
                # 14
                elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time1 = (in1 - out1).seconds

            # -----------------out2 in2
            if out2 != '-' and in2 != '-':
                # 1
                if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                    space_time2 = (in2 - out2).seconds
                # 2
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 3
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds - 3600
                # 4
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds - 3600
                # 5
                elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 3600 - 900
                # 6
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                    space_time2 = 0
                # 7
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                # 8
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - datetime.datetime.strptime(
                        clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
                # 9
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds - 900
                # 10
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                    space_time2 = (in2 - out2).seconds
                # 11
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                              '%Y%m%d%H%M') - out2).seconds
                # 12
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - out2).seconds - 900
                # 13
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                        int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                    space_time2 = 0
                # 14
                elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                        out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                    space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
                # 15
                else:
                    space_time2 = (in2 - out2).seconds

            work_time -= space_time1
            work_time -= space_time2
            # --------------------------如过午夜加上一天
            #if day_lag == 1:
                #work_time += 86400   删除逻辑错误代码

            over_time = work_time - 28800
            if over_time < 0:
                over_time = 0

        if work_time!='-':
            work_time = round(int(work_time) / 3600, 2)
        if over_time!='-':
            over_time = round(int(over_time) / 3600, 2)

        self.textEdit.setText(clock_in.strftime('%H:%M:%S'))
        if clock_out != '-':
            self.textEdit_2.setText(clock_out.strftime('%H:%M:%S'))
        else:
            self.textEdit_2.setText(clock_out)
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))

        if out1 != '-':
            self.textEdit_6.setText(out1.strftime('%H:%M:%S'))
        else:
            self.textEdit_6.setText(out1)
        if in1 != '-':
            self.textEdit_7.setText(in1.strftime('%H:%M:%S'))
        else:
            self.textEdit_7.setText(in1)

        if out2 != '-':
            self.textEdit_8.setText(out2.strftime('%H:%M:%S'))
        else:
            self.textEdit_8.setText(out2)
        if in2 != '-':
            self.textEdit_9.setText(in2.strftime('%H:%M:%S'))
        else:
            self.textEdit_9.setText(in2)

        self.calendarWidget.setFocus()

    def clock_in(self):
        a=QMessageBox.question(self, 'Confirmation', 'Are you sure to clock in?')
        if a==QMessageBox.No:
            self.calendarWidget.setFocus()
            return
        user_id = str(ID)
        serial = user_id + self.t.strftime('%Y%m%d')
        in_time = self.t
        # serial = user_id + (self.t-datetime.timedelta(days=1)).strftime('%Y%m%d')
        # in_time = self.t-datetime.timedelta(days=1)
        self.cursor_clock = DB.cursor()
        sql = """INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN) VALUES
        (%s,%s,%s)"""
        try:
            try:
                self.cursor_clock.execute(sql, (serial, user_id, in_time))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_clock=DB.cursor()
                self.cursor_clock.execute(sql, (serial, user_id, in_time))
                DB.commit()
            QMessageBox.information(self, 'Info', 'Clock-in completed!')
        except pymysql.err.IntegrityError:
            DB.rollback()
            sql = """SELECT CLOCK_OUT, OUT_1, OUT_2 FROM time_card WHERE SERIAL=%s"""
            self.cursor_clock.execute(sql, (serial))
            res = self.cursor_clock.fetchall()
            if res[0][0] == None:  # CLOCK_OUT==None
                QMessageBox.critical(self, 'Info', 'Sorry, you have already clocked in.')
            else:
                if res[0][1] == None:  # OUT_1==None
                    sql = """UPDATE time_card SET OUT_1=%s, CLOCK_OUT=%s, IN_1=%s WHERE SERIAL=%s"""
                    self.cursor_clock.execute(sql, (res[0][0], None, in_time, serial))
                    DB.commit()
                    QMessageBox.information(self, 'Info', 'Clock-in completed!')
                elif res[0][2] == None:  # OUT_2=None
                    sql = """UPDATE time_card SET OUT_2=%s, CLOCK_OUT=%s, IN_2=%s WHERE SERIAL=%s"""
                    self.cursor_clock.execute(sql, (res[0][0] ,None, in_time, serial))
                    DB.commit()
                    QMessageBox.information(self, 'Info', 'Clock-in completed!')
                else:
                    QMessageBox.critical(self, 'Warning', 'Sorry, you can not clock in more than 3 times per day.')

        self.cursor_clock.close()
        self.refresh_panel()

    def clock_out(self):
        a = QMessageBox.question(self, 'Confirmation', 'Are you sure to clock out?')
        if a == QMessageBox.No:
            self.calendarWidget.setFocus()
            return
        user_id = str(ID)
        if not self.checkBox.isChecked():
            serial = user_id + self.t.strftime('%Y%m%d')
            dt = self.t.strftime('%d/%m/%Y')
            day_lag = None
        else:
            serial = user_id + (self.t - datetime.timedelta(days=1)).strftime('%Y%m%d')
            dt = (self.t - datetime.timedelta(days=1)).strftime('%d/%m/%Y')
            day_lag = 1

        out_time = self.t
        self.cursor_clock = DB.cursor()
        sql = """SELECT SERIAL, CLOCK_OUT FROM time_card WHERE SERIAL=%s"""
        try:
            a = self.cursor_clock.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_clock=DB.cursor()
            a = self.cursor_clock.execute(sql, (serial))

        if a == 0:
            QMessageBox.critical(self, 'Warning', f'Clock-out failed! You did not clocked in for today({dt}).')
            self.cursor_clock.close()
            self.calendarWidget.setFocus()
            return

        temp = self.cursor_clock.fetchall()[0][1]
        if temp != None:
            QMessageBox.critical(self, 'Warning',
                                 f'You have already clocked out, if you want to clock out again, please clock in first, or cancel clock-out first.')
            self.cursor_clock.close()
            self.calendarWidget.setFocus()
            return

        sql = """UPDATE time_card SET CLOCK_OUT=%s, DAY_LAG=%s WHERE SERIAL=%s"""
        self.cursor_clock.execute(sql, (out_time, day_lag, serial))
        DB.commit()
        QMessageBox.information(self, 'Info', 'Clock-out completed!')
        self.cursor_clock.close()
        self.refresh_panel()

    def cancel_clock_out(self):
        user_id = str(ID)
        if not self.checkBox.isChecked():
            serial = user_id + self.t.strftime('%Y%m%d')
            dt = self.t.strftime('%d/%m/%Y')

        else:
            serial = user_id + (self.t - datetime.timedelta(days=1)).strftime('%Y%m%d')
            dt = (self.t - datetime.timedelta(days=1)).strftime('%d/%m/%Y')

        sql = """SELECT CLOCK_OUT FROM time_card WHERE SERIAL=%s"""
        try:
            a = self.cursor.execute(sql, (serial))
        except pymysql.err.ProgrammingError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            a = self.cursor.execute(sql, (serial))
        if a == 0:
            QMessageBox.critical(self, 'Warning', f'You did not clocked out for the latest time.')
            self.calendarWidget.setFocus()
            return

        temp = self.cursor.fetchall()[0][0]
        if temp == None:
            QMessageBox.critical(self, 'Warning', f'You did not clocked out for the latest time.')
            self.calendarWidget.setFocus()
            return

        target_time = temp.strftime('%d/%m/%Y %H:%M:%S')
        res = QMessageBox.question(self, 'Query', f'Are you sure to cancel the clock-out record: {target_time}?')
        if res == QMessageBox.No:
            self.calendarWidget.setFocus()
            return

        sql = """UPDATE time_card SET CLOCK_OUT=%s, DAY_LAG=%s WHERE SERIAL=%s"""
        self.cursor.execute(sql, (None, None, serial))
        DB.commit()
        QMessageBox.information(self, 'Info',
                                f'The latest clock-out record({target_time}) has been canceled successfully')
        self.refresh_panel()

    def initial(self):
        self.label_2.setText(f'Name: {MainWindow.name}  Staff ID:{ID}')
        self.cursor = DB.cursor()
        self.startTimer()
        self.refresh_panel()

    def timerStop(self):
        self.timer.stop()

    def showtime(self):
        if (self.t - self.t0).seconds == 600:
            #self.t0 = self.NTP.get_datetime()
            self.ntp_thread.start()
            #self.t = self.t0
            # print('Reset time!', self.t0.strftime('%H:%M:%S'))
        self.t += datetime.timedelta(seconds=1)
        _time = self.t.strftime('%H:%M:%S')
        # print(_time)
        self.lcdNumber.display(_time)
        _date = self.t.strftime('%d/%m/%Y')
        self.label_9.setText(_date)

    def startTimer(self):
        #self.t0 = self.NTP.get_datetime()
        self.ntp_thread.start()
        #self.timer.start(1000)

    def get_t0(self, t0):
        self.t0 = t0

    def ntp_thread_finished(self):
        self.t = self.t0
        self.timer.start(1000)

    def update_tiplabel(self, text):
        Tipwindow.label.setText(text)

    def tipwindow_show(self):
        Tipwindow.show()

    def tipwindow_close(self):
        Tipwindow.destroy()

    def quit(self):
        TimeCard.close()

    def closeEvent(self, event):
        self.cursor.close()
        MainWindow.show()


class AskForLeave(QMainWindow, Ui_AskForLeave):
    def __init__(self):
        super(AskForLeave, self).__init__()
        self.setupUi(self)

        self.dateEdit_5.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_2.setDate(QDate.currentDate())
        self.dateEdit_7.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit_6.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_6.setCurrentText(month)

        self.geoTab = self.tabWidget.geometry()
        self.geoTable = self.tableWidget.geometry()
        self.geoBTN7 = self.pushButton_7.geometry()
        self.geoBTN16 = self.pushButton_16.geometry()
        self.geoBTN2 = self.pushButton_2.geometry()
        self.geoBTN4 = self.pushButton_4.geometry()
        self.geoFrame = self.frame.geometry()

        self.pushButton_4.clicked.connect(self.quit)
        self.checkBox.clicked.connect(self.is_multiple)
        self.pushButton.clicked.connect(self.request_per_leave)
        self.checkBox_2.clicked.connect(self.is_query)
        self.dateEdit_7.dateChanged.connect(self.show_IDs)
        self.comboBox_2.currentTextChanged.connect(self.input_id)
        self.lineEdit_2.textChanged.connect(self.refresh_data)
        self.pushButton_5.clicked.connect(self.del_single_request)
        self.pushButton_3.clicked.connect(self.request_go_home)
        self.checkBox_4.clicked.connect(self.is_query_hometown)
        self.lineEdit_3.textChanged.connect(self.get_id_hometown)
        self.comboBox_3.currentTextChanged.connect(self.refresh_hometown)
        self.pushButton_6.clicked.connect(self.del_hometown_request)
        self.pushButton_16.clicked.connect(self.download_data)
        self.pushButton_7.clicked.connect(self.cancel_request_on_table)
        self.pushButton_2.clicked.connect(self.export_excel)

        self.dateEdit_6.dateTimeChanged.connect(self.download_data)
        self.comboBox_6.currentTextChanged.connect(self.download_data)
        self.pushButton_16.setVisible(False)

        self.tabWidget.removeTab(self.tabWidget.indexOf(self.tab_2))

        self.tabWidget.currentChanged.connect(self.tableview)

    def export_excel(self):
        try:
            first_cell=self.tableWidget.item(0, 0).text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Table', 'Warning: The table is empty, no need to export the excel file!')
            return

        yearmonth=self.dateEdit_6.text()+self.comboBox_6.currentText()
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s Leave Application({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return
        excel_path = a[0]

        data_frame=[]
        for i in range(self.tableWidget.rowCount()):
            data_line=[]
            for j in range(self.tableWidget.columnCount()):
                data_line.append(self.tableWidget.item(i, j).text())
            data_frame.append(data_line)


        wb=xl.Workbook()
        ws=wb.active
        ws.append(['Request ID', 'Staff ID', 'Staff Name', 'Type', 'Request Date-Time', 'Request Date',
                   'Leave From', 'Time to Start', 'Leave Until', 'Time to End', 'If Single Day', 'Duration','Remarks',
                   'Leader','DM','HR','MD'])  #Updated on 12/9/2023, exchange the positions of "HR" and "MD"

        for line in data_frame:
            ws.append(line)

        try:
            wb.save(filename=excel_path)
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info','Excel file has been exported successfully!')

    def tableview(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        # print(self.tabWidget.currentIndex())
        if self.tabWidget.currentIndex() == 1:
            width0 = W_LEAVE
            height0 = H_LEAVE
            self.desktop = QApplication.desktop()
            screen_count=self.desktop.screenCount() #2 screen version updating
            geometry = self.desktop.geometry()
            if screen_count==1 or screen_count==2:  #2 screen version updating
                width1 = geometry.width()           #2 screen version updating
            else:                                   #2 screen version updating
                width1 = geometry.width()/2         #2 screen version updating
            height1 = geometry.height()
            # print(width0, height0, width1, height1)
            x = width1 - width0
            y = height1 - height0 - 100

            self.setMaximumSize(16777215, 16777215)
            self.showMaximized()

            move_widgets_for_frame(self.tabWidget, x, y)
            move_widgets_for_frame(self.tableWidget, x, y)
            move_widgets_only_vertical(self.pushButton_7, x, y)
            move_widgets_only_horizontal(self.pushButton_16, x, y)
            move_widgets_only_vertical(self.pushButton_2, x, y)
            move_widgets_both(self.pushButton_4, x, y)
            extend_widgets_horizontal(self.frame, x, y)
        else:
            self.tabWidget.setGeometry(self.geoTab.x(), self.geoTab.y(), self.geoTab.width(), self.geoTab.height())
            self.tableWidget.setGeometry(self.geoTable.x(), self.geoTable.y(), self.geoTable.width(),
                                         self.geoTable.height())
            self.pushButton_7.setGeometry(self.geoBTN7.x(), self.geoBTN7.y(), self.geoBTN7.width(),
                                          self.geoBTN7.height())
            self.pushButton_16.setGeometry(self.geoBTN16.x(), self.geoBTN16.y(), self.geoBTN16.width(),
                                           self.geoBTN16.height())
            self.pushButton_4.setGeometry(self.geoBTN4.x(), self.geoBTN4.y(), self.geoBTN4.width(),
                                          self.geoBTN4.height())
            self.pushButton_2.setGeometry(self.geoBTN2)
            self.frame.setGeometry(self.geoFrame)
            self.showNormal()
            self.setMaximumSize(W_LEAVE, H_LEAVE)

    def initializing(self):
        self.dateEdite_style_highlight = "QDateEdit{ font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 138, 163);}"
        self.comboBox_style_highlight = "QComboBox{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:9pt;background-color:rgb(255, 138, 163)}"
        self.textEdit_style_highlight = "QTextEdit{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 138, 163);}"
        self.radio_style_highlight = "QRadioButton{font-family:Microsoft Yahei;font-size:12pt;color:rgb(74, 74, 74); background-color:rgb(255, 138, 163);}"
        self.lineEdite_style_highlight = "QLineEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:black;   font-size:12pt;	background-color:rgb(200, 200, 200);}QLineEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QLineEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"

        self.dateEdite_style_normal = "QDateEdit{font-family:Microsoft Yahei;color:rgb(74, 74, 74);font-size:12pt;background-color:rgb(255, 255, 255);}QLineEdit:hover{background-color:rgb(229, 242, 255);}QLineEdit:focus{background-color:rgb(229, 242, 255);}"
        self.comboBox_style_normal = "QComboBox    {font-family:Microsoft Yahei;    /*字体大小为20点*/	color:rgb(74, 74, 74);   font-size:9pt;}QComboBox:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QComboBox:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"
        self.textEdit_style_normal = "QTextEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTextEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QTextEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"
        self.radio_style_normal = "QRadioButton{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/    font-size:12pt;    /*字体颜色为白色*/        color:rgb(74, 74, 74);    /*背景颜色*/  }/*按钮停留态*/QRadioButton:hover{    /*背景颜色*/      background-color:rgb(21, 175, 0);	font-size:14pt;    padding-left:-3pt;	font-weight:bold;	color:white;    /*上内边距为3像素，让按下时字向下移动3像素*/      padding-top:-3pt;}/*按钮按下态*/QRadioButton:pressed{    /*背景颜色*/      background-color:rgb(0, 113, 0);	font-weight:bold;	font-size:14pt;	color:white;    /*左内边距为3像素，让按下时字向右移动3像素*/      padding-left:3pt;    /*上内边距为3像素，让按下时字向下移动3像素*/      padding-top:3pt;}"
        self.lineEdite_style_normal = "QLineEdit{    /*字体为微软雅黑*/    font-family:Microsoft Yahei;    /*字体大小为20点*/	color:grey;   font-size:12pt;	background-color:rgb(255, 255, 255);}QLineEdit:hover{    /*背景颜色*/     background-color:rgb(229, 242, 255);}QLineEdit:focus{    /*背景颜色*/     background-color:rgb(229, 242, 255);}"

        self.cursor = DB.cursor()
        sql = """SELECT AN_DAYS, SICK_DAYS, HOME_TOWN, PERSONAL_DAYS FROM akt_staff_ WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (str(ID)))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (str(ID)))

        res = self.cursor.fetchall()
        self.annual_remain=res[0][0]
        self.label_10.setText(f'{self.annual_remain} DAYS')
        self.sick_remain=res[0][1]
        self.label_18.setText(f'{self.sick_remain} DAYS')
        self.is_query()
        self.home_town = res[0][2]
        self.label_25.setText(f'Remains {self.home_town} time(s) for traveling to hometown in this year.')
        self.personal_remain=res[0][3]         #------------Updated on 9/11/2022
        self.label_23.setText(f'{self.personal_remain} DAYS')    #-----------------Updated on 9/11/2022

        self.download_data()

    def cancel_request_on_table(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return
        try:
            request_id = self.tableWidget.item(index, 0).text()
            md=self.tableWidget.item(index, 16).text()

        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if md=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return
        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.is_query()
        self.is_query_hometown()

        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.download_data()

    def download_data(self):
        self.tableWidget.clearContents()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setRowCount(1)
        if self.comboBox_6.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit_6.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_6.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit_6.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit_6.text() + self.comboBox_6.currentText() + '01',
                                                  '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit_6.text() + str(int(self.comboBox_6.currentText()) + 1) + '01', '%Y%m%d')

        sql = """SELECT * FROM leave_request WHERE USER_ID=%s AND (APPLY_DTTM>=%s AND APPLY_DTTM<%s)"""
        try:
            self.cursor.execute(sql, (ID, date_min, date_max))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID, date_min, date_max))

        res = self.cursor.fetchall()
        if res == ():
            #QMessageBox.information(self, 'Empty Record', 'No record in the selected time range.')
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            for j in range(len(res[0])):
                #=========================Updated on 12/9/2023, switching HR and MD
                if j == 15:
                    temp_j = 16
                elif j == 16:
                    temp_j = 15
                else:
                    temp_j = j

                value=res[i][temp_j]
                if temp_j in [13, 14, 15, 16]:
                    if res[i][temp_j]==None:
                        value='Unconfirmed'
                    elif res[i][temp_j]==1:
                        value='OK'
                    else:
                        value='Declined'
                #==========================Updated on 12/9/2023, switching HR and MD (End)

                if j == 5:
                    apply_date_temp = datetime.datetime.strptime(value, '%d/%m/%Y')
                    value = datetime.datetime.strftime(apply_date_temp, '%Y-%m-%d')
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(value)))



        for i in range(17):
            if i == 12:
                continue
            self.tableWidget.resizeColumnToContents(i)
        self.tableWidget.setSortingEnabled(True)

    def del_hometown_request(self):
        if self.comboBox_3.currentText() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.comboBox_3.currentText())
        sql = """SELECT SERIAL, TYPE, APPLY_DTTM, START_DT,START_LEN,END_DT, END_LEN, DURING, REMARKS, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()
        request_id = res[0][0]
        type = res[0][1]
        apply_dt = res[0][2]
        start_dt = res[0][3]
        start_len = res[0][4]
        end_dt = res[0][5]
        end_len = res[0][6]
        during = res[0][7]
        destination = res[0][8]
        leader = res[0][9]
        dm = res[0][10]
        md = res[0][11]
        hr = res[0][12]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(md)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        end_dt_func = lambda y: None if y == None else y.strftime("%d/%m/%Y")
        msm = f'Are you sure to delete this request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {apply_dt}\n' \
              f'Destination: {destination}\n' \
              f'Type: {type}\n' \
              f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Start from: {start_len}\n' \
              f'Until(Date): {end_dt_func(end_dt)}\n' \
              f'Until(time): {end_len}\n' \
              f'Duration: {during} day(s)\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'HR approving: {approve_func(hr)}\n' \
              f'MD approving: {approve_func(md)}\n'
        #Updated on 12/9/2023, switching md and hr
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        self.cursor.execute(sql, (serial))
        DB.commit()
        self.is_query_hometown()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been deleted successfully!')

    def default_css_hometown(self):
        self.lineEdit.setStyleSheet(self.lineEdite_style_normal)
        self.dateEdit_3.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit_4.setStyleSheet(self.dateEdite_style_normal)
        self.pushButton_12.setText('―')
        self.pushButton_13.setText('―')
        self.pushButton_14.setText('―')
        self.pushButton_15.setText('―')

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_12.setStyleSheet(btn_yellow_css)
        self.pushButton_13.setStyleSheet(btn_yellow_css)
        self.pushButton_14.setStyleSheet(btn_yellow_css)
        self.pushButton_15.setStyleSheet(btn_yellow_css)

        self.label_38.setText('WAITING...')
        self.label_38.setStyleSheet("QLabel{color:rgb(255, 170, 0);}")

    def refresh_hometown(self):
        self.default_css_hometown()
        if self.comboBox_3.currentText() == '':
            return
        sql = """SELECT REMARKS, START_DT, END_DT, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (self.comboBox_3.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (self.comboBox_3.currentText()))

        res = self.cursor.fetchall()
        destination = res[0][0]
        start_dt = res[0][1]
        end_dt = res[0][2]
        leader = res[0][3]
        dm = res[0][4]
        md = res[0][5]
        hr = res[0][6]

        self.lineEdit.setText(destination)
        self.lineEdit.setStyleSheet(self.lineEdite_style_highlight)
        self.dateEdit_3.setDate(start_dt)
        self.dateEdit_3.setStyleSheet(self.dateEdite_style_highlight)
        self.dateEdit_4.setDate(end_dt)
        self.dateEdit_4.setStyleSheet(self.dateEdite_style_highlight)

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"

        if leader == None:
            self.pushButton_12.setText('―')
            self.pushButton_12.setStyleSheet(btn_yellow_css)
        elif leader == 1:
            self.pushButton_12.setText('〇')
            self.pushButton_12.setStyleSheet(btn_green_css)
        else:
            self.pushButton_12.setText('×')
            self.pushButton_12.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_13.setText('―')
            self.pushButton_13.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_13.setText('〇')
            self.pushButton_13.setStyleSheet(btn_green_css)
        else:
            self.pushButton_13.setText('×')
            self.pushButton_13.setStyleSheet(btn_red_css)

        #=============================Updated on 12/9/2023, switching hr and md
        if hr == None:
            self.pushButton_14.setText('―')
            self.pushButton_14.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_14.setText('〇')
            self.pushButton_14.setStyleSheet(btn_green_css)
        else:
            self.pushButton_14.setText('×')
            self.pushButton_14.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_15.setText('―')
            self.pushButton_15.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_15.setText('〇')
            self.pushButton_15.setStyleSheet(btn_green_css)
        else:
            self.pushButton_15.setText('×')
            self.pushButton_15.setStyleSheet(btn_red_css)
        #===========================Updated on 12/9/2023, switching hr and md (End)

        if self.pushButton_12.text() == '×' or self.pushButton_13.text() == '×' or self.pushButton_14.text() == '×' or self.pushButton_15.text() == '×':
            self.label_38.setText('DECLINED')
            self.label_38.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_15.text() == '〇':
            self.label_38.setText('OK')
            self.label_38.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_38.setText('WAITING...')
            self.label_38.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def get_id_hometown(self):
        self.comboBox_3.clear()
        if self.lineEdit_3.text() == '':
            self.cursor_2 = DB.cursor()
            sql = """SELECT SERIAL FROM leave_request WHERE USER_ID=%s AND TYPE=%s"""
            try:
                self.cursor_2.execute(sql, (ID, 'hometown'))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_2=DB.cursor()
                self.cursor_2.execute(sql, (ID, 'hometown'))

            res = self.cursor_2.fetchall()
            if res == ():
                self.cursor_2.close()
                return

            id_lst = []
            for each in res:
                id_lst.append(str(each[0]))
            self.comboBox_3.addItems(id_lst)
            self.cursor_2.close()
            return

        year = self.lineEdit_3.text()
        try:
            dttm0 = datetime.datetime.strptime(year + '0101', '%Y%m%d')
            dttm1 = datetime.datetime.strptime(str(int(year) + 1) + '0101', '%Y%m%d')
        except:
            return
        self.cursor_2 = DB.cursor()
        sql = """SELECT SERIAL FROM leave_request WHERE TYPE=%s AND APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID=%s"""
        self.cursor_2.execute(sql, ('hometown', dttm0, dttm1, ID))
        res = self.cursor_2.fetchall()
        if res == ():
            self.cursor_2.close()
            return
        id_lst = []
        for each in res:
            id_lst.append(str(each[0]))
        self.comboBox_3.addItems(id_lst)
        self.cursor_2.close()

    def query_mode_hometown(self):
        self.lineEdit.setEnabled(False)
        self.dateEdit_3.setEnabled(False)
        self.dateEdit_4.setEnabled(False)
        self.pushButton_3.setEnabled(False)

        self.comboBox_3.setEnabled(True)
        self.lineEdit_3.setEnabled(True)
        self.pushButton_6.setEnabled(True)
        self.pushButton_12.setEnabled(True)
        self.pushButton_13.setEnabled(True)
        self.pushButton_14.setEnabled(True)
        self.pushButton_15.setEnabled(True)

        self.get_id_hometown()

    def apply_mode_hometown(self):
        self.lineEdit.setEnabled(True)
        self.dateEdit_3.setEnabled(True)
        self.dateEdit_4.setEnabled(True)
        self.pushButton_3.setEnabled(True)

        self.comboBox_3.setEnabled(False)
        self.lineEdit_3.setEnabled(False)
        self.pushButton_6.setEnabled(False)
        self.pushButton_12.setEnabled(False)
        self.pushButton_13.setEnabled(False)
        self.pushButton_14.setEnabled(False)
        self.pushButton_15.setEnabled(False)

        self.default_css_hometown()

    def is_query_hometown(self):
        self.default_css_hometown()
        if self.checkBox_4.isChecked():
            self.query_mode_hometown()
        else:
            self.apply_mode_hometown()
            self.panel_to_default()

    def request_go_home(self):
        if self.home_town==0:
            QMessageBox.warning(self, 'Request Denied', 'Request denied: 0 time remains for traveling to hometown in this year!')
            return

        if self.lineEdit.text() == '':
            QMessageBox.critical(self, 'Error', 'Please input the "Destination"(Your hometown)!')
            return

        start_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        end_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        if int(start_dt.strftime('%Y%m%d')) >= int(end_dt.strftime('%Y%m%d')):
            QMessageBox.critical(self, 'Date Error', 'Date error: end date must be later than start date!')
            return

        id = ID
        name = MainWindow.name
        type = 'Hometown'
        apply_dttm = TimeCard.t
        apply_dt = TimeCard.t.strftime('%d/%m/%Y')
        start_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        start_len = 'all'
        end_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        end_len = 'all'
        single_day = 0
        during = (end_dt - start_dt).days + 1
        remarks = self.lineEdit.text()

        msm = f'Submit this hometown traveling request?\n' \
              f'Leave type: Hometown traveling\n' \
              f'Start from: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Until: {end_dt.strftime("%d/%m/%Y")}\n' \
              f'Duration: {during} days'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0: #如果存在LEADER
            current_to = res[0][0]
            current_po = 'LEADER'
        else:  #如果不存在LEADER
            if res[0][1] != 0: #如果存在DM
                current_to = res[0][1]
                current_po = 'DM'
            else:  #如果不存在DM
                current_to = 8888     #Updated on 12/9/2023
                current_po = 'HR'

        sql = """INSERT INTO leave_request (USER_ID, USER_NAME, TYPE, APPLY_DTTM, APPLY_DT,START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        self.cursor.execute(sql, (
        id, name, type, apply_dttm, apply_dt, start_dt, start_len, end_dt, end_len, single_day, during, remarks,
        current_to, current_po))
        DB.commit()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='leave')
        # ===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Hometown traveling request has been sent successfully! Please wait for the approvements.')
        self.panel_to_default()

    def del_single_request(self):
        if self.lineEdit_2.text() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.lineEdit_2.text())
        sql = """SELECT SERIAL, TYPE, START_DT, START_LEN, END_DT, END_LEN, DURING, REMARKS, LEADER, DM, MD, HR, APPLY_DTTM FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()

        request_id = res[0][0]
        type = res[0][1]
        start_dt = res[0][2]
        start_len = res[0][3]
        end_dt = res[0][4]
        end_len = res[0][5]
        during = res[0][6]
        remarks = res[0][7]
        leader = res[0][8]
        dm = res[0][9]
        md = res[0][10]
        hr = res[0][11]
        request_date = res[0][12]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(md)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        end_dt_func = lambda y: None if y == None else y.strftime("%d/%m/%Y")
        msm = f'Are you sure to delete this request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {request_date}\n' \
              f'Leave type: {type}\n' \
              f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
              f'Start from: {start_len}\n' \
              f'Until(Date): {end_dt_func(end_dt)}\n' \
              f'Until(time): {end_len}\n' \
              f'Duration: {during} day(s)\n' \
              f'Remarks: {remarks}\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'HR approving: {approve_func(hr)}\n' \
              f'MD approving: {approve_func(md)}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """DELETE FROM leave_request WHERE SERIAL=%s"""
        self.cursor.execute(sql, (serial))
        DB.commit()
        self.lineEdit_2.setText('')
        self.input_id()
        self.show_IDs()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been deleted successfully!')

    def input_id(self):
        txt = self.comboBox_2.currentText()
        #if txt == '':
            #return
        self.lineEdit_2.setText(txt)

    def default_css(self):
        self.dateEdit_5.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit.setStyleSheet(self.dateEdite_style_normal)
        self.dateEdit_2.setStyleSheet(self.dateEdite_style_normal)
        self.comboBox_5.setStyleSheet(self.comboBox_style_normal)
        self.comboBox.setStyleSheet(self.comboBox_style_normal)
        self.comboBox_4.setStyleSheet(self.comboBox_style_normal)
        self.textEdit.setStyleSheet(self.textEdit_style_normal)
        self.radioButton.setStyleSheet(self.radio_style_normal)
        self.radioButton_2.setStyleSheet(self.radio_style_normal)
        self.radioButton_3.setStyleSheet(self.radio_style_normal)
        self.radioButton_4.setStyleSheet(self.radio_style_normal)
        self.radioButton_5.setStyleSheet(self.radio_style_normal)
        self.radioButton_6.setStyleSheet(self.radio_style_normal)
        self.radioButton_7.setStyleSheet(self.radio_style_normal)

        self.pushButton_8.setText('―')
        self.pushButton_9.setText('―')
        self.pushButton_10.setText('―')
        self.pushButton_11.setText('―')

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_8.setStyleSheet(btn_yellow_css)
        self.pushButton_9.setStyleSheet(btn_yellow_css)
        self.pushButton_10.setStyleSheet(btn_yellow_css)
        self.pushButton_11.setStyleSheet(btn_yellow_css)

        self.label_21.setText('WAITING...')
        self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def refresh_data(self):
        self.default_css()

        self.cursor_2 = DB.cursor()
        sql = """SELECT TYPE, START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, REMARKS, LEADER, DM, MD, HR FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_2.execute(sql, (self.comboBox_2.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_2=DB.cursor()
            self.cursor_2.execute(sql, (self.comboBox_2.currentText()))

        res = self.cursor_2.fetchall()

        self.cursor_2.close()

        if res == ():
            return
        type = res[0][0]
        start_date = res[0][1]
        start_len = res[0][2]
        end_date = res[0][3]
        end_len = res[0][4]
        single_day = res[0][5]
        remarks = res[0][6]
        leader = res[0][7]
        dm = res[0][8]
        md = res[0][9]
        hr = res[0][10]

        len_func = lambda x: 'Morning leave' if x == 'morning' else (
            'Afternoon leave' if x == 'afternoon' else 'All day leave')
        if single_day == 1:
            self.checkBox.setChecked(False)

            self.dateEdit_5.setDate(start_date)
            self.dateEdit_5.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox_5.setCurrentText(len_func(start_len))
            self.comboBox_5.setStyleSheet(self.comboBox_style_highlight)

        else:
            self.checkBox.setChecked(True)

            self.dateEdit.setDate(start_date)
            self.dateEdit.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox.setCurrentText(len_func(start_len))
            self.comboBox.setStyleSheet(self.comboBox_style_highlight)

            self.dateEdit_2.setDate(end_date)
            self.dateEdit_2.setStyleSheet(self.dateEdite_style_highlight)

            self.comboBox_4.setCurrentText(len_func(end_len))
            self.comboBox_4.setStyleSheet(self.comboBox_style_highlight)

        self.textEdit.setText(remarks)
        self.textEdit.setStyleSheet(self.textEdit_style_highlight)

        if type == 'Annual leave':
            self.radioButton.setChecked(True)
            self.radioButton.setStyleSheet(self.radio_style_highlight)
        elif type == 'Leave without pay':
            self.radioButton_5.setChecked(True)
            self.radioButton_5.setStyleSheet(self.radio_style_highlight)
        elif type == 'Sick leave':
            self.radioButton_3.setChecked(True)
            self.radioButton_3.setStyleSheet(self.radio_style_highlight)
        elif type == 'Compensatory leave':
            self.radioButton_6.setChecked(True)
            self.radioButton_6.setStyleSheet(self.radio_style_highlight)
        elif type == 'Maternity leave':
            self.radioButton_2.setChecked(True)
            self.radioButton_2.setStyleSheet(self.radio_style_highlight)
        elif type == 'Personal leave':
            self.radioButton_7.setChecked(True)
            self.radioButton_7.setStyleSheet(self.radio_style_highlight)
        else:
            self.radioButton_4.setChecked(True)
            self.radioButton_4.setStyleSheet(self.radio_style_highlight)

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"
        if leader == None:
            self.pushButton_8.setText('―')
            self.pushButton_8.setStyleSheet(btn_yellow_css)

        elif leader == 1:
            self.pushButton_8.setText('〇')
            self.pushButton_8.setStyleSheet(btn_green_css)
        else:
            self.pushButton_8.setText('×')
            self.pushButton_8.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_9.setText('―')
            self.pushButton_9.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_9.setText('〇')
            self.pushButton_9.setStyleSheet(btn_green_css)
        else:
            self.pushButton_9.setText('×')
            self.pushButton_9.setStyleSheet(btn_red_css)

        if hr == None:
            self.pushButton_10.setText('―')
            self.pushButton_10.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_10.setText('〇')
            self.pushButton_10.setStyleSheet(btn_green_css)
        else:
            self.pushButton_10.setText('×')
            self.pushButton_10.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_11.setText('―')
            self.pushButton_11.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_11.setText('〇')
            self.pushButton_11.setStyleSheet(btn_green_css)
        else:
            self.pushButton_11.setText('×')
            self.pushButton_11.setStyleSheet(btn_red_css)

        if self.pushButton_8.text() == '×' or self.pushButton_9.text() == '×' or self.pushButton_10.text() == '×' or self.pushButton_11.text() == '×':
            self.label_21.setText('DECLINED')
            self.label_21.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_11.text() == '〇':
            self.label_21.setText('OK')
            self.label_21.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_21.setText('WAITING...')
            self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def show_IDs(self):
        self.comboBox_2.clear()
        sql = """SELECT SERIAL FROM leave_request WHERE USER_ID=%s AND APPLY_DT=%s AND TYPE!=%s"""
        try:
            self.cursor.execute(sql, (ID, self.dateEdit_7.text(), 'hometown'))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID, self.dateEdit_7.text(), 'hometown'))

        res = self.cursor.fetchall()
        if res == ():
            return
        id_lst = []
        for each in res:
            id_lst.append(str(each[0]))
        self.comboBox_2.addItems(id_lst)

    def request_per_leave(self):
        if self.checkBox.isChecked():
            start_dt = datetime.datetime.strptime(self.dateEdit.text(), '%d/%m/%Y')
            if self.comboBox.currentText() == 'All day leave':
                start_len = 'all'
            else:
                start_len = 'afternoon'

            end_dt = datetime.datetime.strptime(self.dateEdit_2.text(), '%d/%m/%Y')
            if self.comboBox_4.currentText() == 'All day leave':
                end_len = 'all'
            else:
                end_len = 'morning'

            if int(start_dt.strftime('%Y%m%d')) >= int(end_dt.strftime('%Y%m%d')):
                QMessageBox.critical(self, 'Date Error', 'Date error: end date must be later than start date!')
                return

            single = 0
            during = (end_dt - start_dt).days + 1
            if start_len == 'afternoon':
                during -= 0.5
            if end_len == 'morning':
                during -= 0.5

        else:
            start_dt = datetime.datetime.strptime(self.dateEdit_5.text(), '%d/%m/%Y')
            # print(self.comboBox_5.currentText())
            if self.comboBox_5.currentText() == 'All day leave':
                start_len = 'all'
            elif self.comboBox_5.currentText() == 'Morning leave':
                start_len = 'morning'
            else:
                start_len = 'afternoon'

            single = 1
            if start_len == 'all':
                during = 1
            else:
                during = 0.5

            end_dt = None
            end_len = None
            # print(during)

        if self.radioButton_7.isChecked(): #Personal leave --Updated on 8/11/2022
            #print(int(start_dt.strftime('%Y%m%d'))-int(TimeCard.t.strftime('%Y%m%d')))
            days_before=int(start_dt.strftime('%Y%m%d'))-int(TimeCard.t.strftime('%Y%m%d'))
            #if days_before<3:
            #    QMessageBox.critical(self, 'Warning',
            #                         "Personal leave must be applied at least 3 days beforehand!")
            #    return

        remarks = self.textEdit.toPlainText()
        id = ID
        name = MainWindow.name

        if self.radioButton.isChecked():
            if float(self.annual_remain) < float(during): #--------Updated on 9/11/2022
                QMessageBox.critical(self, 'Warning',
                                     "Amount of remaining annual leave days is not enough, please change the type of leave that you are requesting.")
                return
        if self.radioButton_3.isChecked():
            if float(self.sick_remain) < float(during):   #----------Updated on 9/11/2022
                QMessageBox.critical(self, 'Warning',
                                     "Amount of remaining sick leave days is not enough, please change the type of leave that you are requesting.")
                return

        if self.radioButton_7.isChecked():
            if float(self.personal_remain) < float(during):  #-----------Updated on 9/11/2022
                QMessageBox.critical(self, 'Warning',        #-----------Updated on 9/11/2022
                                     "Amount of remaining personal leave days is not enough, please change the type of leave that you are requesting.")
                return                    #------------Updated on 9/11/2022

        if self.radioButton.isChecked():
            type = 'Annual leave'
        elif self.radioButton_5.isChecked():
            type = 'Leave without pay'
        elif self.radioButton_3.isChecked():
            type = 'Sick leave'
        elif self.radioButton_6.isChecked():
            type = 'Compensatory leave'
        elif self.radioButton_2.isChecked():
            type = 'Maternity leave'
        elif self.radioButton_7.isChecked():
            type = 'Personal leave'
        else:
            type = 'Sterilisation leave'

        if type == 'Sick leave' or type == 'Personal leave':
            if str(remarks).strip() == '':
                QMessageBox.warning(self, 'Warning', 'Please input the remarks content before sending request!')
                return

        if single == 1:
            msm = f'Submit this leave request?\n' \
                  f'Leave type: {type}\n' \
                  f'Leave date: {start_dt.strftime("%d/%m/%Y")}\n' \
                  f'Leave time: {start_len}\n' \
                  f'Duration: {during} day\n' \
                  f'Remarks: {remarks}'
        else:
            msm = f'Submit this leave request?\n' \
                  f'Leave type: {type}\n' \
                  f'Start date: {start_dt.strftime("%d/%m/%Y")}\n' \
                  f'Start from: {start_len}\n' \
                  f'Until(Date): {end_dt.strftime("%d/%m/%Y")}\n' \
                  f'Until(time): {end_len}\n' \
                  f'Duration: {during} day(s)\n' \
                  f'Remarks: {remarks}'

        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 8888
                current_po = 'HR'

        sql = """INSERT INTO leave_request (USER_ID, USER_NAME, TYPE, APPLY_DTTM, APPLY_DT,START_DT, START_LEN, END_DT, END_LEN, SINGLE_DAY, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        self.cursor.execute(sql, (
        id, name, type, TimeCard.t, TimeCard.t.strftime('%d/%m/%Y'), start_dt, start_len, end_dt, end_len, single,
        during, remarks, current_to, current_po))
        DB.commit()

        #Mail sending==============================================
        info_lst=query_email(id=current_to)
        if info_lst==-1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='leave')
        #===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Leave request has been sent successfully! Please wait for the approvements.')

        self.panel_to_default()

    def query_mode(self):
        self.dateEdit_5.setEnabled(False)
        self.checkBox.setEnabled(False)
        self.dateEdit.setEnabled(False)
        self.dateEdit_2.setEnabled(False)
        self.comboBox_5.setEnabled(False)
        self.comboBox.setEnabled(False)
        self.comboBox_4.setEnabled(False)
        self.textEdit.setEnabled(False)
        self.radioButton.setEnabled(False)
        self.radioButton_2.setEnabled(False)
        self.radioButton_3.setEnabled(False)
        self.radioButton_4.setEnabled(False)
        self.radioButton_5.setEnabled(False)
        self.radioButton_6.setEnabled(False)
        self.pushButton.setEnabled(False)

        self.comboBox_2.setEnabled(True)
        self.dateEdit_7.setEnabled(True)
        self.pushButton_5.setEnabled(True)

        self.lineEdit_2.setText('')
        self.input_id()
        self.show_IDs()

    def apply_mode(self):
        self.dateEdit_5.setEnabled(True)
        self.checkBox.setEnabled(True)
        self.dateEdit.setEnabled(True)
        self.dateEdit_2.setEnabled(True)
        self.comboBox_5.setEnabled(True)
        self.comboBox.setEnabled(True)
        self.comboBox_4.setEnabled(True)
        self.textEdit.setEnabled(True)
        self.radioButton.setEnabled(True)
        self.radioButton_2.setEnabled(True)
        self.radioButton_3.setEnabled(True)
        self.radioButton_4.setEnabled(True)
        self.radioButton_5.setEnabled(True)
        self.radioButton_6.setEnabled(True)
        self.pushButton.setEnabled(True)

        self.comboBox_2.setEnabled(False)
        self.dateEdit_7.setEnabled(False)
        self.pushButton_5.setEnabled(False)

        self.is_multiple()

    def is_query(self):
        self.default_css()
        if self.checkBox_2.isChecked():
            self.query_mode()
        else:
            self.apply_mode()
            self.panel_to_default()
            self.default_css()

    def is_multiple(self):
        if self.checkBox.isChecked():
            self.dateEdit.setEnabled(True)
            self.dateEdit_2.setEnabled(True)
            self.comboBox.setEnabled(True)
            self.comboBox_4.setEnabled(True)

            self.dateEdit_5.setEnabled(False)
            self.comboBox_5.setEnabled(False)
        else:
            self.dateEdit.setEnabled(False)
            self.dateEdit_2.setEnabled(False)
            self.comboBox.setEnabled(False)
            self.comboBox_4.setEnabled(False)

            self.dateEdit_5.setEnabled(True)
            self.comboBox_5.setEnabled(True)

    def panel_to_default(self):
        self.dateEdit_5.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_2.setDate(QDate.currentDate())
        self.dateEdit_7.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit_6.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_6.setCurrentText(month)
        self.textEdit.setText('')
        self.lineEdit.setText('')
        self.radioButton.setChecked(True)
        self.checkBox_2.setChecked(False)
        self.apply_mode()
        self.checkBox_4.setChecked(False)
        self.apply_mode_hometown()

    def quit(self):
        AskForLeave.close()

    def closeEvent(self, event):
        try:
            self.cursor.close()
        except:
            pass
        MainWindow.show()


class OTApplication(QMainWindow, Ui_OTApplication):
    def __init__(self):
        super(OTApplication, self).__init__()
        self.setupUi(self)

        self.label_31.setText('')
        self.timeEdit.setTime(QTime.fromString('17:15', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('18:15', 'hh:mm'))
        self.initHours()
        self.textEdit_5.setReadOnly(True)
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.geoTab = self.tabWidget.geometry()
        self.geoTable = self.tableWidget.geometry()
        self.geoBTN6 = self.pushButton_6.geometry()
        self.geoBTN2 = self.pushButton_2.geometry()
        self.geoBTN4 = self.pushButton_4.geometry()
        self.geoFrame2 = self.frame_2.geometry()
        self.geoBTN16 = self.pushButton_16.geometry()

        self.pushButton.clicked.connect(self.submit_request)
        self.timeEdit.timeChanged.connect(self.initHours)
        self.timeEdit_2.timeChanged.connect(self.initHours)
        self.checkBox.clicked.connect(self.is_query)
        self.pushButton_4.clicked.connect(self.quit)
        self.dateEdit_4.dateChanged.connect(self.show_key)
        self.comboBox_2.currentTextChanged.connect(self.refresh_panel)
        self.pushButton_5.clicked.connect(self.cancel_apply)
        self.pushButton_16.clicked.connect(self.download_data)
        self.pushButton_6.clicked.connect(self.cancel_request_on_table)
        self.pushButton_2.clicked.connect(self.export_excel)

        self.dateEdit.dateTimeChanged.connect(self.download_data)
        self.comboBox_3.currentTextChanged.connect(self.download_data)
        self.pushButton_16.setVisible(False)

        self.tabWidget.currentChanged.connect(self.tableview)

    def export_excel(self):
        try:
            first_cell = self.tableWidget.item(0, 0).text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Table', 'Warning: The table is empty, no need to export the excel file!')
            return

        yearmonth = self.dateEdit.text() + self.comboBox_3.currentText()
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./{MainWindow.name}\'s OT Application({yearmonth})',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return
        excel_path = a[0]

        data_frame = []
        for i in range(self.tableWidget.rowCount()):
            data_line = []
            for j in range(self.tableWidget.columnCount()):
                data_line.append(self.tableWidget.item(i, j).text())
            data_frame.append(data_line)

        wb = xl.Workbook()
        ws = wb.active
        ws.append(['Request ID', 'Staff ID', 'Staff Name', 'Request Date-Time', 'Request Date',
                   'Date of OT', 'Start Time', 'End Time', 'OT Hours', 'Leader', 'DM', 'HR', 'MD','Remarks'])

        for line in data_frame:
            ws.append(line)

        try:
            wb.save(filename=excel_path)
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been exported successfully!')

    def tableview(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        if self.tabWidget.currentIndex() == 1:
            width0 = W_OT
            height0 = H_OT
            self.desktop = QApplication.desktop()
            screen_count = self.desktop.screenCount()  #2 screen version updating
            geometry = self.desktop.geometry()
            if screen_count == 1 or screen_count==2:                      #2 screen version updating
                width1 = geometry.width()              #2 screen version updating
            else:                                      #2 screen version updating
                width1 = geometry.width()/2            #2 screen version updating
            height1 = geometry.height()
            x = width1 - width0
            y = height1 - height0 - 100

            self.setMaximumSize(16777215, 16777215)
            self.showMaximized()

            move_widgets_for_frame(self.tabWidget, x, y)
            move_widgets_for_frame(self.tableWidget, x, y)
            move_widgets_only_vertical(self.pushButton_6, x, y)
            move_widgets_only_vertical(self.pushButton_2, x, y)
            move_widgets_only_horizontal(self.pushButton_16, x, y)
            move_widgets_both(self.pushButton_4, x, y)
            extend_widgets_horizontal(self.frame_2, x, y)
        else:
            self.tabWidget.setGeometry(self.geoTab.x(), self.geoTab.y(), self.geoTab.width(), self.geoTab.height())
            self.tableWidget.setGeometry(self.geoTable.x(), self.geoTable.y(), self.geoTable.width(),
                                         self.geoTable.height())
            self.pushButton_6.setGeometry(self.geoBTN6.x(), self.geoBTN6.y(), self.geoBTN6.width(),
                                          self.geoBTN6.height())
            self.pushButton_2.setGeometry(self.geoBTN2.x(), self.geoBTN2.y(), self.geoBTN2.width(),
                                          self.geoBTN2.height())
            self.pushButton_4.setGeometry(self.geoBTN4)
            self.pushButton_16.setGeometry(self.geoBTN16.x(), self.geoBTN16.y(), self.geoBTN16.width(),
                                           self.geoBTN16.height())
            self.frame_2.setGeometry(self.geoFrame2)
            self.showNormal()
            self.setMaximumSize(W_OT, H_OT)

    def hours_css_control(self, mode):
        if mode == 0:
            self.textEdit_5.setStyleSheet(
                'QTextEdit{ font-family:Microsoft Yahei; color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 251, 201);}')
            self.label_31.setText('')
        else:
            self.textEdit_5.setStyleSheet(
                'QTextEdit{  font-family:Microsoft Yahei;	color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 120, 120);}')
            self.label_31.setText('Please input the correct time range!')

    def initHours(self):
        seconds = self.timeEdit.time().secsTo(self.timeEdit_2.time())
        hours = seconds / 3600
        hours = round(hours, 2)
        self.textEdit_5.setText(str(hours))
        if hours < 0:
            self.hours_css_control(mode=1)
        else:
            self.hours_css_control(mode=0)

    def initializing(self):
        self.apply_mode()

        self.dateEdit_css_highlighted = 'QDateEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'
        self.timeEdit_css_highlighted = 'QTimeEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'
        self.textEdit_css_highlighted = 'QTextEdit{    font-family:Microsoft Yahei;   color:rgb(50, 50, 50);   font-size:12pt;   background-color:rgb(255, 163, 144);}'
        self.remarks_css_highlighted = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 163, 144);}'

        self.dateEdit_css_normal = 'QDateEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255,255);}QDateEdit:hover{   background-color:rgb(229, 242, 255);}QDateEdit:focus{   background-color:rgb(229, 242, 255);}'
        self.timeEdit_css_normal = 'QTimeEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTimeEdit:hover{   background-color:rgb(229, 242, 255);}QTimeEdit:focus{   background-color:rgb(229, 242, 255);}'
        self.textEdit_css_normal = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(50, 50, 50);   font-size:12pt;	background-color:rgb(255, 251, 201);}'
        self.remarks_css_normal = 'QTextEdit{    font-family:Microsoft Yahei;	color:rgb(74, 74, 74);   font-size:12pt;	background-color:rgb(255, 255, 255);}QTextEdit:hover{   background-color:rgb(229, 242, 255);}QTextEdit:focus{   background-color:rgb(229, 242, 255);}'

        self.download_data()

    def cancel_request_on_table(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            md=self.tableWidget.item(index, 12).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if md=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.show_key()

        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.download_data()
        self.cursor.close()

    def download_data(self):
        self.tableWidget.clearContents()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM ot_request WHERE OT_DT>=%s AND OT_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            #QMessageBox.information(self, 'Empty Record', 'No record in the selected time range.')
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            for j in range(len(res[0])):
                if j == 11:
                    temp_j = 12
                elif j == 12:
                    temp_j = 11
                else:
                    temp_j = j
                value=res[i][temp_j]
                if j in [9,10,11,12]:
                    if res[i][temp_j]==None:
                        value='Unconfirmed'
                    elif res[i][temp_j]==1:
                        value='OK'
                    else:
                        value='Declined'

                if j == 4:
                    apply_dt_temp = datetime.datetime.strptime(value, '%d/%m/%Y')
                    value = datetime.datetime.strftime(apply_dt_temp, '%Y-%m-%d')

                self.tableWidget.setItem(i, j, QTableWidgetItem(str(value)))

        for i in [0, 3, 9, 10, 11, 12, 13]:
            # if i==13:
            #   continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()
        self.tableWidget.setSortingEnabled(True)

    def cancel_apply(self):
        if self.comboBox_2.currentText() == '':
            QMessageBox.critical(self, 'Error', 'Please select the request ID first!')
            return

        serial = int(self.comboBox_2.currentText())
        self.cursor = DB.cursor()

        sql = """SELECT SERIAL, OT_DT, START_TM, END_TM, DURING, REMARKS, LEADER, DM, MD, HR, APPLY_DTTM FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))

        res = self.cursor.fetchall()

        request_id = res[0][0]
        ot_dt = res[0][1]
        start_tm = res[0][2]
        end_tm = res[0][3]
        during = res[0][4]
        remarks = res[0][5]
        leader = res[0][6]
        dm = res[0][7]
        md = res[0][8]
        hr = res[0][9]
        request_date = res[0][10]

        approve_func = lambda x: 'Unconfirmed' if x == None else ('OK' if x == 1 else 'Declined')
        if approve_func(md)=='OK':
            QMessageBox.critical(self, 'Notice',
                                 'You can not cancel the request because it has been approved completely by all departments!')
            return

        msm = f'Are you sure to delete this OT request?\n' \
              f'Request ID: {request_id}\n' \
              f'Requested on: {request_date}\n' \
              f'OT date: {ot_dt}\n' \
              f'Start time: {start_tm}\n' \
              f'End time: {end_tm}\n' \
              f'Duration: {during} hour(s)\n' \
              f'Remarks: {remarks}\n' \
              f'Leader approving: {approve_func(leader)}\n' \
              f'DM approving: {approve_func(dm)}\n' \
              f'HR approving: {approve_func(hr)}\n' \
              f'MD approving: {approve_func(md)}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            self.cursor.close()
            return

        sql = """DELETE FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (serial))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (serial))
            DB.commit()

        self.show_key()

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'Selected OT request has been deleted successfully!')

    def highlight_on(self):
        self.dateEdit_3.setStyleSheet(self.dateEdit_css_highlighted)
        self.timeEdit.setStyleSheet(self.timeEdit_css_highlighted)
        self.timeEdit_2.setStyleSheet(self.timeEdit_css_highlighted)
        self.textEdit_5.setStyleSheet(self.textEdit_css_highlighted)
        self.textEdit.setStyleSheet(self.remarks_css_highlighted)

    def highlight_off(self):
        self.dateEdit_3.setStyleSheet(self.dateEdit_css_normal)
        self.timeEdit.setStyleSheet(self.timeEdit_css_normal)
        self.timeEdit_2.setStyleSheet(self.timeEdit_css_normal)
        self.textEdit_5.setStyleSheet(self.textEdit_css_normal)
        self.textEdit.setStyleSheet(self.remarks_css_normal)

        self.pushButton_8.setText('―')
        self.pushButton_9.setText('―')
        self.pushButton_10.setText('―')
        self.pushButton_11.setText('―')
        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        self.pushButton_8.setStyleSheet(btn_yellow_css)
        self.pushButton_9.setStyleSheet(btn_yellow_css)
        self.pushButton_10.setStyleSheet(btn_yellow_css)
        self.pushButton_11.setStyleSheet(btn_yellow_css)

        self.label_21.setText('WAITING...')
        self.label_21.setStyleSheet('QLabel{color:rgb(255, 170, 0);}')

    def refresh_panel(self):
        if self.comboBox_2.currentText() == '':
            self.highlight_off()
            return
        self.cursor = DB.cursor()
        sql = """SELECT OT_DT, START_TM, END_TM, DURING, REMARKS,LEADER, DM, MD, HR FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (self.comboBox_2.currentText()))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (self.comboBox_2.currentText()))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return
        ot_dt = res[0][0]
        start_tm = res[0][1]
        end_tm = res[0][2]
        during = res[0][3]
        remarks = res[0][4]
        leader = res[0][5]
        dm = res[0][6]
        md = res[0][7]
        hr = res[0][8]

        self.dateEdit_3.setDate(ot_dt)
        self.timeEdit.setDateTime(datetime.datetime.strptime(str(start_tm), '%H:%M:%S'))
        self.timeEdit_2.setDateTime(datetime.datetime.strptime(str(end_tm), '%H:%M:%S'))
        self.textEdit_5.setText(str(during))
        self.textEdit.setText(remarks)
        self.highlight_on()
        self.cursor.close()

        btn_yellow_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(255, 170, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(230, 153, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(197, 128, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_green_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(21, 199, 1);    border-radius:10px;}QPushButton:hover{    background-color:rgb(21, 175, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(0, 113, 0);    padding-left:3pt;    padding-top:3pt;}"
        btn_red_css = "QPushButton{    font-family:Microsoft Yahei;    font-size:20pt;    color:white;    background-color:rgb(200, 0, 0);    border-radius:10px;}QPushButton:hover{    background-color:rgb(176, 0, 0);    padding-left:-3pt;    padding-top:-3pt;}QPushButton:pressed{    background-color:rgb(147, 0, 0);    padding-left:3pt;    padding-top:3pt;}"
        if leader == None:
            self.pushButton_8.setText('―')
            self.pushButton_8.setStyleSheet(btn_yellow_css)

        elif leader == 1:
            self.pushButton_8.setText('〇')
            self.pushButton_8.setStyleSheet(btn_green_css)
        else:
            self.pushButton_8.setText('×')
            self.pushButton_8.setStyleSheet(btn_red_css)

        if dm == None:
            self.pushButton_9.setText('―')
            self.pushButton_9.setStyleSheet(btn_yellow_css)
        elif dm == 1:
            self.pushButton_9.setText('〇')
            self.pushButton_9.setStyleSheet(btn_green_css)
        else:
            self.pushButton_9.setText('×')
            self.pushButton_9.setStyleSheet(btn_red_css)

        if hr == None:
            self.pushButton_10.setText('―')
            self.pushButton_10.setStyleSheet(btn_yellow_css)
        elif hr == 1:
            self.pushButton_10.setText('〇')
            self.pushButton_10.setStyleSheet(btn_green_css)
        else:
            self.pushButton_10.setText('×')
            self.pushButton_10.setStyleSheet(btn_red_css)

        if md == None:
            self.pushButton_11.setText('―')
            self.pushButton_11.setStyleSheet(btn_yellow_css)
        elif md == 1:
            self.pushButton_11.setText('〇')
            self.pushButton_11.setStyleSheet(btn_green_css)
        else:
            self.pushButton_11.setText('×')
            self.pushButton_11.setStyleSheet(btn_red_css)

        if self.pushButton_8.text() == '×' or self.pushButton_9.text() == '×' or self.pushButton_10.text() == '×' or self.pushButton_11.text() == '×':
            self.label_21.setText('DECLINED')
            self.label_21.setStyleSheet("QLabel {color:rgb(170, 0, 0);}")
        elif self.pushButton_11.text() == '〇':
            self.label_21.setText('OK')
            self.label_21.setStyleSheet("QLabel {color:rgb(0, 170, 0);}")
        else:
            self.label_21.setText('WAITING...')
            self.label_21.setStyleSheet("QLabel {color:rgb(255, 170, 0);}")

    def show_key(self):
        self.comboBox_2.clear()
        ot_dt = datetime.datetime.strptime(self.dateEdit_4.text(), '%d/%m/%Y')
        self.cursor2 = DB.cursor()
        sql = """SELECT SERIAL FROM ot_request WHERE OT_DT=%s AND USER_ID=%s"""
        try:
            self.cursor2.execute(sql, (ot_dt, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor2=DB.cursor()
            self.cursor2.execute(sql, (ot_dt, ID))

        res = self.cursor2.fetchall()
        if res == ():
            self.cursor2.close()
            return
        keys = []
        for each in res:
            keys.append(str(each[0]))
        self.comboBox_2.addItems(keys)
        self.cursor2.close()

    def query_mode(self):
        self.dateEdit_3.setEnabled(False)
        self.timeEdit.setEnabled(False)
        self.timeEdit_2.setEnabled(False)
        self.textEdit_5.setEnabled(False)
        self.pushButton.setEnabled(False)
        self.textEdit.setEnabled(False)

        self.comboBox_2.setEnabled(True)
        self.pushButton_5.setEnabled(True)
        self.dateEdit_4.setEnabled(True)

        self.show_key()

    def apply_mode(self):
        self.dateEdit_3.setEnabled(True)
        self.timeEdit.setEnabled(True)
        self.timeEdit_2.setEnabled(True)
        self.textEdit_5.setEnabled(True)
        self.pushButton.setEnabled(True)
        self.textEdit.setEnabled(True)

        self.comboBox_2.setEnabled(False)
        self.pushButton_5.setEnabled(False)
        self.dateEdit_4.setEnabled(False)

    def is_query(self):
        if self.checkBox.isChecked():
            self.query_mode()
        else:
            self.apply_mode()
            self.to_default_panel()
            self.highlight_off()

    def submit_request(self):
        user_id = ID
        user_name = MainWindow.name
        apply_dttm = TimeCard.t
        apply_dt = TimeCard.t.strftime('%d/%m/%Y')
        ot_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        start_tm = datetime.datetime.strptime(self.timeEdit.text(), '%H:%M')
        end_tm = datetime.datetime.strptime(self.timeEdit_2.text(), '%H:%M')
        during = float(self.textEdit_5.toPlainText())
        remarks = self.textEdit.toPlainText()

        if str(remarks).strip() == "":
            QMessageBox.critical(self, 'Warning',
                                 'Please input the remarks content before sending the request!')
            return

        self.cursor_confirm=DB.cursor()
        sql="""SELECT * FROM ot_request WHERE USER_ID=%s AND OT_DT=%s"""
        try:
            self.cursor_confirm.execute(sql, (ID, ot_dt))
        except:
            reconnect_DB(self)
            self.cursor_confirm=DB.cursor()
            self.cursor_confirm.execute(sql, (ID, ot_dt))
        if_exist=self.cursor_confirm.fetchall()
        self.cursor_confirm.close()
        if if_exist!=():
            QMessageBox.critical(self, 'Request Denied', f'Request denied: You have already submitted the request for the date of {self.dateEdit_3.text()}! If you want to request again, please cancel the old request first.')
            return

        msm = f'Submit this OT application?\n' \
              f'OT date: {ot_dt}\n' \
              f'Start from: {start_tm.strftime("%H:%M")}\n' \
              f'Until: {end_tm.strftime("%H:%M")}\n' \
              f'Duration: {during} hour(s)\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()

        sql = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        res = self.cursor.fetchall()
        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 8888
                current_po = 'HR'

        sql = """INSERT INTO ot_request (USER_ID, USER_NAME, APPLY_DTTM, APPLY_DT, OT_DT, START_TM, END_TM, DURING, REMARKS, CURRENT_TO, CURRENT_PO) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        self.cursor.execute(sql, (
        user_id, user_name, apply_dttm, apply_dt, ot_dt, start_tm, end_tm, during, remarks, current_to, current_po))
        DB.commit()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='ot')
        # ===========================================================

        self.download_data()
        QMessageBox.information(self, 'Info',
                                'OT application has been submitted! Please wait for the approvements.')
        self.cursor.close()
        self.to_default_panel()

    def to_default_panel(self):
        self.label_31.setText('')
        self.timeEdit.setTime(QTime.fromString('17:15', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('18:15', 'hh:mm'))
        self.initHours()
        self.dateEdit_3.setDate(QDate.currentDate())
        self.dateEdit_4.setDate(QDate.currentDate())
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)
        self.textEdit.setText('')

    def quit(self):
        OTApplication.close()

    def closeEvent(self, event):
        MainWindow.show()


class BookMeetingRoom(QMainWindow, Ui_BookMeetingRoom):
    def __init__(self):
        super(BookMeetingRoom, self).__init__()
        self.setupUi(self)

        self.tableWidget.setRowCount(20)
        self.label_31.setText(' ')
        self.timeEdit.setTime(QTime.fromString('09:00', 'hh:mm'))
        self.timeEdit_2.setTime(QTime.fromString('09:30', 'hh:mm'))

        self.timeEdit.timeChanged.connect(self.is_valid_time)
        self.timeEdit_2.timeChanged.connect(self.is_valid_time)
        self.pushButton_5.clicked.connect(self.quit)
        self.pushButton_3.clicked.connect(self.submit_request)
        self.calendarWidget.clicked.connect(self.show_on_table)
        self.pushButton_4.clicked.connect(self.cancel_booking)

        self.tableWidget.horizontalHeader().sectionClicked.connect(self.sortTable)
        self.calendarSorting = {1 : None,
                                3 : None}

    def sortTable(self, logicalIndex):
        #print(logicalIndex)
        if logicalIndex not in [1, 3]:
            return
        item = self.tableWidget.horizontalHeaderItem(logicalIndex)
        if '^' in item.text():
            item.setText(str(item.text()).replace('^','') + 'v')
            self.calendarSorting[logicalIndex] = Qt.DescendingOrder
        elif 'v' in item.text():
            item.setText(str(item.text()).replace('v', ''))
            self.calendarSorting[logicalIndex] = None
        else:
            item.setText(item.text() + '^')
            self.calendarSorting[logicalIndex] = Qt.AscendingOrder

        # 保存第一列的排序状态
        first_column_sort_order = self.calendarSorting.get(1)

        header = self.tableWidget.horizontalHeader()
        sortOrder = self.calendarSorting[logicalIndex]
        #print(sortOrder)

        if sortOrder == Qt.AscendingOrder:
            header.setSortIndicator(logicalIndex, Qt.AscendingOrder)
        elif sortOrder == Qt.DescendingOrder:
            header.setSortIndicator(logicalIndex, Qt.DescendingOrder)
        else:
            header.setSortIndicator(0, Qt.AscendingOrder)

        for col in self.calendarSorting:
            if self.calendarSorting[col] is not None:
                self.tableWidget.sortItems(col, self.calendarSorting[col])
            else:
                if col == 3:
                    if self.calendarSorting[1] is None:
                        self.tableWidget.sortItems(0, Qt.AscendingOrder)
                else:
                    self.tableWidget.sortItems(0, Qt.AscendingOrder)

        # 恢复第一列的排序状态
        if first_column_sort_order is not None:
            self.tableWidget.sortItems(1, first_column_sort_order)

    def initializing(self):
        self.show_on_table()

    def cancel_booking(self):
        self.calendarWidget.setFocus()
        index = self.tableWidget.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Warning', 'Please select the booking record that you want to cancel.')
            return

        try:
            booking_id = self.tableWidget.item(index, 0).text()
        except:
            QMessageBox.critical(self, 'Empty Selection',
                                 'Empty selection! Please select the booking record that you want to cancel.')
            return

        self.cursor = DB.cursor()

        SQL = """SELECT USER_ID FROM book_meeting_room_ WHERE SERIAL=%s"""
        try:
            self.cursor.execute(SQL, (booking_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(SQL, (booking_id))

        res = self.cursor.fetchall()
        user_id = res[0][0]
        if user_id != ID:
            QMessageBox.critical(self, 'Access Denied',
                                 'Warning: You can only cancel the bookings submitted by yourself.')
            self.cursor.close()
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected booking?\n'
                                                       f'Booking ID: {booking_id}')
        if a == QMessageBox.No:
            self.cursor.close()
            return

        sql = """DELETE FROM book_meeting_room_ WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (booking_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (booking_id))
            DB.commit()
        self.cursor.close()

        self.show_on_table()
        QMessageBox.information(self, 'Info', f'Request ID: {booking_id}, has been canceled successfully!')

    def submit_request(self):
        self.calendarWidget.setFocus()
        if self.label_31.text() == 'Please input the correct time range!':
            QMessageBox.critical(self, 'Time Error', 'Please input the correct time range first!')
            return

        if self.radioButton.isChecked()==False and self.radioButton_2.isChecked()==False and self.radioButton_3.isChecked()==False:  #-------Updated on 9/11/2022, 16/6/2023
            QMessageBox.critical(self, 'Room not selected!', 'Please select the meeting room first!')  #--------Updated on 9/11/2022
            return    #----------Updated on 9/11/2022

        apply_dttm = TimeCard.t
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')
        start_time = datetime.datetime.strptime(self.timeEdit.time().toString('hh:mm'), '%H:%M')
        end_time = datetime.datetime.strptime(self.timeEdit_2.time().toString('hh:mm'), '%H:%M')
        if self.radioButton.isChecked():  #------Updated on 9/11/2022
            room_no = 1  #------Updated on 9/11/2022
        elif self.radioButton_2.isChecked():  #------Updated on 9/11/2022, 16/6/2023
            room_no = 2  #------Updated on 9/11/2022
        else:            #-------Updated on 16/6/2023
            room_no = 3 #-------Updated on 16/6/2023

        user_id = ID
        applier = MainWindow.name
        division = self.get_userDIV()
        contents = self.textEdit.toPlainText()

        judge = self.judge_validation(start_tm=start_time, end_tm=end_time, room_no=room_no)  #------Updated on 9/11/2022
        if judge == False:
            QMessageBox.warning(self, 'Warning',
                                'The time range you selected is occupied by another meeting! Please change the time range or change the room number, then submit the booking again.')  #------Updated on 9/11/2022
            self.show_on_table()
            return

        msm = f'Are you sure to submit the booking?\n' \
              f'Meeting Date: {meeting_dt.strftime("%d/%m/%Y")}\n' \
              f'Room No.: Room {str(room_no)}\n'\
              f'Start Time: {start_time.strftime("%H:%M")}\n' \
              f'Finish Time: {end_time.strftime("%H:%M")}\n' \
              f'Meeting Contents: {contents}\n'  #------Updated on 9/11/2022

        a = QMessageBox.question(self, 'Query', msm)
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """INSERT INTO book_meeting_room_ (APPLY_DTTM, ROOM_NO, MEETING_DT, START_TM, END_TM, USER_ID, USER_NAME, DIVISION, CONTENTS) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""  #------Updated on 9/11/2022
        try:
            self.cursor.execute(sql, (apply_dttm, room_no, meeting_dt, start_time, end_time, user_id, applier, division, contents))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql,
                                (apply_dttm, room_no, meeting_dt, start_time, end_time, user_id, applier, division, contents))  #------Updated on 9/11/2022
            DB.commit()

        QMessageBox.information(self, 'Info', 'Meeting room booking has been completed successfully!')
        self.cursor.close()
        self.show_on_table()

    def judge_validation(self, start_tm, end_tm, room_no):  #------Updated on 9/11/2022
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')
        #self.cursor_judge = DB.cursor()
        sql = """SELECT START_TM, END_TM FROM book_meeting_room_ WHERE MEETING_DT=%s and ROOM_NO=%s"""  #------Updated on 9/11/2022
        #try:
            #self.cursor_judge.execute(sql, (meeting_dt, room_no))  #--------Updated on 9/11/2022
        #except pymysql.err.OperationalError:
        reconnect_DB(self)
        self.cursor_judge=DB.cursor()
        self.cursor_judge.execute(sql, (meeting_dt, room_no))  #--------Updated on 9/11/2022

        res = self.cursor_judge.fetchall()
        self.cursor_judge.close()
        if res == ():
            return True
        for each in res:
            if not (end_tm < datetime.datetime.strptime(str(each[0]),
                                                        '%H:%M:%S') or start_tm > datetime.datetime.strptime(
                    str(each[1]), '%H:%M:%S')):
                return False
        return True


    def show_on_table(self):
        meeting_dt = datetime.datetime.strptime(self.calendarWidget.selectedDate().toString('yyyy/MM/dd'), '%Y/%m/%d')

        self.tableWidget.horizontalHeader().setSortIndicator(-1, Qt.AscendingOrder)
        self.tableWidget.sortItems(-1)
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        #print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM book_meeting_room_ WHERE MEETING_DT=%s"""
        try:
            self.cursor.execute(sql, (meeting_dt))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (meeting_dt))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res) + 20)
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 3, 4, 5, 7, 8, 9]:  #---------Updateded on 9/11/2022
                if j == 4 or j == 5:
                    total_seconds = res[i][j].total_seconds()
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60

                    delta_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
                    self.tableWidget.setItem(i, col, QTableWidgetItem(delta_str))
                else:
                    self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()
        self.tableWidget.horizontalHeaderItem(1).setText('Room No.')
        self.tableWidget.horizontalHeaderItem(3).setText('Start Time.')

    # self.tableWidget.sortItems(2, QtCore.Qt.AscendingOrder)

    def get_userDIV(self):
        self.cursor_staff = DB.cursor()
        sql = """SELECT DIVISION FROM akt_staff_ WHERE ID=%s"""
        try:
            self.cursor_staff.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_staff=DB.cursor()
            self.cursor_staff.execute(sql, (ID))

        res = self.cursor_staff.fetchall()
        self.cursor_staff.close()
        return res[0][0]

    def is_valid_time(self):
        if self.timeEdit_2.time() <= self.timeEdit.time():
            self.label_31.setText('Please input the correct time range!')
        else:
            self.label_31.setText(' ')

    def quit(self):
        BookMeetingRoom.close()

    def closeEvent(self, event):
        MainWindow.show()


class PassWindow(QWidget, Ui_PassWindow):

    def __init__(self):
        super(PassWindow, self).__init__()
        self.setupUi(self)

        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)

        self.pushButton_2.clicked.connect(self.change_psswd)
        self.pushButton_3.clicked.connect(self.quit)

    def initializing(self):
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_4.setText('')

    def change_psswd(self):
        for each_char in self.lineEdit_3.text():
            if ord(each_char) not in range(32, 127):
                QMessageBox.warning(self, 'Warning', 'Please only use A-Z, a-z, 0-9 and keyboard symbols(such as "!","$","_"... and so on...) for your new password!')
                return
        for each_char in self.lineEdit_4.text():
            if ord(each_char) not in range(32, 127):
                QMessageBox.warning(self, 'Warning',
                                    'Please only use A-Z, a-z, 0-9 or keyboard symbols(such as "!","$","_"... and so on...) for your new password!')
                return

        if self.lineEdit_2.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your old password!')
            return
        if self.lineEdit_3.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your new password!')
            return
        if self.lineEdit_4.text() == '':
            QMessageBox.warning(self, 'Warning', 'Please input your new password again!')
            return
        if self.lineEdit_3.text() != self.lineEdit_4.text():
            QMessageBox.warning(self, 'Warning',
                                'The new passwords that you input are not the same, please check it again!')
            return

        self.cursor = DB.cursor()
        sql = """SELECT PASSWORD FROM login_pass WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor=DB.cursor()
            self.cursor.execute(sql, (ID))

        pw_old = self.cursor.fetchall()[0][0]
        if self.lineEdit_2.text() != pw_old:
            QMessageBox.warning(self, 'Warning', 'The old password is not correct!')
            self.cursor.close()
            return

        sql = """UPDATE login_pass SET PASSWORD=%s WHERE ID=%s"""
        try:
            self.cursor.execute(sql, (self.lineEdit_3.text(), ID))
            DB.commit()
            QMessageBox.information(self, 'Info', 'Your password has been changed successfully!')
            self.quit()
        except:
            DB.rollback()
            QMessageBox.critical(self, 'Warning', 'Operation failed! Please check the network.')
        self.cursor.close()

    def quit(self):
        PassWindow.close()

    def closeEvent(self, event):
        MainWindow.setEnabled(True)


class ApprovePanel(QMainWindow, Ui_ApprovePanel):
    def __init__(self):
        super(ApprovePanel, self).__init__()
        self.setupUi(self)
        self.binding_btn()

    def binding_btn(self):
        self.dateEdit.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.dateEdit.setEnabled(False)
        self.comboBox_3.setEnabled(False)

        self.pushButton_8.clicked.connect(self.quit)
        self.tableWidget_3.clicked.connect(self.show_leave_panel)
        self.tableWidget_2.clicked.connect(self.show_ot_panel)
        self.tableWidget_4.clicked.connect(self.show_late_panel)
        self.tableWidget_8.clicked.connect(self.show_forget_panel)

        self.pushButton_2.clicked.connect(self.leave_accepted)
        self.pushButton_7.clicked.connect(self.leave_all_accepted)
        self.pushButton_5.clicked.connect(self.leave_declined)
        #self.pushButton_5.clicked.connect(self.show_leave_history)
        self.pushButton.clicked.connect(self.ot_accepted)
        self.pushButton_9.clicked.connect(self.ot_all_accepted)
        self.pushButton_4.clicked.connect(self.ot_declined)
        self.pushButton_3.clicked.connect(self.late_accepted)
        self.pushButton_10.clicked.connect(self.late_all_accepted)
        self.pushButton_6.clicked.connect(self.late_declined)
        self.pushButton_15.clicked.connect(self.forget_accepted)
        self.pushButton_11.clicked.connect(self.forget_all_accepted)
        self.pushButton_14.clicked.connect(self.forget_declined)

        self.checkBox.stateChanged.connect(self.mode_switch)

        self.dateEdit.dateChanged.connect(self.show_leave_history)
        self.dateEdit.dateChanged.connect(self.show_ot_history)
        self.dateEdit.dateChanged.connect(self.show_late_history)
        self.dateEdit.dateChanged.connect(self.show_forget_history)

        self.comboBox_3.currentTextChanged.connect(self.show_leave_history)
        self.comboBox_3.currentTextChanged.connect(self.show_ot_history)
        self.comboBox_3.currentTextChanged.connect(self.show_late_history)
        self.comboBox_3.currentTextChanged.connect(self.show_forget_history)

    def initializing(self):
        def move_widgets_for_frame(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_only_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def move_widgets_both(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[0] += x
            sizes[1] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_horizontal(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[2] += x
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        def extend_widgets_vertical(ob, x, y):
            geometry = ob.geometry()
            sizes = [geometry.x(), geometry.y(), geometry.width(), geometry.height()]
            sizes[3] += y
            ob.setGeometry(sizes[0], sizes[1], sizes[2], sizes[3])

        width0 = W_APPROVE
        height0 = H_APPROVE
        self.desktop = QApplication.desktop()
        screen_count = self.desktop.screenCount()  #2 screen version updating
        geometry = self.desktop.geometry()
        if screen_count == 1 or screen_count == 2:                      #2 screen version updating
            width1 = geometry.width()              #2 screen version updating
        else:                                      #2 screen version updating
            width1 = geometry.width()/2            #2 screen version updating
        height1 = geometry.height()
        x = width1 - width0
        y = height1 - height0 - 100

        self.setMaximumSize(16777215, 16777215)
        self.setMinimumSize(width1, height1)
        self.showMaximized()

        move_widgets_for_frame(self.tabWidget, x, y)

        move_widgets_for_frame(self.tableWidget_2, x, y)
        move_widgets_for_frame(self.tableWidget_3, x, y)
        move_widgets_for_frame(self.tableWidget_4, x, y)
        move_widgets_for_frame(self.tableWidget_8, x, y)

        move_widgets_only_vertical(self.pushButton_7, x, y)
        move_widgets_only_vertical(self.pushButton_9, x, y)
        move_widgets_only_vertical(self.pushButton_10, x, y)
        move_widgets_only_vertical(self.pushButton_11, x, y)

        move_widgets_both(self.pushButton, x/3, y)
        move_widgets_both(self.pushButton_2, x/3, y)
        move_widgets_both(self.pushButton_3, x/3, y)
        move_widgets_both(self.pushButton_15, x/3, y)

        move_widgets_both(self.pushButton_4, x/2, y)
        move_widgets_both(self.pushButton_5, x/2, y)
        move_widgets_both(self.pushButton_6, x/2, y)
        move_widgets_both(self.pushButton_14, x/2, y)

        move_widgets_only_horizontal(self.pushButton_8, x, y)
        extend_widgets_horizontal(self.widget, x, y)

        extend_widgets_horizontal(self.pushButton, 50, y)
        extend_widgets_horizontal(self.pushButton_2, 50, y)
        extend_widgets_horizontal(self.pushButton_4, 50, y)
        extend_widgets_horizontal(self.pushButton_5, 50, y)
        extend_widgets_horizontal(self.pushButton_3, 50, y)
        extend_widgets_horizontal(self.pushButton_15, 50, y)
        extend_widgets_horizontal(self.pushButton_6, 50, y)
        extend_widgets_horizontal(self.pushButton_14, 50, y)

        extend_widgets_horizontal(self.pushButton_7, 50, y)
        extend_widgets_horizontal(self.pushButton_9, 50, y)
        extend_widgets_horizontal(self.pushButton_10, 50, y)
        extend_widgets_horizontal(self.pushButton_11, 50, y)

        extend_widgets_vertical(self.textEdit_2, x, y)
        extend_widgets_vertical(self.textEdit, x, y)
        extend_widgets_vertical(self.textEdit_3, x, y)
        extend_widgets_vertical(self.textEdit_40, x, y)

        self.show_leave_contents()
        self.show_ot_contents()
        self.show_late_contents()
        self.show_forget_contents()

    def mode_switch(self):
        if self.checkBox.isChecked():
            self.show_leave_history()
            self.show_ot_history()
            self.show_late_history()
            self.show_forget_history()

            self.dateEdit.setEnabled(True)
            self.comboBox_3.setEnabled(True)

            self.pushButton_2.setEnabled(False)
            if HR_MODE:
                self.pushButton_5.setEnabled(True)
                self.pushButton_5.setText('Cancel Leave Request')
            else:
                self.pushButton_5.setEnabled(False)
                self.pushButton_5.setText('DECLINE')

            self.pushButton.setEnabled(False)
            self.pushButton_4.setEnabled(False)
            self.pushButton_3.setEnabled(False)
            self.pushButton_6.setEnabled(False)
            self.pushButton_15.setEnabled(False)
            self.pushButton_14.setEnabled(False)
            self.pushButton_7.setEnabled(False)
            self.pushButton_9.setEnabled(False)
            self.pushButton_10.setEnabled(False)
            self.pushButton_11.setEnabled(False)
        else:
            self.show_leave_contents()
            self.show_ot_contents()
            self.show_late_contents()
            self.show_forget_contents()

            self.dateEdit.setEnabled(False)
            self.comboBox_3.setEnabled(False)

            self.pushButton_2.setEnabled(True)
            self.pushButton_5.setEnabled(True)
            self.pushButton_5.setText('DECLINE')

            self.pushButton.setEnabled(True)
            self.pushButton_4.setEnabled(True)
            self.pushButton_3.setEnabled(True)
            self.pushButton_6.setEnabled(True)
            self.pushButton_15.setEnabled(True)
            self.pushButton_14.setEnabled(True)
            self.pushButton_7.setEnabled(True)
            self.pushButton_9.setEnabled(True)
            self.pushButton_10.setEnabled(True)
            self.pushButton_11.setEnabled(True)

    def show_leave_contents(self):
        self.tableWidget_3.clearContents()
        self.tableWidget_3.setSortingEnabled(False)
        self.tableWidget_3.setRowCount(0)
        self.tableWidget_3.setColumnCount(11)
        self.cursor_filling = DB.cursor()
        if HR_MODE:
            SQL = """SELECT * from leave_request WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (8888))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling = DB.cursor()
                self.cursor_filling.execute(SQL, (8888))
        else:
            SQL = """SELECT * from leave_request WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (ID))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling=DB.cursor()
                self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_3.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 6, 7, 8, 9, 11, 12]:
                self.tableWidget_3.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                #if col != 4 and col != 5:
                    #self.tableWidget_3.resizeColumnToContents(col)
                col += 1
            self.tableWidget_3.resizeRowToContents(i)

        self.cursor_filling.close()
        self.tableWidget_3.setSortingEnabled(True)

    def show_ot_contents(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.setSortingEnabled(False)
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.setColumnCount(9)
        self.cursor_filling = DB.cursor()

        if HR_MODE:
            SQL = """SELECT * from ot_request WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (8888))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling = DB.cursor()
                self.cursor_filling.execute(SQL, (8888))
        else:
            SQL = """SELECT * from ot_request WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (ID))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling=DB.cursor()
                self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_2.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 5, 6, 7, 8, 13]:
                self.tableWidget_2.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                # self.tableWidget_2.resizeColumnToContents(col)
                col += 1
            self.tableWidget_2.resizeRowToContents(i)

        self.cursor_filling.close()
        self.tableWidget_2.setSortingEnabled(True)

    def show_late_contents(self):
        self.tableWidget_4.clearContents()
        self.tableWidget_4.setSortingEnabled(False)
        self.tableWidget_4.setRowCount(0)
        self.tableWidget_4.setColumnCount(7)
        self.cursor_filling = DB.cursor()
        if HR_MODE:
            SQL = """SELECT * from apply_late WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (8888))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling = DB.cursor()
                self.cursor_filling.execute(SQL, (8888))
        else:
            SQL = """SELECT * from apply_late WHERE CURRENT_TO=%s"""
            try:
                self.cursor_filling.execute(SQL, (ID))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_filling = DB.cursor()
                self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_4.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 5, 6]:
                self.tableWidget_4.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                if col == 4:
                    self.tableWidget_4.resizeColumnToContents(col)
                col += 1
            self.tableWidget_4.resizeRowToContents(i)

        self.cursor_filling.close()
        self.tableWidget_4.setSortingEnabled(True)

    def show_forget_contents(self):
        self.tableWidget_8.clearContents()
        self.tableWidget_8.setSortingEnabled(False)
        self.tableWidget_8.setRowCount(0)
        self.tableWidget_8.setColumnCount(12)
        self.cursor_filling = DB.cursor()
        SQL = """SELECT * from forget_record WHERE CURRENT_TO=%s"""
        try:
            self.cursor_filling.execute(SQL, (ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_filling = DB.cursor()
            self.cursor_filling.execute(SQL, (ID))

        res = self.cursor_filling.fetchall()
        if res == ():
            self.cursor_filling.close()
            return

        self.tableWidget_8.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in [0, 2, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                value=str(res[i][j])
                if j in [5,6,7,8,9,10]:
                    if res[i][j]!=None:
                        value=res[i][j].strftime("%H:%M:%S")

                self.tableWidget_8.setItem(i, col, QTableWidgetItem(value))
                #if col != 4 and col != 5:
                #self.tableWidget_8.resizeColumnToContents(col)
                col += 1
            self.tableWidget_8.resizeRowToContents(i)

        self.cursor_filling.close()
        self.tableWidget_8.setSortingEnabled(True)
#-----------------------------------------------------
    def show_leave_history(self):
        self.tableWidget_3.clearContents()
        self.tableWidget_3.setSortingEnabled(False)
        self.tableWidget_3.setRowCount(0)
        self.tableWidget_3.setColumnCount(15)
        self.tableWidget_3.setHorizontalHeaderLabels(['No.','Staff Name','Staff ID','Type','Submitted On','Start Date','Start Time','End Date','End Time','Duration(Day)','Remarks','Leader','DM','HR','MD']) #Updated on 12/9/2023, switching HR and MD
        cur = DB.cursor()
        SQL = """SELECT ID from team_stru WHERE LEADER_ID=%s OR DM_ID=%s OR MD_ID=%s"""
        try:
            cur.execute(SQL, (ID, ID, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL, (ID, ID, ID))
        res = cur.fetchall()
        if res == () and HR_MODE == 0:
            cur.close()
            return
        staff_lst=[]
        for i in range(len(res)):
            staff_lst.append(res[i][0])

        self.staff_lst=staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')


        if HR_MODE:
            SQL = """SELECT * FROM leave_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s"""
            try:
                cur.execute(SQL, (date_min, date_max))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max))

        else:
            SQL = """SELECT * FROM leave_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
            try:
                cur.execute(SQL, (date_min, date_max, staff_lst))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_3.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4, 6, 7, 8, 9, 11, 12,13,14,16,15]:  #Updated on 12/9/2023, switching 15 and 16
                value = str(res[i][j])
                if j in [13, 14,15,16]:
                    if res[i][j] == None:
                        value = "Unconfirmed"
                    elif res[i][j] == 1:
                        value = "OK"
                    else:
                        value = "Declined"

                self.tableWidget_3.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            #if i == 10:
                #continue
            #self.tableWidget_3.resizeColumnToContents(i)
            self.tableWidget_3.resizeRowToContents(i)
        cur.close()
        self.tableWidget_3.setSortingEnabled(True)

    def show_ot_history(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.setSortingEnabled(False)
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.setColumnCount(13)
        self.tableWidget_2.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'OT Date', 'Start Time', 'Finish Time',
            'Duration(h)', 'Remarks', 'Leader', 'DM', 'HR', 'MD'])   #Updated on 12/9/2023, switching HR and MD
        cur = DB.cursor()

        if HR_MODE:
            staff_lst=[]
        else:
            staff_lst=self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        if HR_MODE:
            SQL = """SELECT * FROM ot_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s"""
            try:
                cur.execute(SQL, (date_min, date_max))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max))
        else:
            SQL = """SELECT * FROM ot_request WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
            try:
                cur.execute(SQL, (date_min, date_max, staff_lst))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_2.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 5, 6, 7, 8, 13,9,10,12,11]:  #Updated on 12/9/2023, switching 11 and 12
                value = str(res[i][j])
                if j in [9, 10, 11, 12]:
                    if res[i][j] == None:
                        value = "Unconfirmed"
                    elif res[i][j] == 1:
                        value = "OK"
                    else:
                        value = "Declined"

                self.tableWidget_2.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            #if i == 8:
                #continue
            #self.tableWidget_2.resizeColumnToContents(i)
            self.tableWidget_2.resizeRowToContents(i)
        cur.close()
        self.tableWidget_2.setSortingEnabled(True)

    def show_late_history(self):
        self.tableWidget_4.clearContents()
        self.tableWidget_4.setSortingEnabled(False)
        self.tableWidget_4.setRowCount(0)
        self.tableWidget_4.setColumnCount(10)
        self.tableWidget_4.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'Date of Late Clock-In', 'Clock-In Time', 'Remarks', 'Leader','DM','HR'])
        cur = DB.cursor()

        if HR_MODE:
            staff_lst=[]
        else:
            staff_lst = self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        if HR_MODE:
            SQL = """SELECT * FROM apply_late WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s"""
            try:
                cur.execute(SQL, (date_min, date_max))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max))
        else:
            SQL = """SELECT * FROM apply_late WHERE APPLY_DTTM>=%s AND APPLY_DTTM<%s AND USER_ID IN %s"""
            try:
                cur.execute(SQL, (date_min, date_max, staff_lst))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_4.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4,5, 6, 7, 8, 9]:
                value = str(res[i][j])
                if j in [7,8,9]:
                    if res[i][j] == None:
                        value = "Unconfirmed"

                self.tableWidget_4.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        self.tableWidget_4.resizeColumnToContents(4)
        for i in range(len(res)):
            # if i == 8:
            # continue
            # self.tableWidget_4.resizeColumnToContents(i)
            self.tableWidget_4.resizeRowToContents(i)
        cur.close()
        self.tableWidget_4.setSortingEnabled(True)

    def show_forget_history(self):
        self.tableWidget_8.clearContents()
        self.tableWidget_8.setSortingEnabled(False)
        self.tableWidget_8.setRowCount(0)
        self.tableWidget_8.setColumnCount(15)
        self.tableWidget_8.setHorizontalHeaderLabels(
            ['No.', 'Staff Name', 'Staff ID', 'Submitted On', 'Clock Date', 'Clock In', 'Clock Out',
             'Out-1', 'In-1','Out-2','In-2','Remarks','Leader','DM','HR'])
        cur = DB.cursor()

        if HR_MODE:
            staff_lst = []
        else:
            staff_lst = self.staff_lst

        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        if HR_MODE:
            SQL = """SELECT * FROM forget_record WHERE REQUEST_DTTM>=%s AND REQUEST_DTTM<%s"""
            try:
                cur.execute(SQL, (date_min, date_max))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max))
        else:
            SQL = """SELECT * FROM forget_record WHERE REQUEST_DTTM>=%s AND REQUEST_DTTM<%s AND USER_ID IN %s"""
            try:
                cur.execute(SQL, (date_min, date_max, staff_lst))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (date_min, date_max, staff_lst))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        self.tableWidget_8.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 2, 1, 3, 4, 5, 6, 7, 8, 9,10,11,12,13,14]:
                value = str(res[i][j])
                if j in [12, 13, 14]:
                    if res[i][j] == None:
                        value = "Unconfirmed"

                self.tableWidget_8.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(len(res)):
            # if i == 8:
            # continue
            # self.tableWidget_8.resizeColumnToContents(i)
            self.tableWidget_8.resizeRowToContents(i)
        cur.close()
        self.tableWidget_8.setSortingEnabled(True)

#-----------------------------------------------------
    def show_leave_panel(self):
        self.textEdit_11.setText('')
        self.textEdit_12.setText('')
        self.textEdit_16.setText('')
        self.textEdit_13.setText('')
        self.textEdit_14.setText('')
        self.textEdit_15.setText('')
        self.textEdit_2.setText('')

        index = self.tableWidget_3.currentRow()
        try:
            staff_name = self.tableWidget_3.item(index, 1).text()
            staff_id = self.tableWidget_3.item(index, 2).text()
            type = self.tableWidget_3.item(index, 3).text()
            start_dt = self.tableWidget_3.item(index, 5).text()
            start_len = self.tableWidget_3.item(index, 6).text()
            end_dt = self.tableWidget_3.item(index, 7).text()
            end_len = self.tableWidget_3.item(index, 8).text()
            during = self.tableWidget_3.item(index, 9).text()
            remarks = self.tableWidget_3.item(index, 10).text()
        except:
            return

        self.textEdit_11.setText(staff_name)
        self.textEdit_12.setText(staff_id)
        self.textEdit_16.setText(type)
        self.textEdit_13.setText((str(start_dt) + ' (' + str(start_len) + ')').replace('all', 'all day'))
        if end_dt == 'None':
            self.textEdit_14.setText('(Single day leave)')
        else:
            self.textEdit_14.setText((str(end_dt) + ' (' + str(end_len) + ')').replace('all', 'all day'))
        self.textEdit_15.setText(during)
        self.textEdit_2.setText(remarks)

    def show_ot_panel(self):
        self.textEdit_10.setText('')
        self.textEdit_9.setText('')
        self.textEdit_6.setText('')
        self.textEdit_7.setText('')
        self.textEdit_8.setText('')
        self.textEdit_5.setText('')
        self.textEdit.setText('')

        index = self.tableWidget_2.currentRow()
        try:
            staff_name = self.tableWidget_2.item(index, 1).text()
            staff_id = self.tableWidget_2.item(index, 2).text()
            ot_dt = self.tableWidget_2.item(index, 4).text()
            start_tm = self.tableWidget_2.item(index, 5).text()
            end_tm = self.tableWidget_2.item(index, 6).text()
            during = self.tableWidget_2.item(index, 7).text()
            remarks = self.tableWidget_2.item(index, 8).text()
        except:
            return

        self.textEdit_10.setText(staff_name)
        self.textEdit_9.setText(staff_id)
        self.textEdit_6.setText(ot_dt)
        self.textEdit_7.setText(start_tm)
        self.textEdit_8.setText(end_tm)
        self.textEdit_5.setText(during)
        self.textEdit.setText(remarks)

    def show_late_panel(self):
        self.textEdit_17.setText('')
        self.textEdit_21.setText('')
        self.textEdit_22.setText('')
        self.textEdit_18.setText('')

        index = self.tableWidget_4.currentRow()
        try:
            staff_name = self.tableWidget_4.item(index, 1).text()
            staff_id = self.tableWidget_4.item(index, 2).text()
            late_dt = self.tableWidget_4.item(index, 4).text()
            clockin_tm = self.tableWidget_4.item(index, 5).text()
            remarks = self.tableWidget_4.item(index, 6).text()
        except:
            return

        self.textEdit_17.setText(staff_name)
        self.textEdit_21.setText(staff_id)
        self.textEdit_22.setText(late_dt)
        self.textEdit_18.setText(clockin_tm)
        self.textEdit_3.setText(remarks)

    def show_forget_panel(self):
        self.textEdit_39.setText('')
        self.textEdit_44.setText('')
        self.textEdit_45.setText('')
        self.textEdit_41.setText('')
        self.textEdit_42.setText('')
        self.textEdit_43.setText('')
        self.textEdit_46.setText('')
        self.textEdit_48.setText('')
        self.textEdit_47.setText('')
        self.textEdit_40.setText('')

        index = self.tableWidget_8.currentRow()
        try:
            staff_name = self.tableWidget_8.item(index, 1).text()
            staff_id = self.tableWidget_8.item(index, 2).text()
            clock_date = self.tableWidget_8.item(index, 4).text()
            clock_in = self.tableWidget_8.item(index, 5).text()
            clock_out = self.tableWidget_8.item(index, 6).text()
            out1=self.tableWidget_8.item(index, 7).text()
            in1=self.tableWidget_8.item(index, 8).text()
            out2=self.tableWidget_8.item(index, 9).text()
            in2=self.tableWidget_8.item(index, 10).text()
            remarks=self.tableWidget_8.item(index, 11).text()
        except:
            return

        self.textEdit_39.setText(staff_name)
        self.textEdit_44.setText(staff_id)
        self.textEdit_45.setText(clock_date)
        self.textEdit_41.setText(clock_in)
        self.textEdit_42.setText(clock_out)
        self.textEdit_43.setText(out1)
        self.textEdit_46.setText(in1)
        self.textEdit_48.setText(out2)
        self.textEdit_47.setText(in2)
        self.textEdit_40.setText(remarks)
#------------------------------------------------------
    def leave_accepted(self):
        index = self.tableWidget_3.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_3.item(index, 0).text()
        staff_id_ = self.tableWidget_3.item(index, 2).text()  #updated for 1.1
        type_ = self.tableWidget_3.item(index, 3).text() #updated for 1.1
        during_ = float(self.tableWidget_3.item(index, 9).text()) #updated for 1.1

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:  #如果存在DM
                if res[0][0] == res[0][1]:  #如果LEADER和DM是同一个人的话
                    if res[0][1] == res[0][2]:  #如果DM和MD又是同一个人的话
                        #current_to = 9999
                        #current_po = 'HR'
                        current_to = 8888
                        current_po = 'HR'
                    else:  #如果LEADER和DM是同一个人，但DM和MD不是同一个人
                        #current_to = res[0][2]
                        #current_po = 'MD'
                        current_to = 8888
                        current_po = 'HR'

                else:  #如果LEADER和DM不是同一个人
                    current_to = res[0][1]
                    current_po = 'DM'
            else:   #如果不存在DM
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 8888
                current_po = 'HR'

        elif current_po == 'DM':
            if res[0][2] == res[0][1]:  #如果DM和MD是同一人
                #current_to = 9999
                #current_po = 'HR'
                current_to = 8888
                current_po = 'HR'
            else:   #如果DM和MD不是同一人
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 8888
                current_po = 'HR'

        elif current_po == 'HR':
            if res[0][2] == res[0][1]:  #如果DM和MD是同一人
                current_to = 9999
                current_po = 'PA'
            else:  # 如果DM和MD不是同一人
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 9999
                current_po = 'PA'

        else:
            current_to = 9999
            current_po = 'PA'

        if current_po == 'DM':
            SQL = """UPDATE leave_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))

        elif current_po == 'HR':
            SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))

        elif current_po == 'MD':
            SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, current_to, current_po, request_id))
        else:
            SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

        DB.commit()

        self.cursor_approve.close()
        self.show_leave_contents()
        self.show_leave_panel()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2=query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='leave')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:

            if type_== 'Annual leave':
                self.cursor_calc = DB.cursor()
                SQL="""SELECT AN_DAYS FROM akt_staff_ WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res=self.cursor_calc.fetchall()
                if res==():
                    pass
                else:
                    an_days=res[0][0]
                    an_days-=float(during_)
                    SQL="""UPDATE akt_staff_ SET AN_DAYS=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (an_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
            elif type_=='Sick leave':
                self.cursor_calc = DB.cursor()
                SQL = """SELECT SICK_DAYS FROM akt_staff_ WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res = self.cursor_calc.fetchall()
                if res == ():
                    pass
                else:
                    sick_days = res[0][0]
                    sick_days -= float(during_)
                    SQL = """UPDATE akt_staff_ SET SICK_DAYS=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (sick_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
            elif type_=='Hometown':
                self.cursor_calc = DB.cursor()
                SQL = """SELECT HOME_TOWN FROM akt_staff_ WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res = self.cursor_calc.fetchall()
                if res == ():
                    pass
                else:
                    home_days = res[0][0]
                    home_days -= 1
                    SQL = """UPDATE akt_staff_ SET HOME_TOWN=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (home_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()
            elif type_=='Personal leave':
                self.cursor_calc = DB.cursor()
                SQL = """SELECT PERSONAL_DAYS FROM akt_staff_ WHERE ID=%s"""
                self.cursor_calc.execute(SQL, (staff_id_))
                res = self.cursor_calc.fetchall()
                if res == ():
                    pass
                else:
                    personal_days = res[0][0]
                    personal_days -= float(during_)
                    SQL = """UPDATE akt_staff_ SET PERSONAL_DAYS=%s WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (personal_days, staff_id_))
                    DB.commit()
                self.cursor_calc.close()

            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='leave')
        # ===========================================================


        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def leave_all_accepted(self):
        data = []
        for i in range(self.tableWidget_3.rowCount()):
            line = []
            for j in range(self.tableWidget_3.columnCount()):
                line.append(self.tableWidget_3.item(i, j).text())
            data.append(line)

        if not data:
            QMessageBox.warning(self, 'Warning', 'No unconfirmed leave request remains on the table!')
            return

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept all of the leave requests?')
        if a == QMessageBox.No:
            return

        self.allLeave_accepted = Leave_All_Accepted(data=data)
        self.allLeave_accepted.finish_box.connect(self.finish_msgbox)
        self.allLeave_accepted.update_label.connect(Monitor.update_text)
        self.allLeave_accepted.update_progress.connect(Monitor.update_progressbar)
        self.allLeave_accepted.monitor_close.connect(Monitor.monitor_close_approve)
        self.allLeave_accepted.monitor_open.connect(self.monitor_show)

        self.allLeave_accepted.start()
        self.monitor_show()

    def leave_declined(self):
        index = self.tableWidget_3.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        if self.pushButton_5.text()=="Cancel Leave Request":
            request_id = self.tableWidget_3.item(index, 0).text()
            a = QMessageBox.question(self, 'Confirmation',
                                     f'Are you sure to cancel the selected request?\nRequest ID: {request_id}\n')
            if a == QMessageBox.No:
                return

            self.cursor_cancel = DB.cursor()
            SQL = """SELECT TYPE, DURING, USER_ID, CURRENT_TO FROM leave_request WHERE SERIAL=%s"""
            try:
                self.cursor_cancel.execute(SQL, (request_id))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_cancel=DB.cursor()
                self.cursor_cancel.execute(SQL, (request_id))

            res_po = self.cursor_cancel.fetchall()
            leave_type = str(res_po[0][0]).strip()
            duration = float(res_po[0][1])
            user_id = int(res_po[0][2])
            current_to = int(res_po[0][3])

            #print(leave_type, duration, user_id)

            if leave_type=='Annual leave' and current_to == 9999:
                SQL = """SELECT AN_DAYS FROM akt_staff_ WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (user_id))
                res_po = self.cursor_cancel.fetchall()
                an_days = float(res_po[0][0])
                an_days += duration
                #print(an_days)
                SQL = """UPDATE akt_staff_ SET AN_DAYS=%s WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (an_days, user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (an_days, user_id))

            elif leave_type=='Sick leave' and current_to == 9999:
                SQL = """SELECT SICK_DAYS FROM akt_staff_ WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (user_id))
                res_po = self.cursor_cancel.fetchall()
                sick_days = float(res_po[0][0])
                sick_days += duration
                #print(sick_days)
                SQL = """UPDATE akt_staff_ SET SICK_DAYS=%s WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (sick_days, user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (sick_days, user_id))

            elif leave_type=='Personal leave' and current_to == 9999:
                SQL = """SELECT PERSONAL_DAYS FROM akt_staff_ WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (user_id))
                res_po = self.cursor_cancel.fetchall()
                personal_days = float(res_po[0][0])
                personal_days += duration
                #print(personal_days)
                SQL = """UPDATE akt_staff_ SET PERSONAL_DAYS=%s WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (personal_days, user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (personal_days, user_id))

            elif leave_type=='Hometown' and current_to == 9999:
                SQL = """SELECT HOME_TOWN FROM akt_staff_ WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (user_id))
                res_po = self.cursor_cancel.fetchall()
                home_town = int(res_po[0][0])
                home_town += 1
                #print(home_town)
                SQL = """UPDATE akt_staff_ SET HOME_TOWN=%s WHERE ID=%s"""
                try:
                    self.cursor_cancel.execute(SQL, (home_town, user_id))
                except pymysql.err.OperationalError:
                    reconnect_DB(self)
                    self.cursor_cancel = DB.cursor()
                    self.cursor_cancel.execute(SQL, (home_town, user_id))

            DB.commit()

            SQL="""DELETE FROM leave_request WHERE SERIAL=%s"""
            try:
                self.cursor_cancel.execute(SQL, (int(request_id)))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_cancel = DB.cursor()
                self.cursor_cancel.execute(SQL, (int(request_id)))
                DB.commit()

            self.cursor_cancel.close()
            self.show_leave_history()
            self.show_leave_panel()

            return

        request_id = self.tableWidget_3.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM leave_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE leave_request SET LEADER=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        elif current_po == 'DM':
            SQL = """UPDATE leave_request SET DM=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        elif current_po == 'HR':
            SQL = """UPDATE leave_request SET HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        else:
            SQL = """UPDATE leave_request SET MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 9998, 'NG', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_leave_contents()
        self.show_leave_panel()

        # Mail sending 2==============================================
        info_lst2=query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='leave')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def ot_accepted(self):
        index = self.tableWidget_2.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_2.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:  #如果存在DM
                if res[0][0] == res[0][1]: #如果LEADER和DM是同一人
                    if res[0][1] == res[0][2]: #如果DM和MD是同一人
                        #current_to = 9999
                        #current_po = 'HR'
                        current_to = 8888
                        current_po = 'HR'
                    else:   #如果DM和MD不是同一人
                        #current_to = res[0][2]
                        #current_po = 'MD'
                        current_to = 8888
                        current_po = 'HR'
                else: #如果LEADER和DM不是同一人
                    current_to = res[0][1]
                    current_po = 'DM'
            else:    #如果不存在DM
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 8888
                current_po = 'HR'
        elif current_po == 'DM':
            if res[0][2] == res[0][1]:  #如果DM和MD是同一人
                #current_to = 9999
                #current_po = 'HR'
                current_to = 8888
                current_po = 'HR'
            else:  #如果DM和MD不是同一人
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 8888
                current_po = 'HR'

        elif current_po == 'HR':
            if res[0][2] == res[0][1]:  # 如果DM和MD是同一人
                current_to = 9999
                current_po = 'PA'
            else:  # 如果DM和MD不是同一人
                #current_to = res[0][2]
                #current_po = 'MD'
                current_to = 9999
                current_po = 'PA'

        else:
            current_to = 9999
            current_po = 'PA'

        if current_po == 'DM':
            SQL = """UPDATE ot_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))
        elif current_po == 'HR':
            SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))
        elif current_po == 'MD':
            SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, current_to, current_po, request_id))
        else:
            SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_ot_contents()
        self.show_ot_panel()
        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2= query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='ot')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='ot')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def ot_all_accepted(self):
        #QMessageBox.information(self, 'Info', 'Sorry, this function is under development, please wait for the next version.')
        data = []
        for i in range(self.tableWidget_2.rowCount()):
            line = []
            for j in range(self.tableWidget_2.columnCount()):
                line.append(self.tableWidget_2.item(i, j).text())
            data.append(line)

        if not data:
            QMessageBox.warning(self, 'Warning', 'No unconfirmed OT request remains on the table!')
            return

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept all of the OT requests?')
        if a == QMessageBox.No:
            return

        self.allOt_accepted = Ot_All_Accepted(data=data)
        self.allOt_accepted.finish_box.connect(self.finish_msgbox)
        self.allOt_accepted.update_label.connect(Monitor.update_text)
        self.allOt_accepted.update_progress.connect(Monitor.update_progressbar)
        self.allOt_accepted.monitor_close.connect(Monitor.monitor_close_approve)
        self.allOt_accepted.monitor_open.connect(self.monitor_show)

        self.allOt_accepted.start()
        self.monitor_show()

    def ot_declined(self):
        index = self.tableWidget_2.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_2.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM ot_request WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve=DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE ot_request SET LEADER=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        elif current_po == 'DM':
            SQL = """UPDATE ot_request SET DM=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        elif current_po == 'HR':
            SQL = """UPDATE ot_request SET HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 0, 9998, 'NG', request_id))
        else:
            SQL = """UPDATE ot_request SET MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, (0, 9998, 'NG', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_ot_contents()
        self.show_ot_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='ot')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def late_accepted(self):
        index = self.tableWidget_4.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_4.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0: #如果存在DM
                if res[0][0] == res[0][1]: #如果LEADER跟DM是同一人
                    #current_to = 9999
                    #current_po = 'HR'
                    current_to = 8888
                    current_po = 'HR'

                else: #如果LEADER跟DM不是同一人
                    current_to = res[0][1]
                    current_po = 'DM'

            else: #如果不存在DM
                #current_to = 9999
                #current_po = 'HR'
                current_to = 8888
                current_po = 'HR'

        elif current_po == 'DM':
            current_to = 8888
            current_po = 'HR'

        else:
            current_to = 9999
            current_po = 'PA'

        if current_po == 'DM':
            SQL = """UPDATE apply_late SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))
        elif current_po == 'HR':
            SQL = """UPDATE apply_late SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', 'OK', current_to, current_po, request_id))
        else:
            SQL = """UPDATE apply_late SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_late_contents()
        self.show_late_panel()
        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2 = query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='late')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='late')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def late_all_accepted(self):
        #QMessageBox.information(self, 'Info', 'Sorry, this function is under development, please wait for the next version.')
        data = []
        for i in range(self.tableWidget_4.rowCount()):
            line = []
            for j in range(self.tableWidget_4.columnCount()):
                line.append(self.tableWidget_4.item(i, j).text())
            data.append(line)

        if not data:
            QMessageBox.warning(self, 'Warning', 'No unconfirmed late clock-in request remains on the table!')
            return

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept all of the late clock-in requests?')
        if a == QMessageBox.No:
            return

        self.allLate_accepted = Late_All_Accepted(data=data)
        self.allLate_accepted.finish_box.connect(self.finish_msgbox)
        self.allLate_accepted.update_label.connect(Monitor.update_text)
        self.allLate_accepted.update_progress.connect(Monitor.update_progressbar)
        self.allLate_accepted.monitor_close.connect(Monitor.monitor_close_approve)
        self.allLate_accepted.monitor_open.connect(self.monitor_show)

        self.allLate_accepted.start()
        self.monitor_show()


    def late_declined(self):
        index = self.tableWidget_4.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_4.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE apply_late SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 9998, 'NG', request_id))
        elif current_po == 'DM':
            SQL = """UPDATE apply_late SET DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 9998, 'NG', request_id))
        else:
            SQL = """UPDATE apply_late SET HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 9998, 'NG', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_late_contents()
        self.show_late_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='late')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def forget_accepted(self):
        index = self.tableWidget_8.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_8.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        self.cursor_approve.execute(SQL, (staff_id))
        res = self.cursor_approve.fetchall()

        if current_po == 'LEADER':
            if res[0][1] != 0:
                if res[0][0] == res[0][1]:
                    current_to = 9999
                    current_po = 'HR'
                else:
                    current_to = res[0][1]
                    current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'
        else:
            current_to = 9999
            current_po = 'HR'

        if current_po == 'DM':
            SQL = """UPDATE forget_record SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))

        else:
            SQL = """UPDATE forget_record SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_forget_contents()
        self.show_forget_panel()

        if current_to == 9999:
            SQL="""SELECT USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2 FROM forget_record WHERE SERIAL=%s"""
            self.cursor_migrant=DB.cursor()
            self.cursor_migrant.execute(SQL, (request_id))
            data=self.cursor_migrant.fetchall()
            user_id=data[0][0]
            clockin=data[0][1]
            clockout=data[0][2]
            out1=data[0][3]
            in1=data[0][4]
            out2=data[0][5]
            in2=data[0][6]

            try:
                serial_timecard=str(user_id)+clockin.strftime('%Y%m%d')
            except AttributeError:
                serial_timecard=str(user_id)+clockout.strftime('%Y%m%d')

            if clockin==None:
                SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data = self.cursor_migrant.fetchall()
                if data != ():
                    SQL = """UPDATE time_card SET CLOCK_OUT=%s WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (clockout, serial_timecard))
                    DB.commit()
                else:
                    SQL="""INSERT INTO time_card (SERIAL,USER_ID,CLOCK_OUT) VALUES (%s,%s,%s)"""
                    self.cursor_migrant.execute(SQL,(serial_timecard,user_id,clockout))
                    DB.commit()

            elif clockout==None:
                SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data = self.cursor_migrant.fetchall()
                if data != ():
                    SQL = """UPDATE time_card SET CLOCK_IN=%s WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (clockin, serial_timecard))
                    DB.commit()
                else:
                    SQL = """INSERT INTO time_card (SERIAL,USER_ID,CLOCK_IN) VALUES (%s,%s,%s)"""
                    self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin))
                    DB.commit()

            else:
                SQL="""SELECT * FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                data=self.cursor_migrant.fetchall()
                if data!=():
                    SQL="""DELETE FROM time_card WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (serial_timecard))
                    DB.commit()

                SQL="""INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
                self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin, clockout, out1, in1, out2, in2))
                DB.commit()

            self.cursor_migrant.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        info_lst2 = query_email(id=staff_id)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=info_lst2[0],
                                         mode='forget')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            if info_lst2 == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst2[1],
                                              receiver_name=info_lst2[0],
                                              mode='forget')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been accepted!')

    def forget_all_accepted(self):
        #QMessageBox.information(self, 'Info', 'Sorry, this function is under development, please wait for the next version.')
        data = []
        for i in range(self.tableWidget_8.rowCount()):
            line = []
            for j in range(self.tableWidget_8.columnCount()):
                line.append(self.tableWidget_8.item(i, j).text())
            data.append(line)

        if not data:
            QMessageBox.warning(self, 'Warning', 'No unconfirmed time record adding request remains on the table!')
            return

        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to accept all of the time record adding requests?')
        if a == QMessageBox.No:
            return

        self.allForget_accepted = Forget_All_Accepted(data=data)
        self.allForget_accepted.finish_box.connect(self.finish_msgbox)
        self.allForget_accepted.update_label.connect(Monitor.update_text)
        self.allForget_accepted.update_progress.connect(Monitor.update_progressbar)
        self.allForget_accepted.monitor_close.connect(Monitor.monitor_close_approve)
        self.allForget_accepted.monitor_open.connect(self.monitor_show)

        self.allForget_accepted.start()
        self.monitor_show()


    def forget_declined(self):
        index = self.tableWidget_8.currentRow()
        if index == -1:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a record first!')
            return

        request_id = self.tableWidget_8.item(index, 0).text()
        a = QMessageBox.question(self, 'Confirmation',
                                 f'Are you sure to decline the selected request?\nRequest ID: {request_id}\n')
        if a == QMessageBox.No:
            return

        self.cursor_approve = DB.cursor()
        SQL = """SELECT CURRENT_PO, USER_ID FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor_approve.execute(SQL, (request_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_approve = DB.cursor()
            self.cursor_approve.execute(SQL, (request_id))

        res_po = self.cursor_approve.fetchall()
        current_po = res_po[0][0]
        staff_id = res_po[0][1]

        if current_po == 'LEADER':
            SQL = """UPDATE forget_record SET LEADER=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))
        else:
            SQL = """UPDATE forget_record SET DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
            self.cursor_approve.execute(SQL, ('Declined', 'Declined', 9998, 'HR', request_id))

        DB.commit()
        self.cursor_approve.close()
        self.show_forget_contents()
        self.show_forget_panel()

        # Mail sending 2==============================================
        info_lst2 = query_email(id=staff_id)
        if info_lst2 == -1:
            pass
        else:
            mailsender.send_declined_mail(email_add=info_lst2[1],
                                          receiver_name=info_lst2[0],
                                          mode='forget')
        # ===========================================================
        QMessageBox.information(self, 'Info', 'Request has been declined!')

    def monitor_show(self):
        Monitor.show()
        Monitor.initializing()
        self.setEnabled(False)

    def finish_msgbox(self, title, text):
        QMessageBox.information(self, title, text)

    def quit(self):
        ApprovePanel.close()

    def closeEvent(self, event):
        self.setupUi(self)
        self.binding_btn()
        MainWindow.show()

class ApplyLateClockIn(QDialog, Ui_ApplyLateClockIn):
    def __init__(self):
        super(ApplyLateClockIn, self).__init__()
        self.setupUi(self)

        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        self.timeEdit_2.setTime(QTime.fromString('09:00', 'hh:mm'))
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)

        self.pushButton_3.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.apply_late)
        self.dateEdit.dateChanged.connect(self.show_tableData)
        self.comboBox_3.currentTextChanged.connect(self.show_tableData)
        self.pushButton_2.clicked.connect(self.withdraw_apply)

    def initialize(self):
        self.show_tableData()

    def show_tableData(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM apply_late WHERE LATE_DT>=%s AND LATE_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            table_col=0
            for j in [0, 3, 4, 5, 6, 7, 8, 9]:
                value=str(res[i][j])
                if j in [7, 8, 9]:
                    if str(res[i][j])=="None":
                        value="Unconfirmed"
                    elif str(res[i][j])=="OK":
                        value="OK"
                    else:
                        value="Declined"

                self.tableWidget.setItem(i, table_col, QTableWidgetItem(value))
                table_col+=1

        for i in range(8):
            if i==4:
                continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()

    def apply_late(self):
        late_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        late_dt_compare=late_dt.strftime('%Y%m%d')
        now_dt_compare=TimeCard.t.strftime('%Y%m%d')
        if late_dt_compare<=now_dt_compare:
            QMessageBox.critical(self, 'Warning','You can only apply for late clock-in 1 day beforehand.')
            return

        user_id = ID
        user_name = MainWindow.name
        apply_dttm = TimeCard.t
        clockin_tm = datetime.datetime.strptime(self.timeEdit_2.time().toString('hh:mm'), '%H:%M')
        remarks = self.textEdit.toPlainText()

        msm = f'Apply late clock-in?\n' \
              f'Late date: {late_dt.strftime("%d/%m/%Y")}\n' \
              f'Clock-in time: {clockin_tm.strftime("%H:%M:%S")}\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor_apply=DB.cursor()
        SQL="""SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor_apply.execute(SQL, (user_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_apply=DB.cursor()
            self.cursor_apply.execute(SQL, (user_id))

        res=self.cursor_apply.fetchall()

        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 8888
                current_po = 'HR'

        if current_to!=9999:
            sql = """INSERT INTO apply_late (USER_ID, USER_NAME, APPLY_DTTM, LATE_DT, CLOCKIN_TM, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
                (%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (user_id, user_name, apply_dttm, late_dt, clockin_tm, remarks, current_to, current_po))
            DB.commit()
        else:
            sql = """INSERT INTO apply_late (USER_ID, USER_NAME, APPLY_DTTM, LATE_DT, CLOCKIN_TM, REMARKS,LEADER, DM, HR, CURRENT_TO, CURRENT_PO) VALUES
                            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
            user_id, user_name, apply_dttm, late_dt, clockin_tm, remarks, 'OK','OK','OK',current_to, current_po))
            DB.commit()

        self.cursor_apply.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='late')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to==9999:
            info_lst = query_email(id=ID)
            if info_lst == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst[1],
                                              receiver_name=info_lst[0],
                                              mode='late')
        # ===========================================================

        QMessageBox.information(self, 'Info',
                                'Late clock-in request has been sent successfully! Please wait for the approvements.')
        self.show_tableData()

    def withdraw_apply(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 7).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM apply_late WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.cursor.close()
        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.show_tableData()

    def quit(self):
        ApplyLateClockIn.close()

    def closeEvent(self, event):
        TimeCard.show()
        TimeCard.calendarWidget.setFocus()

class ForgetRecord(QDialog, Ui_ForgetRecord):
    def __init__(self):
        super(ForgetRecord, self).__init__()
        self.setupUi(self)

        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit_3.setDate(QDate.currentDate())
        month = time.strftime("%m", time.localtime(time.time()))
        self.comboBox_3.setCurrentText(month)
        self.timeEdit_2.setTime(QTime.fromString('08:30', 'hh:mm'))
        self.timeEdit_3.setTime(QTime.fromString('17:45', 'hh:mm'))
        self.timeEdit_4.setTime(QTime.fromString('12:00', 'hh:mm'))
        self.timeEdit_5.setTime(QTime.fromString('13:00', 'hh:mm'))
        self.timeEdit_6.setTime(QTime.fromString('14:00', 'hh:mm'))
        self.timeEdit_7.setTime(QTime.fromString('15:00', 'hh:mm'))
        self.checkBox_2.setEnabled(False)
        self.timeEdit_4.setEnabled(False)
        self.timeEdit_5.setEnabled(False)
        self.timeEdit_6.setEnabled(False)
        self.timeEdit_7.setEnabled(False)
        self.checkBox.setChecked(False)
        self.label_31.setText('')

        clockin = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_2.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        clockout = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_3.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')

        data_pre_input = [None, None, None, None, None, None, None, None, None]

        data_pre_input[2] = clockin
        data_pre_input[3] = clockout

        data_set = calculate_without_approved_ot(data_line=data_pre_input)
        work_time = data_set[0]
        over_time = data_set[1]
        self.textEdit_3.setText(str(work_time))
        self.textEdit_4.setText(str(over_time))


        self.timeEdit_2.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_3.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_4.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_5.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_6.timeChanged.connect(self.time_stamp_check)
        self.timeEdit_7.timeChanged.connect(self.time_stamp_check)
        self.checkBox.stateChanged.connect(self.time_stamp_check)
        self.checkBox_2.stateChanged.connect(self.time_stamp_check)
        self.checkBox_3.stateChanged.connect(self.time_stamp_check)
        self.checkBox_4.stateChanged.connect(self.time_stamp_check)
        self.checkBox_3.stateChanged.connect(self.singleTime_or_not)
        self.checkBox_4.stateChanged.connect(self.singleTime_or_not)

        self.checkBox_3.stateChanged.connect(self.uncheck_clock_in)
        self.checkBox_4.stateChanged.connect(self.uncheck_clock_out)

        self.checkBox.stateChanged.connect(self.if_in_out_1)
        self.checkBox_2.stateChanged.connect(self.if_in_out_2)
        self.pushButton_3.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.submit)
        self.dateEdit.dateChanged.connect(self.show_tableData)
        self.comboBox_3.currentTextChanged.connect(self.show_tableData)
        self.pushButton_2.clicked.connect(self.withdraw_apply)

    def test(self):
        print(self.timeEdit_2.text())
        print(self.timeEdit_2.text().replace(':',''))
        print(type(self.timeEdit_2.text()))

    def initialize(self):
        self.show_tableData()

    def show_tableData(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(1)
        if self.comboBox_3.currentText() == 'All':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '0101', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        elif self.comboBox_3.currentText() == '12':
            date_min = datetime.datetime.strptime(self.dateEdit.text() + '1201', '%Y%m%d')
            date_max = datetime.datetime.strptime(str(int(self.dateEdit.text()) + 1) + '0101', '%Y%m%d')

        else:
            date_min = datetime.datetime.strptime(self.dateEdit.text() + self.comboBox_3.currentText() + '01', '%Y%m%d')
            date_max = datetime.datetime.strptime(
                self.dateEdit.text() + str(int(self.comboBox_3.currentText()) + 1) + '01', '%Y%m%d')

        self.cursor = DB.cursor()
        sql = """SELECT * FROM forget_record WHERE CLOCK_DT>=%s AND CLOCK_DT<%s AND USER_ID=%s"""
        try:
            self.cursor.execute(sql, (date_min, date_max, ID))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (date_min, date_max, ID))

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            table_col = 0
            for j in [0, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                value = str(res[i][j])
                if j in [12, 13, 14]:
                    if str(res[i][j]) == "None":
                        value = "Unconfirmed"
                    elif str(res[i][j]) == "OK":
                        value = "OK"
                    else:
                        value = "Declined"

                if j in [5,6,7,8,9,10]:
                    if res[i][j]!=None:
                        value=res[i][j].strftime("%H:%M:%S")

                self.tableWidget.setItem(i, table_col, QTableWidgetItem(value))
                table_col += 1

        for i in range(13):
            if i == 9:
                continue
            self.tableWidget.resizeColumnToContents(i)
        self.cursor.close()

    def submit(self):
        if self.label_31.text()!='':
            QMessageBox.critical(self, 'Error', 'Error: Please input the correct time range!')
            return

        forgot_dt = datetime.datetime.strptime(self.dateEdit_3.text(), '%d/%m/%Y')
        forgot_dt_compare = forgot_dt.strftime('%Y%m%d')
        now_dt_compare = TimeCard.t.strftime('%Y%m%d')
        if forgot_dt_compare > now_dt_compare:
            QMessageBox.critical(self, 'Error', 'Error: You can not submit forget-time-record report for the future\'s date!')
            return

        user_id = ID
        user_name = MainWindow.name
        request_dttm = TimeCard.t
        clock_dt=forgot_dt
        clockin = datetime.datetime.strptime(self.dateEdit_3.date().toString('yyyy/MM/dd/')+self.timeEdit_2.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        clockout = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_3.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out1=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_4.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in1=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_5.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out2=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_6.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in2=datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_7.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        remarks = self.textEdit.toPlainText()

        if not self.checkBox.isChecked():
            out1=None
            in1=None

        if not self.checkBox_2.isChecked():
            out2=None
            in2=None

        if not self.checkBox_3.isChecked():
            clockin=None

        if not self.checkBox_4.isChecked():
            clockout=None

        time_func = lambda y: None if y == None else y.strftime("%H:%M:%S")
        msm = f'Submit request?\n' \
              f'Date of clock in/out forgotten: {clock_dt.strftime("%d/%m/%Y")}\n' \
              f'Clock-in: {time_func(clockin)}\n' \
              f'Clock-out: {time_func(clockout)}\n' \
              f'Out-1: {time_func(out1)}\n' \
              f'In-1: {time_func(in1)}\n' \
              f'Out-2: {time_func(out2)}\n' \
              f'In-2: {time_func(in2)}\n' \
              f'Work time: {self.textEdit_3.toPlainText()}\n' \
              f'OT time: {self.textEdit_4.toPlainText()}\n' \
              f'Remarks: {remarks}\n'
        a = QMessageBox.question(self, 'Confirmation', msm)
        if a == QMessageBox.No:
            return

        self.cursor_apply = DB.cursor()
        SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
        try:
            self.cursor_apply.execute(SQL, (user_id))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor_apply = DB.cursor()
            self.cursor_apply.execute(SQL, (user_id))

        res = self.cursor_apply.fetchall()

        if res[0][0] != 0:
            current_to = res[0][0]
            current_po = 'LEADER'
        else:
            if res[0][1] != 0:
                current_to = res[0][1]
                current_po = 'DM'
            else:
                current_to = 9999
                current_po = 'HR'

        if current_to!=9999:
            sql = """INSERT INTO forget_record (USER_ID, USER_NAME, REQUEST_DTTM, CLOCK_DT, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2, REMARKS, CURRENT_TO, CURRENT_PO) VALUES
                        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
            user_id, user_name, request_dttm, clock_dt, clockin, clockout, out1, in1, out2, in2, remarks, current_to, current_po))
            DB.commit()
        else:
            sql = """INSERT INTO forget_record (USER_ID, USER_NAME, REQUEST_DTTM, CLOCK_DT, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2, REMARKS,LEADER, DM, HR, CURRENT_TO, CURRENT_PO) VALUES
                                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_apply.execute(sql, (
                user_id, user_name, request_dttm, clock_dt, clockin, clockout, out1, in1, out2, in2, remarks, 'OK','OK','OK',
                current_to, current_po))
            DB.commit()

        self.cursor_apply.close()

        if current_to == 9999:
            self.cursor_migrant = DB.cursor()
            serial_timecard = str(user_id) + clockin.strftime('%Y%m%d')

            SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
            self.cursor_migrant.execute(SQL, (serial_timecard))
            data = self.cursor_migrant.fetchall()
            if data != ():
                SQL = """DELETE FROM time_card WHERE SERIAL=%s"""
                self.cursor_migrant.execute(SQL, (serial_timecard))
                DB.commit()

            SQL = """INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
            self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin, clockout, out1, in1, out2, in2))
            DB.commit()

            self.cursor_migrant.close()

        # Mail sending==============================================
        info_lst = query_email(id=current_to)
        if info_lst == -1:
            pass
        else:
            mailsender.send_request_mail(email_add=info_lst[1],
                                         receiver_name=info_lst[0],
                                         sender_name=MainWindow.name,
                                         mode='forget')
        # ===========================================================
        # Mail sending 2==============================================
        if current_to == 9999:
            info_lst = query_email(id=ID)
            if info_lst == -1:
                pass
            else:
                mailsender.send_approved_mail(email_add=info_lst[1],
                                              receiver_name=info_lst[0],
                                              mode='forget')
        # ===========================================================

        QMessageBox.information(self, 'Info',
                                'Clock in/out forgotten report has been sent successfully! Please wait for the approvements.')
        self.show_tableData()

    def withdraw_apply(self):
        index = self.tableWidget.currentRow()
        if index == -1:
            return

        try:
            request_id = self.tableWidget.item(index, 0).text()
            hr=self.tableWidget.item(index, 12).text()
        except:
            QMessageBox.critical(self, 'Empty Selection', 'Empty selection!')
            return

        if hr=='OK':
            QMessageBox.critical(self, 'Notice','You can not cancel the request because it has been approved completely by all departments!')
            return

        a = QMessageBox.question(self, 'Confirmation', f'Are you sure to cancel the selected request?\n'
                                                       f'Request ID: {request_id}')
        if a == QMessageBox.No:
            return

        self.cursor = DB.cursor()
        sql = """DELETE FROM forget_record WHERE SERIAL=%s"""
        try:
            self.cursor.execute(sql, (request_id))
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql, (request_id))
            DB.commit()

        self.cursor.close()
        QMessageBox.information(self, 'Info', f'Request ID: {request_id}, has been canceled successfully!')
        self.show_tableData()

    def uncheck_clock_in(self):
        if not self.checkBox_3.isChecked():
            if not self.checkBox_4.isChecked():
                self.checkBox_4.setChecked(True)

    def uncheck_clock_out(self):
        if not self.checkBox_4.isChecked():
            if not self.checkBox_3.isChecked():
                self.checkBox_3.setChecked(True)

    def singleTime_or_not(self):
        if self.checkBox_3.isChecked()==False or self.checkBox_4.isChecked()==False:
            self.checkBox.setChecked(False)
            self.checkBox.setEnabled(False)

            self.textEdit_3.setText('-')
            self.textEdit_4.setText('-')
        else:
            self.checkBox.setEnabled(True)

        if self.checkBox_3.isChecked() == False:
            self.timeEdit_2.setEnabled(False)
        else:
            self.timeEdit_2.setEnabled(True)

        if self.checkBox_4.isChecked() == False:
            self.timeEdit_3.setEnabled(False)
        else:
            self.timeEdit_3.setEnabled(True)

    def time_stamp_check(self):
        if self.checkBox_3.isChecked() == False or self.checkBox_4.isChecked() == False:
            self.label_31.setText('')
            return

        in_=int(self.timeEdit_2.text().replace(':',''))
        out_1=int(self.timeEdit_4.text().replace(':',''))
        in_1 = int(self.timeEdit_5.text().replace(':', ''))
        out_2 = int(self.timeEdit_6.text().replace(':', ''))
        in_2 = int(self.timeEdit_7.text().replace(':', ''))
        out_=int(self.timeEdit_3.text().replace(':',''))

        clockin = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_2.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        clockout = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_3.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out1 = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_4.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in1 = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_5.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        out2 = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_6.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')
        in2 = datetime.datetime.strptime(
            self.dateEdit_3.date().toString('yyyy/MM/dd/') + self.timeEdit_7.time().toString('hh:mm'), '%Y/%m/%d/%H:%M')

        data_pre_input = [None, None, None, None, None, None, None, None, None]

        data_pre_input[2] = clockin
        data_pre_input[3] = clockout
        data_pre_input[4] = out1
        data_pre_input[5] = in1
        data_pre_input[6] = out2
        data_pre_input[7] = in2

        if self.checkBox.isChecked()==False:
            if not in_<out_:
                self.label_31.setText('Please input the correct time range!')
                self.textEdit_3.setText('-')
                self.textEdit_4.setText('-')
            else:
                self.label_31.setText('')
                data_pre_input[4] = None
                data_pre_input[5] = None
                data_pre_input[6] = None
                data_pre_input[7] = None

                data_set = calculate_without_approved_ot(data_line=data_pre_input)
                work_time = data_set[0]
                over_time = data_set[1]
                self.textEdit_3.setText(str(work_time))
                self.textEdit_4.setText(str(over_time))

        else:
            if self.checkBox_2.isChecked()==False:
                if not in_<out_1<in_1<out_:
                    self.label_31.setText('Please input the correct time range!')
                    self.textEdit_3.setText('-')
                    self.textEdit_4.setText('-')
                else:
                    self.label_31.setText('')
                    data_pre_input[6] = None
                    data_pre_input[7] = None

                    data_set = calculate_without_approved_ot(data_line=data_pre_input)
                    work_time = data_set[0]
                    over_time = data_set[1]
                    self.textEdit_3.setText(str(work_time))
                    self.textEdit_4.setText(str(over_time))

            else:
                if not in_<out_1<in_1<out_2<in_2<out_:
                    self.label_31.setText('Please input the correct time range!')
                    self.textEdit_3.setText('-')
                    self.textEdit_4.setText('-')
                else:
                    self.label_31.setText('')

                    data_set = calculate_without_approved_ot(data_line=data_pre_input)
                    work_time = data_set[0]
                    over_time = data_set[1]
                    self.textEdit_3.setText(str(work_time))
                    self.textEdit_4.setText(str(over_time))

    def if_in_out_1(self):
        if self.checkBox.isChecked():
            self.timeEdit_4.setEnabled(True)
            self.timeEdit_5.setEnabled(True)
            self.checkBox_2.setEnabled(True)
        else:
            self.timeEdit_4.setEnabled(False)
            self.timeEdit_5.setEnabled(False)
            self.checkBox_2.setChecked(False)
            self.checkBox_2.setEnabled(False)

    def if_in_out_2(self):
        if self.checkBox_2.isChecked():
            self.timeEdit_6.setEnabled(True)
            self.timeEdit_7.setEnabled(True)
        else:
            self.timeEdit_6.setEnabled(False)
            self.timeEdit_7.setEnabled(False)

    def quit(self):
        ForgetRecord.close()

    def closeEvent(self, event):
        TimeCard.show()
        TimeCard.calendarWidget.setFocus()

class AdminMain(QMainWindow, Ui_AdminMain):
    def __init__(self):
        super(AdminMain, self).__init__()
        self.setupUi(self)

        self.pushButton_7.clicked.connect(self.quit)
        self.pushButton.clicked.connect(self.to_staff_manage)
        self.pushButton_2.clicked.connect(self.to_team_stru)
        self.pushButton_3.clicked.connect(self.to_login_pass)
        self.pushButton_9.clicked.connect(self.to_calendar_setting)
        self.pushButton_4.clicked.connect(self.to_ot_sheet)

    def to_ot_sheet(self):
        OTSheet.show()
        OTSheet.initializing()
        OTSheet.load_names()
        self.setVisible(False)

    def to_calendar_setting(self):
        CalendarSetting.show()
        CalendarSetting.show_on_table()
        self.setVisible(False)

    def to_login_pass(self):
        LoginPass.show()
        LoginPass.show_on_table()
        self.setVisible(False)

    def to_staff_manage(self):
        StaffManage.show()
        StaffManage.show_on_table()
        self.setVisible(False)

    def to_team_stru(self):
        TeamStructure.show()
        TeamStructure.show_on_table()
        self.setVisible(False)

    def quit(self):
        self.pushButton_7.setAttribute(Qt.WA_UnderMouse, False)
        AdminMain.destroy()
        MainWindow.setVisible(True)

    def closeEvent(self, event):
        self.quit()

class StaffManage(QMainWindow, Ui_StaffManage):
    def __init__(self):
        super(StaffManage, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

        self.pushButton.clicked.connect(self.export_selected_rows)
        self.pushButton_5.clicked.connect(self.import_partial_rows)

        #self.tableWidget.itemSelectionChanged.connect(self.on_selection)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        # print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM akt_staff_"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(11):   #-------Updated on 9/11/2022
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def on_selection(self):
        selected_rows = [index.row() for index in self.tableWidget.selectionModel().selectedRows()]
        #print("Selected rows:", selected_rows)

        data = []
        for row in selected_rows:
            line = []
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                line.append(item.text())
                #print(f"Row {row + 1}, Col {col + 1}: {item.text()}")
            data.append(line)

        #print(data)
        return data

    def export_selected_rows(self):
        data = self.on_selection()
        if not data:
            QMessageBox.warning(self, 'warning', 'Please select at least one row before click this button!')
            return

        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Staff Information(Selected Rows Only)',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'DIVISION', 'EMAIL', 'ANNUAL LEAVE', 'SICK LEAVE', 'PERSONAL LEAVE',
                   'HOMETOWN TRAVEL', 'REMARKS', 'CONTRACT OR NOT']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            wb.close()
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def import_partial_rows(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'ID' \
                or ws.cell(row=1, column=2).value.strip() != 'NAME' \
                or ws.cell(row=1, column=3).value.strip() != 'POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'DIVISION' \
                or ws.cell(row=1, column=5).value.strip() != 'EMAIL' \
                or ws.cell(row=1, column=6).value.strip() != 'ANNUAL LEAVE' \
                or ws.cell(row=1, column=7).value.strip() != 'SICK LEAVE' \
                or ws.cell(row=1, column=8).value.strip() != 'PERSONAL LEAVE' \
                or ws.cell(row=1, column=9).value.strip() != 'HOMETOWN TRAVEL' \
                or ws.cell(row=1, column=10).value.strip() != 'REMARKS' \
                or ws.cell(row=1, column=11).value.strip() != 'CONTRACT OR NOT':  # ------Updated on 9/11/2022
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column + 1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        wb.close()

        cur = DB.cursor()

        for each in data:
            SQL = """UPDATE akt_staff_ SET NAME=%s, POSITION=%s, DIVISION=%s, EMAIL=%s, AN_DAYS=%s, SICK_DAYS=%s, PERSONAL_DAYS=%s, HOME_TOWN=%s, REMARK=%s, CONTRACT_OR_NOT=%s WHERE ID=%s"""  #-------Updated on 9/11/2022
            try:
                cur.execute(SQL, (
                each[1], each[2], each[3], each[4], each[5], each[6], each[7],each[8], each[9], each[10], each[0]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur=DB.cursor()
                cur.execute(SQL, (
                    each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9], each[10], each[0]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Staff management database has been updated successfully!')
        cur.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb=xl.load_workbook(a[0])
        ws=wb.active
        if ws.cell(row=1, column=1).value.strip()!='ID'\
                or ws.cell(row=1, column=2).value.strip()!='NAME'\
                or ws.cell(row=1, column=3).value.strip()!='POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'DIVISION' \
                or ws.cell(row=1, column=5).value.strip() != 'EMAIL' \
                or ws.cell(row=1, column=6).value.strip() != 'ANNUAL LEAVE' \
                or ws.cell(row=1, column=7).value.strip() != 'SICK LEAVE' \
                or ws.cell(row=1, column=8).value.strip() != 'PERSONAL LEAVE' \
                or ws.cell(row=1, column=9).value.strip() != 'HOMETOWN TRAVEL' \
                or ws.cell(row=1, column=10).value.strip() != 'REMARKS'\
                or ws.cell(row=1, column=11).value.strip() != 'CONTRACT OR NOT':     #------Updated on 9/11/2022
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        if ws.max_row < 25:
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data=[]
        for i in range(2, ws.max_row+1):
            line=[]
            for j in range(1, ws.max_column+1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        wb.close()

        cur = DB.cursor()
        SQL="""DELETE FROM akt_staff_ WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO akt_staff_ (ID, NAME, POSITION, DIVISION, EMAIL, AN_DAYS, SICK_DAYS, PERSONAL_DAYS, HOME_TOWN, REMARK, CONTRACT_OR_NOT) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""  #-------Updated on 9/11/2022
            try:
                cur.execute(SQL, (
                each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7],each[8], each[9], each[10]))  #-------Updated on 9/11/2022
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur=DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9], each[10]))  #-------Updated on 9/11/2022
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Staff management database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Staff Information',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(11):  #--------Updated on 9/11/2022
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'DIVISION', 'EMAIL', 'ANNUAL LEAVE', 'SICK LEAVE', 'PERSONAL LEAVE', 'HOMETOWN TRAVEL', 'REMARKS', 'CONTRACT OR NOT']  #--------Updated on 9/11/2022
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            wb.close()
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        StaffManage.destroy()

    def closeEvent(self, event):
        self.quit()

class TeamStructure(QMainWindow, Ui_TeamStructure):
    def __init__(self):
        super(TeamStructure, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        # print(meeting_dt, type(meeting_dt))
        self.cursor = DB.cursor()
        sql = """SELECT * FROM team_stru"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(10):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'ID' \
                or ws.cell(row=1, column=2).value.strip() != 'NAME' \
                or ws.cell(row=1, column=3).value.strip() != 'POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'DIVISION' \
                or ws.cell(row=1, column=5).value.strip() != 'LEADER ID' \
                or ws.cell(row=1, column=6).value.strip() != 'LEADER NAME' \
                or ws.cell(row=1, column=7).value.strip() != 'DM ID' \
                or ws.cell(row=1, column=8).value.strip() != 'DM NAME' \
                or ws.cell(row=1, column=9).value.strip() != 'MD ID'\
                or ws.cell(row=1, column=10).value.strip() != 'MD NAME':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column+1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM team_stru WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO team_stru (ID, NAME, POSITION, DIVISION, LEADER_ID, LEADER_NAME, DM_ID, DM_NAME, MD_ID, MD_NAME) VALUES (%s, %s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4], each[5], each[6], each[7], each[8], each[9]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Team structure database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Team Structure',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(10):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'DIVISION', 'LEADER ID', 'LEADER NAME', 'DM ID', 'DM NAME', 'MD ID', 'MD NAME',]
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        TeamStructure.destroy()

    def closeEvent(self, event):
        self.quit()

class LoginPass(QMainWindow, Ui_LoginPass):
    def __init__(self):
        super(LoginPass, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        self.cursor = DB.cursor()
        sql = """SELECT * FROM login_pass"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(5):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'ID' \
                or ws.cell(row=1, column=2).value.strip() != 'NAME' \
                or ws.cell(row=1, column=3).value.strip() != 'POSITION' \
                or ws.cell(row=1, column=4).value.strip() != 'PASSWORD' \
                or ws.cell(row=1, column=5).value.strip() != 'PRIORITY':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column + 1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM login_pass WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO login_pass (ID, NAME, POSITION, PASSWORD, PRIORITY) VALUES (%s, %s,%s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3], each[4]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Staff account database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Staff Accounts',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(5):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['ID', 'NAME', 'POSITION', 'PASSWORD', 'PRIORITY']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        LoginPass.destroy()

    def closeEvent(self, event):
        self.quit()

class CalendarSetting(QMainWindow, Ui_CalendarSetting):
    def __init__(self):
        super(CalendarSetting, self).__init__()
        self.setupUi(self)

        self.pushButton_2.clicked.connect(self.export_excel)
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton_4.clicked.connect(self.quit)

    def show_on_table(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(20)

        self.cursor = DB.cursor()
        sql = """SELECT * FROM calendar"""
        try:
            self.cursor.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            self.cursor = DB.cursor()
            self.cursor.execute(sql)

        res = self.cursor.fetchall()
        if res == ():
            self.cursor.close()
            return

        self.tableWidget.setRowCount(len(res))
        for i in range(len(res)):
            col = 0
            for j in range(4):
                self.tableWidget.setItem(i, col, QTableWidgetItem(str(res[i][j])))
                self.tableWidget.resizeColumnToContents(col)
                self.tableWidget.resizeRowToContents(i)
                col += 1
        self.cursor.close()

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'Please select the source excel file.',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])
        ws = wb.active
        if ws.cell(row=1, column=1).value.strip() != 'DATE' \
                or ws.cell(row=1, column=2).value.strip() != 'WEEKDAY' \
                or ws.cell(row=1, column=3).value.strip() != 'IF WORK' \
                or ws.cell(row=1, column=4).value.strip() != 'REMARKS':
            QMessageBox.critical(self, 'Wrong File Format!', 'Wrong Format! Please select the right source file!')
            return

        data = []
        for i in range(2, ws.max_row + 1):
            line = []
            for j in range(1, ws.max_column + 1):
                line.append(str(ws.cell(row=i, column=j).value).strip())
            data.append(line)

        cur = DB.cursor()
        SQL = """DELETE FROM calendar WHERE 1=1"""
        try:
            cur.execute(SQL)
            DB.commit()
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(SQL)
            DB.commit()

        for each in data:
            SQL = """INSERT INTO calendar (DATE, WEEKDAY, IF_WORK, REMARKS) VALUES (%s, %s,%s,%s)"""
            try:
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3]))
                DB.commit()
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(SQL, (
                    each[0], each[1], each[2], each[3]))
                DB.commit()

        self.show_on_table()
        QMessageBox.information(self, 'Info', 'Calendar database has been updated successfully!')
        cur.close()

    def export_excel(self):
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the excel file path.',
                                        f'./Calendar',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        data = []
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(4):
                line.append(self.tableWidget.item(i, j).text())
            data.append(line)

        wb = xl.Workbook()
        ws = wb.active
        headers = ['DATE', 'WEEKDAY', 'IF WORK', 'REMARKS']
        ws.append(headers)
        for each_line in data:
            ws.append(each_line)

        try:
            wb.save(filename=a[0])
        except:
            QMessageBox.critical(self, 'Error', 'Please close the excel file with the same name first!')
            return
        wb.close()
        QMessageBox.information(self, 'Info', 'Excel file has been created successfully!')

    def quit(self):
        AdminMain.setVisible(True)
        CalendarSetting.destroy()

    def closeEvent(self, event):
        self.quit()

class OTSheet(QMainWindow, Ui_OT_Sheet):
    def __init__(self):
        super(OTSheet, self).__init__()
        self.setupUi(self)

        self.pushButton.clicked.connect(self.add_one)
        self.pushButton_2.clicked.connect(self.add_all)
        self.pushButton_3.clicked.connect(self.remove_one)
        self.pushButton_4.clicked.connect(self.remove_all)
        self.lineEdit.textChanged.connect(self.search)
        self.pushButton_5.clicked.connect(self.export_ot_sheet)
        self.pushButton_6.clicked.connect(self.quit)


    def initializing(self):
        year2 = time.strftime("%Y", time.localtime(time.time()))
        month2 = time.strftime("%m", time.localtime(time.time()))
        if month2 == '1':
            month1 = '12'
            year1 = str(int(year2) - 1)
        else:
            month1 = str(int(month2) - 1)
            year1 = year2

        start_dt = year1 +'-'+ month1+ '-' + '16'
        end_dt = year2 + '-' + month2 + '-' + '15'
        self.dateEdit.setDate(QDate.fromString(start_dt, 'yyyy-M-d'))
        self.dateEdit_2.setDate(QDate.fromString(end_dt, 'yyyy-M-d'))

    def load_names(self):
        self.listWidget.clear()
        cur = DB.cursor()
        sql = """SELECT ID, NAME FROM team_stru"""
        try:
            cur.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(sql)

        res = cur.fetchall()
        if res == ():
            cur.close()
            return

        for each in res:
            self.listWidget.addItem(str(each[0])+'----'+str(each[1]))

        cur.close()

    def export_ot_sheet(self):
        if (self.dateEdit_2.date().toPyDate()-self.dateEdit.date().toPyDate()).days+1>33:
            QMessageBox.critical(self, 'Error', 'Too long time range! The maximum time range is 33 days, please adjust the Start Date and Last Date again.')
            return

        if self.listWidget_2.count()==0:
            QMessageBox.critical(self, 'Empty staff list!', 'Empty staff list! Please select the staff ID-NAME from the left box first!')
            return

        start_dt=self.dateEdit.text().replace('/','_')
        end_dt=self.dateEdit_2.text().replace('/','_')
        a = QFileDialog.getSaveFileName(self,
                                        'Please select the OT sheet file path.',
                                        f'./OT sheet_{start_dt}-{end_dt}',
                                        'Excel Files (*.xlsx);;All Files (*)')
        if a[0] == '':
            return

        wb = xl.Workbook()
        try:
            wb.save(filename=a[0])
        except PermissionError:
            QMessageBox.critical(self, 'Permission denied!',
                                 'Permission denied! Please close the excel file with the same filename first!')
            wb.close()
            return

        wb.close()

        self.exp_ot_sheet=Exp_OT_Sheet(a=a, dateEdit=self.dateEdit, dateEdit_2=self.dateEdit_2, listWidget_2=self.listWidget_2)
        self.exp_ot_sheet.finish_box.connect(self.finish_msgbox)
        self.exp_ot_sheet.update_label.connect(Monitor.update_text)
        self.exp_ot_sheet.update_progress.connect(Monitor.update_progressbar)
        self.exp_ot_sheet.monitor_close.connect(Monitor.close_monitor)
        self.exp_ot_sheet.monitor_open.connect(self.monitor_show)


        self.exp_ot_sheet.start()
        self.monitor_show()




    def monitor_show(self):
        Monitor.show()
        Monitor.initializing()
        self.hide()

    def finish_msgbox(self, title, text):
        QMessageBox.information(self, title, text)

    def search(self):
        text=self.lineEdit.text().lower()
        for i in range(self.listWidget.count()):
            if text in self.listWidget.item(i).text().lower():
                self.listWidget.setCurrentRow(i)
                return

    def add_one(self):
        try:
            a=self.listWidget.currentItem().text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a name on the left list-box first!')
            return

        b=self.listWidget_2.findItems(a,Qt.MatchExactly)

        if b!=[]:
            self.listWidget_2.setCurrentItem(b[0])
            return
        self.listWidget_2.addItem(a)

    def add_all(self):
        self.listWidget_2.clear()
        count=self.listWidget.count()
        lst=[]
        for i in range(count):
            lst.append(self.listWidget.item(i).text())
        for each in lst:
            self.listWidget_2.addItem(each)

    def remove_one(self):
        try:
            a=self.listWidget_2.currentItem().text()
        except AttributeError:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a name on the right list-box first!')
            return

        self.listWidget_2.takeItem(self.listWidget_2.currentRow())

    def remove_all(self):
        self.listWidget_2.clear()

    def quit(self):
        AdminMain.setVisible(True)
        OTSheet.destroy()

    def closeEvent(self, event):
        self.quit()

class Monitor(QWidget, Ui_Monitor):
    def __init__(self):
        super(Monitor, self).__init__()
        self.setupUi(self)

        self.setWindowFlags(Qt.WindowMaximizeButtonHint | Qt.MSWindowsFixedSizeDialogHint)

    def update_progressbar(self, integer):
        self.progressBar.setValue(integer)

    def update_text(self, text):
        self.label.setText(text)

    def monitor_close_login(self):
        self.destroy()
        loginWindow.show()

    def monitor_close_approve(self):
        self.destroy()
        ApprovePanel.setEnabled(True)

    def close_monitor(self):
        self.destroy()
        OTSheet.show()

    def initializing(self):
        self.label.setText('')
        self.progressBar.setValue(0)

class UPGRADE(QThread):
    monitor_close=pyqtSignal()
    finish_box=pyqtSignal(str, str)
    update_label=pyqtSignal(str)
    update_progress=pyqtSignal(int)
    monitor_open=pyqtSignal()
    def __init__(self):
        super(UPGRADE, self).__init__()
        self.local_file='.\main_online.exe'
        self.remote_file='/main_online.exe'
        self.new_name="main_online.exe"
        self.old_name="main_app.exe"

    def ftpDownload(self):
        self.update_label.emit('Connecting to the server...')
        self.update_progress.emit(1)
        host = r'210.1.31.3'
        # port = 21
        user = 'lu'
        password = 'B*c913ke9'
        LocalFile = self.local_file
        RemoteFile = self.remote_file
        ftp = ftplib.FTP()
        ftp.connect(host)

        self.update_label.emit('Account registering...')
        self.update_progress.emit(2)
        ftp.login(user, password)

        self.update_label.emit('Downloading files...')
        self.update_progress.emit(3)
        judge=self.DownloadFile(LocalFile, RemoteFile, ftp)
        ftp.close()
        if not judge:
            return False

        self.update_label.emit('File downloading is finished. Now restarting AKT HR system...')
        self.update_progress.emit(5)
        return True

    # 下载单个文件
    def DownloadFile(self, LocalFile, RemoteFile, ftp):
        file_handler = open(LocalFile, 'wb')
        #print(file_handler)
        #print('----------', RemoteFile)
        try:
            ftp.retrbinary('RETR ' + RemoteFile, file_handler.write)
        except:
            file_handler.close()
            return False

        file_handler.close()
        return True

    def run(self):
        Monitor.progressBar.setMaximum(5)
        judge=self.ftpDownload()
        self.monitor_close.emit()
        if not judge:
            self.finish_box.emit('Server Error',
                                 'HR-Tool server setting wrong! Please contact the administrator.')

        else:
            WriteUpdateCMD(new_name=self.new_name, old_name=self.old_name)
            sys.exit(app.exec_())


class Exp_OT_Sheet(QThread):
    monitor_close=pyqtSignal()
    finish_box=pyqtSignal(str, str)
    update_label=pyqtSignal(str)
    update_progress=pyqtSignal(int)
    monitor_open=pyqtSignal()
    def __init__(self, a, dateEdit, dateEdit_2, listWidget_2):
        super(Exp_OT_Sheet, self).__init__()
        self.a=a
        self.dateEdit=dateEdit
        self.dateEdit_2=dateEdit_2
        self.listWidget_2=listWidget_2

    def run(self):
        a=self.a
        file_path = a[0]


        #temp=None
        temp = self.dateEdit.text().split('/')
        #strDT1 = None
        strDT1 = temp[2] + temp[1] + temp[0]
        #temp = None
        temp = self.dateEdit_2.text().split('/')
        #strDT2 = None
        strDT2 = temp[2] + temp[1] + temp[0]
        self.update_label.emit('Connecting to Time Card Database...')
        calendar = []
        cur = DB.cursor()
        sql = """SELECT * FROM calendar WHERE DATE>=%s AND DATE<=%s"""
        try:
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))

        res = cur.fetchall()
        if res == ():
            cur.close()
            return
        cur.close()

        for each in res:
            calendar.append(list(each))

        for each in calendar:
            print(each)  # calendar record

        userID_lst = []
        count = self.listWidget_2.count()
        for i in range(count):
            userID_lst.append(self.listWidget_2.item(i).text().split('----')[0].strip())

        print(userID_lst)

        cur = DB.cursor()
        sql = """SELECT ID, NAME, CONTRACT_OR_NOT FROM akt_staff_"""
        try:
            cur.execute(sql)
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(sql)

        data_dict = cur.fetchall()
        cur.close()

        id_name = {}
        id_contract = {}
        for each_line in data_dict:
            id_name[str(each_line[0])] = each_line[1]

        for each_line in data_dict:
            id_contract[str(each_line[0])] = each_line[2]
        print(id_name)
        print(id_contract)


        cur = DB.cursor()
        sql = """SELECT USER_ID, TYPE, START_LEN, END_LEN, START_DT, END_DT FROM leave_request WHERE CURRENT_TO=9999 AND START_DT>=%s AND START_DT<=%s"""
        try:
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))

        res = cur.fetchall()
        cur.close()
        leaverequest_lst = []
        if res == ():
            leaverequest_lst = []
        else:
            for each in res:
                each = list(each)
                if each[3] == None:
                    each[3] = each[2]
                if each[5] == None:
                    each[5] = each[4]
                leaverequest_lst.append(each)

        print(leaverequest_lst)

        #--------------------------UPDATE FOR V1.6BETA(START)
        cur = DB.cursor()
        sql = """SELECT USER_ID, LATE_DT, CLOCKIN_TM FROM apply_late WHERE CURRENT_TO=9999 AND LATE_DT>=%s AND LATE_DT<=%s"""
        try:
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))
        except pymysql.err.OperationalError:
            reconnect_DB(self)
            cur = DB.cursor()
            cur.execute(sql,
                        (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))

        res = cur.fetchall()
        cur.close()
        applylate_lst = []
        if res == ():
            applylate_lst = []
        else:
            for each in res:
                each = list(each)
                applylate_lst.append(each)

        print(applylate_lst)
        # --------------------------UPDATE FOR V1.6BETA(END)

        weekday_dict = {1: 'Mon',
                        2: 'Tue',
                        3: 'Wed',
                        4: 'Thu',
                        5: 'Fri',
                        6: 'Sat',
                        7: 'Sun'}

        wb = xl.Workbook()

        # cur=DB.cursor()
        # sql="""SELECT USER_ID, OT_DT, CURRENT_TO FROM ot_request WHERE OT_DT>%s and OT_DT<%s and CURRENT_TO=9999"""
        # try:
        #    cur.execute(sql, (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))
        # except pymysql.err.OperationalError:
        #    reconnect_DB(self)
        #    cur=DB.cursor()
        #    cur.execute(sql,
        #                (datetime.datetime.strptime(strDT1, '%Y%m%d'), datetime.datetime.strptime(strDT2, '%Y%m%d')))
        # ot_apply_sourceData=cur.fetchall()
        # cur.close()

        # ot_apply_lst=[]
        # if ot_apply_sourceData==():
        # ot_apply_lst=[]
        # else:
        # for each_line in ot_apply_sourceData:
        # ot_apply_lst.append(list(each_line))

        # print(ot_apply_lst)

        Monitor.progressBar.setMaximum(len(userID_lst))
        counter=0
        for each_id in userID_lst:
            counter+=1
            self.update_progress.emit(counter)
            self.update_label.emit(f'Creating OT Sheet for:        Staff ID: {each_id}      Name: {id_name[each_id]}')

            contract_or_not = id_contract[str(each_id)]
            print(each_id, str(each_id) + strDT1, str(each_id) + strDT2)
            cur = DB.cursor()
            sql = """SELECT * FROM time_card WHERE SERIAL>=%s and SERIAL<=%s"""
            try:
                cur.execute(sql, (int(str(each_id) + strDT1), int(str(each_id) + strDT2)))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                cur = DB.cursor()
                cur.execute(sql, (int(str(each_id) + strDT1), int(str(each_id) + strDT2)))

            res = cur.fetchall()
            if res == ():
                cur.close()
                continue
            cur.close()

            time_card = []
            for each in res:
                each = list(each)
                if each[2] != None:
                    each[2] = each[2] + datetime.timedelta(seconds=-each[2].second)
                if each[3] != None:
                    each[3] = each[3] + datetime.timedelta(seconds=-each[3].second)
                if each[4] != None:
                    each[4] = each[4] + datetime.timedelta(seconds=-each[4].second)
                if each[5] != None:
                    each[5] = each[5] + datetime.timedelta(seconds=-each[5].second)
                if each[6] != None:
                    each[6] = each[6] + datetime.timedelta(seconds=-each[6].second)
                if each[7] != None:
                    each[7] = each[7] + datetime.timedelta(seconds=-each[7].second)
                time_card.append(each)

            for each in time_card:
                print(each)  # time card record
                print('--------------------------------------------------------------------')

            set_format(wb, file_path, id_name[each_id])


            ws = wb[id_name[each_id]]
            ws['C2'].value = id_name[each_id]
            ws['C3'].value = str(each_id)
            ws[
                'C4'].value = f"""{datetime.datetime.strptime(strDT1, '%Y%m%d').strftime("%d %b'%y")} to {datetime.datetime.strptime(strDT2, '%Y%m%d').strftime("%d %b'%y")}"""

            i = 7
            for each_line in calendar:
                i += 1
                if each_line[0].day == 1:
                    for j in range(2, 22):
                        ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor='000000')
                    i += 1

                for each in leaverequest_lst:
                    if str(each_id) == str(each[0]) and each_line[0] >= each[4] and each_line[0] <= each[5]:
                        if each_line[0] == each[4]:
                            statement_when = each[2]
                        elif each_line[0] == each[5]:
                            statement_when = each[3]
                        else:
                            statement_when = 'all'

                        leave_type = each[1]
                        if statement_when == 'all':
                            pass
                        elif statement_when == 'morning':
                            leave_type = 'AM ' + leave_type
                        else:
                            leave_type = 'PM ' + leave_type

                        #MODIFIED ON 12/3/2024 --------START
                        if ws[f'V{i}'].value:
                            if ws[f'W{i}'].value:
                                ws[f'X{i}'].value = leave_type
                            else:
                                ws[f'W{i}'].value = leave_type
                        else:
                            ws[f'V{i}'].value = leave_type
                        ws[f'V{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                        ws[f'W{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                        ws[f'X{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                        #break
                        # MODIFIED ON 12/3/2024 --------END

                ws[f'B{i}'].value = each_line[0]
                ws[f'B{i}'].number_format = '[$-409]dd-mmm-yy;@'
                ws[f'C{i}'].value = weekday_dict[each_line[1]]
                if each_line[2] == 'NO':
                    ws[f'C{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")

                # MODIFIED ON 12/3/2024 --------START
                if each_line[3] != 'None':
                    if ws[f'V{i}'].value:
                        if ws[f'W{i}'].value:
                            ws[f'X{i}'].value = each_line[3]
                        else:
                            ws[f'W{i}'].value = each_line[3]
                    else:
                        ws[f'V{i}'].value = each_line[3]
                    ws[f'V{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                    ws[f'W{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                    ws[f'X{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11, color="FF0000")
                # MODIFIED ON 12/3/2024 --------END

            excel_start_row = 8
            for data_line in time_card:
                for i in range(excel_start_row, 42):
                    if data_line[2].date() == ws[f'B{i}'].value:
                        try:
                            ws[f'D{i}'].value = data_line[2].time()
                        except AttributeError:
                            pass
                        try:
                            ws[f'E{i}'].value = data_line[2].time()
                            clockin_sys = data_line[2]
                        except AttributeError:
                            pass
                        if data_line[2] != None:
                            if int(data_line[2].strftime('%H%M')) > 1130 and int(data_line[2].strftime('%H%M')) < 1230:
                                ws[f'E{i}'].value = datetime.time(12, 30)
                                clockin_sys = datetime.datetime(data_line[2].year, data_line[2].month, data_line[2].day,
                                                                12, 30)
                            elif int(data_line[2].strftime('%H%M')) > 1500 and int(
                                    data_line[2].strftime('%H%M')) < 1515:
                                ws[f'E{i}'].value = datetime.time(15, 15)
                                clockin_sys = datetime.datetime(data_line[2].year, data_line[2].month, data_line[2].day,
                                                                15, 15)

                        try:
                            ws[f'F{i}'].value = data_line[4].time()
                        except AttributeError:
                            pass
                        try:
                            ws[f'G{i}'].value = data_line[5].time()
                        except AttributeError:
                            pass
                        try:
                            ws[f'H{i}'].value = data_line[6].time()
                        except AttributeError:
                            pass
                        try:
                            ws[f'I{i}'].value = data_line[7].time()
                        except AttributeError:
                            pass
                        #--------------------------------V1.6BETA UPDATE(START)  ------UPDATED ON 12/3/2024--START
                        if 'PM ' in str(ws[f'V{i}'].value) or 'PM ' in str(ws[f'W{i}'].value) or 'PM ' in str(ws[f'X{i}'].value):
                            if data_line[3] != None:
                                if int(data_line[3].strftime('%H%M')) > 1130:
                                    data_line[3]=datetime.datetime(data_line[3].year, data_line[3].month, data_line[3].day, 11, 30)

                        # --------------------------------V1.6BETA UPDATE(END) ------UPDATED ON 12/3/2024--END
                        try:
                            ws[f'O{i}'].value = data_line[3].time()
                        except AttributeError:
                            pass

                        ws[f'K{i}'].value = datetime.time(11, 30)
                        ws[f'L{i}'].value = datetime.time(12, 30)
                        if data_line[2] == None or data_line[3] == None:
                            ws[f'K{i}'].value = None
                            ws[f'L{i}'].value = None
                        elif int(data_line[2].strftime('%H%M')) >= 1130 or int(data_line[3].strftime('%H%M')) <= 1230:
                            if data_line[2].day==data_line[3].day: #UPDATE FOR V1.6BETA
                                ws[f'K{i}'].value = None
                                ws[f'L{i}'].value = None

                        ws[f'M{i}'].value = datetime.time(15, 0)
                        ws[f'N{i}'].value = datetime.time(15, 15)
                        if data_line[2] == None or data_line[3] == None:
                            ws[f'M{i}'].value = None
                            ws[f'N{i}'].value = None
                        elif int(data_line[2].strftime('%H%M')) >= 1500 or int(data_line[3].strftime('%H%M')) <= 1515:
                            if data_line[2].day==data_line[3].day: #UPDATE FOR V1.6BETA
                                ws[f'M{i}'].value = None
                                ws[f'N{i}'].value = None


                        if data_line[4] == None or data_line[5] == None:
                            out_hour = None
                        else:
                            if data_line[6] == None or data_line[7] == None:
                                out_hour = data_line[5] - data_line[4]
                            else:
                                out_hour = (data_line[5] - data_line[4]) + (data_line[7] - data_line[6])

                        ws[f'J{i}'].value = out_hour

                        if data_line[2] == None or data_line[3] == None:
                            worked_hour = None
                        else:
                            if out_hour==None:   # UPDATE FOR V1.6BETA
                                worked_hour = round((data_line[3] - clockin_sys).seconds / 3600, 2)  # UPDATE FOR V1.6BETA
                            else:   # UPDATE FOR V1.6BETA
                                worked_hour = round((data_line[3] - clockin_sys-out_hour).seconds / 3600, 2) #UPDATE FOR V1.6BETA
                        ws[f'P{i}'].value = worked_hour

                        calculate_result = calculate_worktime(data_line=data_line)

                        excl_break = calculate_result[0]
                        if excl_break == '-':
                            excl_break = 0

                        if excl_break == 0:
                            pass
                        else:
                            ws[f'Q{i}'].value = excl_break

                        if data_line[2] == None or data_line[3] == None:
                            base_hour = None

                        else:
                            base_hour = datetime.time(8, 0)
                        ws[f'R{i}'].value = base_hour

                        ot_hour = calculate_result[1]
                        if ot_hour == '-':
                            ot_hour = 0

                        if float(excl_break) < 8 and float(excl_break) != 0:
                            ot_hour = excl_break - 8
                        elif float(excl_break) == 0:
                            ot_hour = None

                        ws[f'S{i}'].value = ot_hour

                        #--------------------------------UPDATE FOR V1.6BETA(START) --------UPDATED ON 12/3/2024--START
                        if data_line[2].time() > datetime.time(8, 30, 59):
                            if ws[f'V{i}'].value == None or ('AM' not in ws[f'V{i}'].value):
                                #sql = "SELECT USER_ID, LATE_DT, CLOCKIN_TM FROM apply_late WHERE CURRENT_TO=9999 AND LATE_DT>%s AND LATE_DT<%s"
                                if ws[f'V{i}'].value:
                                    ws[f'W{i}'].value = 'Late'
                                else:
                                    ws[f'V{i}'].value = 'Late'
                                ws[f'V{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11,
                                                        color="FF0000")
                                ws[f'W{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11,
                                                        color="FF0000")
                                for each_apply in applylate_lst:
                                    #a=datetime.time(hour=each_apply[2].seconds//3600, minute=(each_apply[2].seconds%3600)//60, second=59)
                                    #print(a)
                                    if str(each_id)==each_apply[0] and each_apply[1]==data_line[2].date() and data_line[2].time()<=datetime.time(hour=each_apply[2].seconds//3600, minute=(each_apply[2].seconds%3600)//60, second=59):
                                        if ws[f'V{i}'].value:
                                            ws[f'W{i}'].value = 'Approved late clockin'
                                        else:
                                            ws[f'V{i}'].value = 'Approved late clockin'
                                        ws[f'V{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11,
                                                                color="FF0000")
                                        ws[f'W{i}'].font = Font(name=u'ＭＳ Ｐゴシック', bold=True, italic=False, size=11,
                                                                color="FF0000")
                                        break

                        # --------------------------------UPDATE FOR V1.6BETA(END)------UPDATED ON 12/3/2024--END

                        if str(contract_or_not).strip().upper() == 'YES' or calculate_result[2] != '0':
                            pass
                        else:
                            if excl_break > 8:
                                t_delta_seconds = (data_line[3] - clockin_sys).seconds - (9.25 * 3600)
                                ws[f'Q{i}'].value = 8
                                ws[f'P{i}'].value = 9.25
                                ws[f'O{i}'].value = (data_line[3] + datetime.timedelta(seconds=-t_delta_seconds)).time()
                                ws[f'S{i}'].value = 0

                        if ws[f'Q{i}'].value:
                            final_hour_exc=ws[f'Q{i}'].value
                            if final_hour_exc >= 10.5:
                                ws[f'T{i}'].value = 1

                            if (data_line[2].time() < datetime.time(8, 31, 0) or (ws[f'V{i}'].value == 'Approved late clockin' or ws[f'W{i}'].value == 'Approved late clockin')) and final_hour_exc >= 8:#UPDATED ON 12/3/2024
                                ws[f'U{i}'].value = 1

                        excel_start_row = i + 1
                        break

            ot_rate = 0
            deduct_rate = 0
            for i in range(8, 42):
                cell_val = ws[f'S{i}'].value
                if cell_val == None:
                    pass
                else:
                    if float(cell_val) > 0:
                        ot_rate += float(cell_val)
                    else:
                        deduct_rate += float(cell_val)

            ws['S43'].value = ot_rate
            ws['S44'].value = deduct_rate

        self.update_label.emit('Saving excel file...')
        wb.save(filename=file_path)
        wb.close()

        self.monitor_close.emit()
        self.finish_box.emit('Info', 'OT Sheet has been created successfully!')


class Forget_All_Accepted(QThread):
    monitor_close = pyqtSignal()
    finish_box = pyqtSignal(str, str)
    update_label = pyqtSignal(str)
    update_progress = pyqtSignal(int)
    monitor_open = pyqtSignal()
    def __init__(self, data):
        super(Forget_All_Accepted, self).__init__()
        self.data = data

    def run(self):
        Monitor.progressBar.setMaximum(len(self.data))
        counter = 0
        for data_line in self.data:
            counter += 1
            self.update_progress.emit(counter)

            request_id = data_line[0]

            self.update_label.emit(
                f'Approving the time record adding request of Request ID: {request_id}, and sending the email...')

            self.cursor_approve = DB.cursor()
            SQL = """SELECT CURRENT_PO, USER_ID FROM forget_record WHERE SERIAL=%s"""
            try:
                self.cursor_approve.execute(SQL, (request_id))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_approve = DB.cursor()
                self.cursor_approve.execute(SQL, (request_id))

            res_po = self.cursor_approve.fetchall()
            current_po = res_po[0][0]
            staff_id = res_po[0][1]

            SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
            self.cursor_approve.execute(SQL, (staff_id))
            res = self.cursor_approve.fetchall()

            if current_po == 'LEADER':
                if res[0][1] != 0:
                    if res[0][0] == res[0][1]:
                        current_to = 9999
                        current_po = 'HR'
                    else:
                        current_to = res[0][1]
                        current_po = 'DM'
                else:
                    current_to = 9999
                    current_po = 'HR'
            else:
                current_to = 9999
                current_po = 'HR'

            if current_po == 'DM':
                SQL = """UPDATE forget_record SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))

            else:
                SQL = """UPDATE forget_record SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

            DB.commit()
            self.cursor_approve.close()

            if current_to == 9999:
                SQL = """SELECT USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2 FROM forget_record WHERE SERIAL=%s"""
                self.cursor_migrant = DB.cursor()
                self.cursor_migrant.execute(SQL, (request_id))
                data = self.cursor_migrant.fetchall()
                user_id = data[0][0]
                clockin = data[0][1]
                clockout = data[0][2]
                out1 = data[0][3]
                in1 = data[0][4]
                out2 = data[0][5]
                in2 = data[0][6]

                try:
                    serial_timecard = str(user_id) + clockin.strftime('%Y%m%d')
                except AttributeError:
                    serial_timecard = str(user_id) + clockout.strftime('%Y%m%d')

                if clockin == None:
                    SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (serial_timecard))
                    data = self.cursor_migrant.fetchall()
                    if data != ():
                        SQL = """UPDATE time_card SET CLOCK_OUT=%s WHERE SERIAL=%s"""
                        self.cursor_migrant.execute(SQL, (clockout, serial_timecard))
                        DB.commit()
                    else:
                        SQL = """INSERT INTO time_card (SERIAL,USER_ID,CLOCK_OUT) VALUES (%s,%s,%s)"""
                        self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockout))
                        DB.commit()

                elif clockout == None:
                    SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (serial_timecard))
                    data = self.cursor_migrant.fetchall()
                    if data != ():
                        SQL = """UPDATE time_card SET CLOCK_IN=%s WHERE SERIAL=%s"""
                        self.cursor_migrant.execute(SQL, (clockin, serial_timecard))
                        DB.commit()
                    else:
                        SQL = """INSERT INTO time_card (SERIAL,USER_ID,CLOCK_IN) VALUES (%s,%s,%s)"""
                        self.cursor_migrant.execute(SQL, (serial_timecard, user_id, clockin))
                        DB.commit()

                else:
                    SQL = """SELECT * FROM time_card WHERE SERIAL=%s"""
                    self.cursor_migrant.execute(SQL, (serial_timecard))
                    data = self.cursor_migrant.fetchall()
                    if data != ():
                        SQL = """DELETE FROM time_card WHERE SERIAL=%s"""
                        self.cursor_migrant.execute(SQL, (serial_timecard))
                        DB.commit()

                    SQL = """INSERT INTO time_card (SERIAL, USER_ID, CLOCK_IN, CLOCK_OUT, OUT_1, IN_1, OUT_2, IN_2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
                    self.cursor_migrant.execute(SQL,
                                                (serial_timecard, user_id, clockin, clockout, out1, in1, out2, in2))
                    DB.commit()

                self.cursor_migrant.close()

            # Mail sending==============================================
            info_lst = query_email(id=current_to)
            info_lst2 = query_email(id=staff_id)
            if info_lst == -1:
                pass
            else:
                mailsender.send_request_mail(email_add=info_lst[1],
                                             receiver_name=info_lst[0],
                                             sender_name=info_lst2[0],
                                             mode='forget')
            # ===========================================================
            # Mail sending 2==============================================
            if current_to == 9999:
                if info_lst2 == -1:
                    pass
                else:
                    mailsender.send_approved_mail(email_add=info_lst2[1],
                                                  receiver_name=info_lst2[0],
                                                  mode='forget')
            # ===========================================================
        #QMessageBox.information(self, 'Info', 'Request has been accepted!')
        ApprovePanel.show_forget_contents()
        # ApprovePanel.show_forget_panel()
        self.monitor_close.emit()
        self.finish_box.emit('Info', 'All of the time record adding requests has been accepted!')


class Late_All_Accepted(QThread):
    monitor_close = pyqtSignal()
    finish_box = pyqtSignal(str, str)
    update_label = pyqtSignal(str)
    update_progress = pyqtSignal(int)
    monitor_open = pyqtSignal()
    def __init__(self, data):
        super(Late_All_Accepted, self).__init__()
        self.data = data

    def run(self):
        Monitor.progressBar.setMaximum(len(self.data))
        counter = 0
        for data_line in self.data:
            counter += 1
            self.update_progress.emit(counter)

            request_id = data_line[0]

            self.update_label.emit(f'Approving the late clock-in request of Request ID: {request_id}, and sending the email...')

            self.cursor_approve = DB.cursor()
            SQL = """SELECT CURRENT_PO, USER_ID FROM apply_late WHERE SERIAL=%s"""
            try:
                self.cursor_approve.execute(SQL, (request_id))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_approve = DB.cursor()
                self.cursor_approve.execute(SQL, (request_id))

            res_po = self.cursor_approve.fetchall()
            current_po = res_po[0][0]
            staff_id = res_po[0][1]

            SQL = """SELECT LEADER_ID, DM_ID FROM team_stru WHERE ID=%s"""
            self.cursor_approve.execute(SQL, (staff_id))
            res = self.cursor_approve.fetchall()

            if current_po == 'LEADER':
                if res[0][1] != 0:  # 如果存在DM
                    if res[0][0] == res[0][1]:  # 如果LEADER跟DM是同一人
                        # current_to = 9999
                        # current_po = 'HR'
                        current_to = 8888
                        current_po = 'HR'

                    else:  # 如果LEADER跟DM不是同一人
                        current_to = res[0][1]
                        current_po = 'DM'

                else:  # 如果不存在DM
                    # current_to = 9999
                    # current_po = 'HR'
                    current_to = 8888
                    current_po = 'HR'

            elif current_po == 'DM':
                current_to = 8888
                current_po = 'HR'

            else:
                current_to = 9999
                current_po = 'PA'

            if current_po == 'DM':
                SQL = """UPDATE apply_late SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, ('OK', current_to, current_po, request_id))
            elif current_po == 'HR':
                SQL = """UPDATE apply_late SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, ('OK', 'OK', current_to, current_po, request_id))
            else:
                SQL = """UPDATE apply_late SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, ('OK', 'OK', 'OK', current_to, current_po, request_id))

            DB.commit()
            self.cursor_approve.close()

            # Mail sending==============================================
            info_lst = query_email(id=current_to)
            info_lst2 = query_email(id=staff_id)
            if info_lst == -1:
                pass
            else:
                mailsender.send_request_mail(email_add=info_lst[1],
                                             receiver_name=info_lst[0],
                                             sender_name=info_lst2[0],
                                             mode='late')
            # ===========================================================
            # Mail sending 2==============================================
            if current_to == 9999:
                if info_lst2 == -1:
                    pass
                else:
                    mailsender.send_approved_mail(email_add=info_lst2[1],
                                                  receiver_name=info_lst2[0],
                                                  mode='late')
            # ===========================================================
        #QMessageBox.information(self, 'Info', 'Request has been accepted!')
        ApprovePanel.show_late_contents()
        #ApprovePanel.show_late_panel()
        self.monitor_close.emit()
        self.finish_box.emit('Info', 'All of the late clock-in requests has been accepted!')

class Ot_All_Accepted(QThread):
    monitor_close = pyqtSignal()
    finish_box = pyqtSignal(str, str)
    update_label = pyqtSignal(str)
    update_progress = pyqtSignal(int)
    monitor_open = pyqtSignal()
    def __init__(self, data):
        super(Ot_All_Accepted, self).__init__()
        self.data = data

    def run(self):
        Monitor.progressBar.setMaximum(len(self.data))
        counter = 0
        for data_line in self.data:
            counter += 1
            self.update_progress.emit(counter)

            request_id = data_line[0]

            self.update_label.emit(f'Approving the OT request of Request ID: {request_id}, and sending the email...')

            self.cursor_approve = DB.cursor()
            SQL = """SELECT CURRENT_PO, USER_ID FROM ot_request WHERE SERIAL=%s"""
            try:
                self.cursor_approve.execute(SQL, (request_id))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_approve = DB.cursor()
                self.cursor_approve.execute(SQL, (request_id))

            res_po = self.cursor_approve.fetchall()
            current_po = res_po[0][0]
            staff_id = res_po[0][1]

            SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
            self.cursor_approve.execute(SQL, (staff_id))
            res = self.cursor_approve.fetchall()

            if current_po == 'LEADER':
                if res[0][1] != 0:  # 如果存在DM
                    if res[0][0] == res[0][1]:  # 如果LEADER和DM是同一人
                        if res[0][1] == res[0][2]:  # 如果DM和MD是同一人
                            # current_to = 9999
                            # current_po = 'HR'
                            current_to = 8888
                            current_po = 'HR'
                        else:  # 如果DM和MD不是同一人
                            # current_to = res[0][2]
                            # current_po = 'MD'
                            current_to = 8888
                            current_po = 'HR'
                    else:  # 如果LEADER和DM不是同一人
                        current_to = res[0][1]
                        current_po = 'DM'
                else:  # 如果不存在DM
                    # current_to = res[0][2]
                    # current_po = 'MD'
                    current_to = 8888
                    current_po = 'HR'
            elif current_po == 'DM':
                if res[0][2] == res[0][1]:  # 如果DM和MD是同一人
                    # current_to = 9999
                    # current_po = 'HR'
                    current_to = 8888
                    current_po = 'HR'
                else:  # 如果DM和MD不是同一人
                    # current_to = res[0][2]
                    # current_po = 'MD'
                    current_to = 8888
                    current_po = 'HR'

            elif current_po == 'HR':
                if res[0][2] == res[0][1]:  # 如果DM和MD是同一人
                    current_to = 9999
                    current_po = 'PA'
                else:  # 如果DM和MD不是同一人
                    #current_to = res[0][2]
                    #current_po = 'MD'
                    current_to = 9999
                    current_po = 'PA'

            else:
                current_to = 9999
                current_po = 'PA'

            if current_po == 'DM':
                SQL = """UPDATE ot_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))
            elif current_po == 'HR':
                SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))
            elif current_po == 'MD':
                SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, 1, current_to, current_po, request_id))
            else:
                SQL = """UPDATE ot_request SET LEADER=%s, DM=%s, HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

            DB.commit()
            self.cursor_approve.close()

            # Mail sending==============================================
            info_lst = query_email(id=current_to)
            info_lst2 = query_email(id=staff_id)
            if info_lst == -1:
                pass
            else:
                mailsender.send_request_mail(email_add=info_lst[1],
                                             receiver_name=info_lst[0],
                                             sender_name=info_lst2[0],
                                             mode='ot')
            # ===========================================================
            # Mail sending 2==============================================
            if current_to == 9999:
                if info_lst2 == -1:
                    pass
                else:
                    mailsender.send_approved_mail(email_add=info_lst2[1],
                                                  receiver_name=info_lst2[0],
                                                  mode='ot')
            # ===========================================================
        #QMessageBox.information(self, 'Info', 'Request has been accepted!')
        ApprovePanel.show_ot_contents()
        # ApprovePanel.show_ot_panel()
        self.monitor_close.emit()
        self.finish_box.emit('Info', 'All of the OT requests has been accepted!')

class Leave_All_Accepted(QThread):
    monitor_close=pyqtSignal()
    finish_box=pyqtSignal(str, str)
    update_label=pyqtSignal(str)
    update_progress=pyqtSignal(int)
    monitor_open=pyqtSignal()
    def __init__(self, data):
        super(Leave_All_Accepted, self).__init__()
        self.data = data

    def run(self):
        Monitor.progressBar.setMaximum(len(self.data))
        counter = 0
        for data_line in self.data:
            counter += 1
            self.update_progress.emit(counter)
            request_id = data_line[0]
            staff_id_ = data_line[2]
            type_ = data_line[3]
            during_ = data_line[9]

            self.update_label.emit(f'Approving the leave request of Request ID: {request_id}, and sending the email...')

            self.cursor_approve = DB.cursor()
            SQL = """SELECT CURRENT_PO, USER_ID FROM leave_request WHERE SERIAL=%s"""
            try:
                self.cursor_approve.execute(SQL, (request_id))
            except pymysql.err.OperationalError:
                reconnect_DB(self)
                self.cursor_approve = DB.cursor()
                self.cursor_approve.execute(SQL, (request_id))

            res_po = self.cursor_approve.fetchall()
            current_po = res_po[0][0]
            staff_id = res_po[0][1]

            SQL = """SELECT LEADER_ID, DM_ID, MD_ID FROM team_stru WHERE ID=%s"""
            self.cursor_approve.execute(SQL, (staff_id))
            res = self.cursor_approve.fetchall()

            if current_po == 'LEADER':
                if res[0][1] != 0:  # 如果存在DM
                    if res[0][0] == res[0][1]:  # 如果LEADER和DM是同一个人的话
                        if res[0][1] == res[0][2]:  # 如果DM和MD又是同一个人的话
                            # current_to = 9999
                            # current_po = 'HR'
                            current_to = 8888
                            current_po = 'HR'
                        else:  # 如果LEADER和DM是同一个人，但DM和MD不是同一个人
                            # current_to = res[0][2]
                            # current_po = 'MD'
                            current_to = 8888
                            current_po = 'HR'

                    else:  # 如果LEADER和DM不是同一个人
                        current_to = res[0][1]
                        current_po = 'DM'
                else:  # 如果不存在DM
                    # current_to = res[0][2]
                    # current_po = 'MD'
                    current_to = 8888
                    current_po = 'HR'

            elif current_po == 'DM':
                if res[0][2] == res[0][1]:  # 如果DM和MD是同一人
                    # current_to = 9999
                    # current_po = 'HR'
                    current_to = 8888
                    current_po = 'HR'
                else:  # 如果DM和MD不是同一人
                    # current_to = res[0][2]
                    # current_po = 'MD'
                    current_to = 8888
                    current_po = 'HR'

            elif current_po == 'HR':
                if res[0][2] == res[0][1]:  # 如果DM和MD是同一人
                    current_to = 9999
                    current_po = 'PA'
                else:  # 如果DM和MD不是同一人
                    #current_to = res[0][2]
                    #current_po = 'MD'
                    current_to = 9999
                    current_po = 'PA'

            else:
                current_to = 9999
                current_po = 'PA'

            if current_po == 'DM':
                SQL = """UPDATE leave_request SET LEADER=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, current_to, current_po, request_id))

            elif current_po == 'HR':
                SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, current_to, current_po, request_id))

            elif current_po == 'MD':
                SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, HR=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, 1, current_to, current_po, request_id))
            else:
                SQL = """UPDATE leave_request SET LEADER=%s, DM=%s, HR=%s, MD=%s, CURRENT_TO=%s, CURRENT_PO=%s WHERE SERIAL=%s"""
                self.cursor_approve.execute(SQL, (1, 1, 1, 1, current_to, current_po, request_id))

            DB.commit()

            self.cursor_approve.close()

            # Mail sending==============================================
            info_lst = query_email(id=current_to)
            info_lst2 = query_email(id=staff_id)
            if info_lst == -1:
                pass
            else:
                mailsender.send_request_mail(email_add=info_lst[1],
                                             receiver_name=info_lst[0],
                                             sender_name=info_lst2[0],
                                             mode='leave')
            # ===========================================================
            # Mail sending 2==============================================
            if current_to == 9999:

                if type_ == 'Annual leave':
                    self.cursor_calc = DB.cursor()
                    SQL = """SELECT AN_DAYS FROM akt_staff_ WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (staff_id_))
                    res = self.cursor_calc.fetchall()
                    if res == ():
                        pass
                    else:
                        an_days = res[0][0]
                        an_days -= float(during_)
                        SQL = """UPDATE akt_staff_ SET AN_DAYS=%s WHERE ID=%s"""
                        self.cursor_calc.execute(SQL, (an_days, staff_id_))
                        DB.commit()
                    self.cursor_calc.close()
                elif type_ == 'Sick leave':
                    self.cursor_calc = DB.cursor()
                    SQL = """SELECT SICK_DAYS FROM akt_staff_ WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (staff_id_))
                    res = self.cursor_calc.fetchall()
                    if res == ():
                        pass
                    else:
                        sick_days = res[0][0]
                        sick_days -= float(during_)
                        SQL = """UPDATE akt_staff_ SET SICK_DAYS=%s WHERE ID=%s"""
                        self.cursor_calc.execute(SQL, (sick_days, staff_id_))
                        DB.commit()
                    self.cursor_calc.close()
                elif type_ == 'Hometown':
                    self.cursor_calc = DB.cursor()
                    SQL = """SELECT HOME_TOWN FROM akt_staff_ WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (staff_id_))
                    res = self.cursor_calc.fetchall()
                    if res == ():
                        pass
                    else:
                        home_days = res[0][0]
                        home_days -= 1
                        SQL = """UPDATE akt_staff_ SET HOME_TOWN=%s WHERE ID=%s"""
                        self.cursor_calc.execute(SQL, (home_days, staff_id_))
                        DB.commit()
                    self.cursor_calc.close()
                elif type_ == 'Personal leave':
                    self.cursor_calc = DB.cursor()
                    SQL = """SELECT PERSONAL_DAYS FROM akt_staff_ WHERE ID=%s"""
                    self.cursor_calc.execute(SQL, (staff_id_))
                    res = self.cursor_calc.fetchall()
                    if res == ():
                        pass
                    else:
                        personal_days = res[0][0]
                        personal_days -= float(during_)
                        SQL = """UPDATE akt_staff_ SET PERSONAL_DAYS=%s WHERE ID=%s"""
                        self.cursor_calc.execute(SQL, (personal_days, staff_id_))
                        DB.commit()
                    self.cursor_calc.close()

                if info_lst2 == -1:
                    pass
                else:
                    mailsender.send_approved_mail(email_add=info_lst2[1],
                                                  receiver_name=info_lst2[0],
                                                  mode='leave')
            # ===========================================================

        ApprovePanel.show_leave_contents()
        # ApprovePanel.show_leave_panel()
        #QMessageBox.information(self, 'Info', 'All of the leave requests has been accepted!')
        self.monitor_close.emit()
        self.finish_box.emit('Info', 'All of the leave requests has been accepted!')


def calculate_worktime(data_line):
    day_lag = data_line[8]

    clock_out = data_line[3]
    if clock_out == None:
        clock_out = '-'

    clock_in = data_line[2]
    try:
        out1 = data_line[4]
    except:
        out1 = '-'
    if out1 == None:
        out1 = '-'
    try:
        in1 = data_line[5]
    except:
        in1 = '-'
    if in1 == None:
        in1 = '-'
    try:
        out2 = data_line[6]
    except:
        out2 = '-'
    if out2 == None:
        out2 = '-'
    try:
        in2 = data_line[7]
    except:
        in2 = '-'
    if in2 == None:
        in2 = '-'

    if clock_out == '-':
        work_time = '-'
        over_time = '-'
    # ---------------------计算实际作业时间
    else:
        if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
            # 1
            if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                work_time = (clock_out - clock_in).seconds
            # 2
            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 3
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
                work_time -= 3600
            # 4
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
                work_time -= 3600
            # 5
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 4500

        elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1230'):
            # 6
            if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = 0
            # 7
            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
            # 8
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230',
                    '%Y%m%d%H%M')).seconds
            # 9
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1500'):
            # 10
            if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
            # 11
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 12
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1515'):
            # 13
            if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = 0
            # 14
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
        else:
            work_time = (clock_out - clock_in).seconds

        # ------------------计算离岗累计时间
        space_time1 = 0
        space_time2 = 0
        if out1 != '-' and in1 != '-':
            # 1
            if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time1 = (in1 - out1).seconds
            # 2
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 3
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds - 3600
            # 4
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds - 3600
            # 5
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 3600 - 900
            # 6
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = 0
            # 7
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds
            # 11
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 12
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 900
            # 13
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = 0
            # 14
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time1 = (in1 - out1).seconds

        # -----------------out2 in2
        if out2 != '-' and in2 != '-':
            # 1
            if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time2 = (in2 - out2).seconds
            # 2
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 3
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds - 3600
            # 4
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds - 3600
            # 5
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 3600 - 900
            # 6
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = 0
            # 7
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds
            # 11
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 12
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 900
            # 13
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = 0
            # 14
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time2 = (in2 - out2).seconds

        work_time -= space_time1
        work_time -= space_time2
        # --------------------------如过午夜加上一天
        #if day_lag == 1:
           # work_time += 86400   删除逻辑错误代码

        over_time = work_time - 28800
        if over_time < 0:
            over_time = 0

    if work_time != '-':
        work_time = round(int(work_time) / 3600, 2)
    if over_time != '-':

        over_time = round(int(over_time) / 3600, 2)

    qdt = clock_in
    dt_for_query = datetime.datetime.strptime(qdt.strftime('%d/%m/%Y'), '%d/%m/%Y')
    sql = """SELECT DURING FROM ot_request WHERE USER_ID=%s AND OT_DT=%s AND CURRENT_TO=%s"""
    cursor_queryhours = DB.cursor()
    try:
        cursor_queryhours.execute(sql, (data_line[1], dt_for_query, '9999'))
    except:
        reconnect_DB(MainWindow)
        cursor_queryhours = DB.cursor()
        cursor_queryhours.execute(sql, (data_line[1], dt_for_query), '9999')
    res = cursor_queryhours.fetchall()
    cursor_queryhours.close()
    if res == ():
        approved_ot='0'
    else:
        approved_ot = str(res[0][0])

    return [work_time, over_time, approved_ot]

def calculate_without_approved_ot(data_line):
    day_lag = data_line[8]

    clock_out = data_line[3]
    if clock_out == None:
        clock_out = '-'

    clock_in = data_line[2]
    try:
        out1 = data_line[4]
    except:
        out1 = '-'
    if out1 == None:
        out1 = '-'
    try:
        in1 = data_line[5]
    except:
        in1 = '-'
    if in1 == None:
        in1 = '-'
    try:
        out2 = data_line[6]
    except:
        out2 = '-'
    if out2 == None:
        out2 = '-'
    try:
        in2 = data_line[7]
    except:
        in2 = '-'
    if in2 == None:
        in2 = '-'

    if clock_out == '-':
        work_time = '-'
        over_time = '-'
    # ---------------------计算实际作业时间
    else:
        if int(clock_in.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
            # 1
            if int(clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                work_time = (clock_out - clock_in).seconds
            # 2
            elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 3
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
                work_time -= 3600
            # 4
            elif int(clock_out.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    clock_out.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
                work_time -= 3600
            # 5
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 4500

        elif int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1230'):
            # 6
            if int(clock_in.strftime('%Y%m%d') + '1130') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1230'):
                work_time = 0
            # 7
            elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
            # 8
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230',
                    '%Y%m%d%H%M')).seconds
            # 9
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                    '%Y%m%d%H%M')).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1500'):
            # 10
            if int(clock_in.strftime('%Y%m%d') + '1230') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1500'):
                work_time = (clock_out - clock_in).seconds
            # 11
            elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                        '%Y%m%d%H%M') - clock_in).seconds
            # 12
            else:
                work_time = (clock_out - clock_in).seconds
                work_time -= 900

        elif int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_in.strftime('%Y%m%d%H%M')) <= int(
                clock_in.strftime('%Y%m%d') + '1515'):
            # 13
            if int(clock_in.strftime('%Y%m%d') + '1500') < int(clock_out.strftime('%Y%m%d%H%M')) <= int(
                    clock_in.strftime('%Y%m%d') + '1515'):
                work_time = 0
            # 14
            else:
                work_time = (clock_out - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                    '%Y%m%d%H%M')).seconds
        else:
            work_time = (clock_out - clock_in).seconds

        # ------------------计算离岗累计时间
        space_time1 = 0
        space_time2 = 0
        if out1 != '-' and in1 != '-':
            # 1
            if int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time1 = (in1 - out1).seconds
            # 2
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 3
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds - 3600
            # 4
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds - 3600
            # 5
            elif int(out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 3600 - 900
            # 6
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time1 = 0
            # 7
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time1 = (in1 - out1).seconds
            # 11
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out1).seconds
            # 12
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - out1).seconds - 900
            # 13
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time1 = 0
            # 14
            elif (int(out1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out1.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in1.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time1 = (in1 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time1 = (in1 - out1).seconds

        # -----------------out2 in2
        if out2 != '-' and in2 != '-':
            # 1
            if int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130'):
                space_time2 = (in2 - out2).seconds
            # 2
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1130',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 3
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds - 3600
            # 4
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds - 3600
            # 5
            elif int(out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 3600 - 900
            # 6
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')):
                space_time2 = 0
            # 7
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds
            # 8
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - datetime.datetime.strptime(
                    clock_in.strftime('%Y%m%d') + '1230', '%Y%m%d%H%M')).seconds
            # 9
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1130') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1230')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1230',
                                                                '%Y%m%d%H%M')).seconds - 900
            # 10
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')):
                space_time2 = (in2 - out2).seconds
            # 11
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = (datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1500',
                                                          '%Y%m%d%H%M') - out2).seconds
            # 12
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1230') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1500')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - out2).seconds - 900
            # 13
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and (
                    int(in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                in2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')):
                space_time2 = 0
            # 14
            elif (int(out2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1500') and int(
                    out2.strftime('%Y%m%d%H%M')) <= int(clock_in.strftime('%Y%m%d') + '1515')) and int(
                in2.strftime('%Y%m%d%H%M')) > int(clock_in.strftime('%Y%m%d') + '1515'):
                space_time2 = (in2 - datetime.datetime.strptime(clock_in.strftime('%Y%m%d') + '1515',
                                                                '%Y%m%d%H%M')).seconds
            # 15
            else:
                space_time2 = (in2 - out2).seconds

        work_time -= space_time1
        work_time -= space_time2
        # --------------------------如过午夜加上一天
        #if day_lag == 1:
           # work_time += 86400   删除逻辑错误代码

        over_time = work_time - 28800
        if over_time < 0:
            over_time = 0

    if work_time != '-':
        work_time = round(int(work_time) / 3600, 2)
    if over_time != '-':

        over_time = round(int(over_time) / 3600, 2)

    return [work_time, over_time]

def query_email(id):
    #==========================Updated on 12/9/2023
    if id == 8888:
        sql = """SELECT * FROM hr_receiver WHERE hr_code=%s"""
        cursor_query=DB.cursor()
        try:
            cursor_query.execute(sql, (id))
        except pymysql.err.OperationalError:
            reconnect_DB(MainWindow)
            cursor_query = DB.cursor()
            cursor_query.execute(sql, (id))
        res = cursor_query.fetchall()
        cursor_query.close()
        if res == ():
            return -1
        id = res[0][1]
    #===========================Updated on 12/9/2023 (End)

    sql = """SELECT NAME, EMAIL FROM akt_staff_ WHERE ID=%s"""
    cursor_query=DB.cursor()
    try:
        cursor_query.execute(sql, (id))
    except pymysql.err.OperationalError:
        reconnect_DB(MainWindow)
        cursor_query = DB.cursor()
        cursor_query.execute(sql, (id))
    res=cursor_query.fetchall()
    cursor_query.close()
    if res==():
        return -1
    name=res[0][0]
    email=res[0][1]
    return [name, email]

def reconnect_DB(form):
    global DB
    try:
        DB.close()
    except:
        pass
    try:
        DB = pymysql.connect(host='210.1.31.3',
                             user='hr',
                             port=3306,
                             passwd='gwP6xTsA',
                             db='akaganeHR')
    except pymysql.err.OperationalError:
        QMessageBox.critical(form, 'Network Error', 'Can not connect to the server, please check your network!')

def WriteUpdateCMD(new_name, old_name):
    b=open('upgrade.bat', 'w')
    temp_list="@echo off\n"
    temp_list+="if not exist "+new_name+" exit \n"
    temp_list+="echo 'Upgrading now...'\n"
    temp_list+="timeout /t 10 /nobreak\n"
    temp_list+="del "+old_name+"\n"
    temp_list+="copy "+new_name+" "+old_name+"\n"
    temp_list+="del "+new_name+"\n"
    temp_list+="echo 'Upgrade completed! Restarting...'\n"
    temp_list+="timeout /t 3 /nobreak\n"
    temp_list+="start "+old_name+"\n"
    temp_list+="exit"

    b.write(temp_list)
    b.close()
    #os.system("start upgrade.bat")
    subprocess.Popen('upgrade.bat')


if __name__ == '__main__':
    DB = None
    ID = -1
    CURRENT_VER=3.1
    HR_MODE = 0

    app = QApplication(sys.argv)

    loginWindow = loginWindow()
    loginWindow.show()

    MainWindow = MainWindow()
    PassWindow = PassWindow()
    TimeCard = TimeCard()
    AskForLeave = AskForLeave()
    OTApplication = OTApplication()
    BookMeetingRoom = BookMeetingRoom()
    ApprovePanel = ApprovePanel()
    ApplyLateClockIn=ApplyLateClockIn()
    ForgetRecord=ForgetRecord()

    mailsender=MailSender()

    AdminMain = AdminMain()
    StaffManage = StaffManage()
    TeamStructure = TeamStructure()
    LoginPass = LoginPass()
    CalendarSetting = CalendarSetting()
    OTSheet = OTSheet()
    Monitor = Monitor()
    Tipwindow = Tipwindow()

    sys.exit(app.exec_())
